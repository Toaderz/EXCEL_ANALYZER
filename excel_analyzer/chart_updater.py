"""
chart_updater.py
================
Funciones para detectar, clasificar y actualizar automáticamente los rangos
de gráficas en archivos Excel cuando se agregan nuevos períodos temporales.

FUNCIONES PRINCIPALES:
  diagnosticar_graficas(archivo)
    → Escanea todas las hojas y devuelve un inventario completo de gráficas
      con su clasificación (ventana temporal, último valor, estática, rota).

  actualizar_graficas(archivo, archivo_datos=None, archivo_salida=None, ...)
    → Actualiza automáticamente los rangos de todas las gráficas:
        - Ventana temporal: desliza el rango para incluir el nuevo período.
        - Último valor: mueve la referencia si apunta a datos directos.
        - Rotas (#REF!): las borra opcionalmente.
        - Títulos dinámicos: actualiza nombres de meses en títulos.

FILOSOFÍA:
  Igual que formula_navigator, todo funciona por mapeo semántico.
  Las gráficas se clasifican automáticamente por la forma de sus
  referencias (horizontal = ventana temporal, vertical = último valor).
  Las celdas se inspeccionan para saber si contienen fórmulas o datos,
  y solo se mueven los rangos que realmente lo necesitan.
"""

import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart.series import DataPoint

# ═══════════════════════════════════════════════════════════════════════════
# 1. Modelo de datos
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class ChartInfo:
    """Información completa de una gráfica detectada."""
    hoja: str
    indice: int
    tipo_chart: str              # BarChart, PieChart, LineChart, etc.
    titulo: str | None
    num_series: int
    clasificacion: str           # "ventana_temporal", "ultimo_valor", "estatica", "rota"
    refs_valores: list[str]      # referencias de valores de cada serie
    refs_categorias: list[str]   # referencias de categorías
    necesita_actualizacion: bool
    detalle: str = ""            # descripción legible del estado


@dataclass
class ChartUpdateResult:
    """Resultado de la actualización de gráficas."""
    archivo_salida: str
    graficas_analizadas: int
    graficas_actualizadas: int
    graficas_borradas: int
    titulos_actualizados: int
    detalle: list[str] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════════════════
# 2. Utilidades internas
# ═══════════════════════════════════════════════════════════════════════════

_REF_RANGE = re.compile(
    r"(?:'([^']+)'|([A-Za-zÁÉÍÓÚáéíóúñÑ0-9_ ]+))?"
    r"!?"
    r"\$?([A-Z]{1,3})\$?(\d+)"
    r"(?::\$?([A-Z]{1,3})\$?(\d+))?"
)

_MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

_MESES_EN = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}

# Inversión: nombre → número
_MES_A_NUM_ES = {v.lower(): k for k, v in _MESES_ES.items()}
_MES_A_NUM_EN = {v.lower(): k for k, v in _MESES_EN.items()}


def _extraer_titulo(chart) -> str | None:
    """Extrae el texto del título de una gráfica."""
    if not chart.title:
        return None
    tx = getattr(chart.title, 'tx', None)
    if not tx:
        return None
    rich = getattr(tx, 'rich', None)
    if not rich:
        return None
    textos = []
    for p in rich.paragraphs:
        for r in p.r:
            if r.t and r.t != 'None':
                textos.append(r.t)
    return ' '.join(textos).strip() if textos else None


def _clasificar_serie(ref: str) -> str:
    """
    Clasifica una referencia de serie:
      "horizontal" = misma fila, varía columna → ventana temporal
      "vertical"   = misma columna, varía fila → snapshot / último valor
      "punto"      = celda única
      "rota"       = contiene #REF!
    """
    if '#REF!' in ref:
        return "rota"

    parte = ref.split('!')[-1]
    if ':' not in parte:
        return "punto"

    start, end = parte.split(':')
    col_start = re.sub(r'[\$\d]', '', start)
    col_end = re.sub(r'[\$\d]', '', end)
    row_start = re.sub(r'[\$A-Z]', '', start)
    row_end = re.sub(r'[\$A-Z]', '', end)

    if col_start == col_end and row_start != row_end:
        return "vertical"
    elif row_start == row_end and col_start != col_end:
        return "horizontal"
    else:
        return "bloque"


def _parsear_ref(ref: str) -> dict:
    """Parsea una referencia como 'Hoja'!$M$30:$Y$30 en componentes."""
    resultado = {"raw": ref, "hoja": None, "col_start": None, "row_start": None,
                 "col_end": None, "row_end": None}

    if '#REF!' in ref:
        return resultado

    # Extraer hoja
    if '!' in ref:
        hoja_part, rango_part = ref.rsplit('!', 1)
        resultado["hoja"] = hoja_part.replace("'", "").strip()
    else:
        rango_part = ref

    parts = rango_part.replace('$', '')
    if ':' in parts:
        start, end = parts.split(':')
    else:
        start = end = parts

    col_s = re.sub(r'\d', '', start)
    row_s = re.sub(r'[A-Z]', '', start)
    col_e = re.sub(r'\d', '', end)
    row_e = re.sub(r'[A-Z]', '', end)

    resultado["col_start"] = col_s
    resultado["row_start"] = int(row_s) if row_s else None
    resultado["col_end"] = col_e
    resultado["row_end"] = int(row_e) if row_e else None

    return resultado


def _reconstruir_ref(hoja: str | None, col_start: str, row_start: int,
                     col_end: str | None, row_end: int | None) -> str:
    """Reconstruye una referencia Excel con formato $COL$ROW."""
    if col_end and row_end:
        rango = f"${col_start}${row_start}:${col_end}${row_end}"
    else:
        rango = f"${col_start}${row_start}"

    if hoja:
        if ' ' in hoja or any(c in hoja for c in "'-+&()"):
            return f"'{hoja}'!{rango}"
        return f"{hoja}!{rango}"
    return rango


def _buscar_fila_fechas(ws, rango_filas=range(1, 50)) -> tuple[int | None, dict]:
    """
    Busca la fila que contiene más fechas (datetime) en una hoja.
    Retorna (fila, {col_num: datetime}).
    """
    mejor_fila = None
    max_fechas = 0
    mejor_map = {}

    for r in rango_filas:
        fechas = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if hasattr(v, 'strftime'):
                fechas[c] = v
        if len(fechas) > max_fechas:
            max_fechas = len(fechas)
            mejor_fila = r
            mejor_map = fechas

    return mejor_fila, mejor_map


def _celda_es_formula(wb, hoja: str, col: str, row: int) -> bool:
    """Verifica si una celda contiene una fórmula."""
    if hoja not in wb.sheetnames:
        return False
    ws = wb[hoja]
    col_num = column_index_from_string(col)
    v = ws.cell(row, col_num).value
    return isinstance(v, str) and v.startswith('=')


# ═══════════════════════════════════════════════════════════════════════════
# 3. Diagnóstico
# ═══════════════════════════════════════════════════════════════════════════

def diagnosticar_graficas(archivo: str) -> list[ChartInfo]:
    """
    Escanea todas las hojas del archivo y devuelve un inventario
    completo de gráficas con su clasificación.

    Clasificaciones:
      - ventana_temporal: rango horizontal que cubre múltiples períodos
      - ultimo_valor: rango vertical o punto que toma snapshot de un período
      - estatica: rangos que no dependen de períodos temporales
      - rota: contiene #REF!

    Returns:
        Lista de ChartInfo con toda la información de cada gráfica.
    """
    wb = openpyxl.load_workbook(archivo)
    resultado: list[ChartInfo] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        charts = ws._charts
        if not charts:
            continue

        for i, chart in enumerate(charts):
            tipo = type(chart).__name__
            titulo = _extraer_titulo(chart)

            refs_val = []
            refs_cat = []
            tipos_serie = set()

            for s in chart.series:
                # Valores
                if s.val and s.val.numRef:
                    ref = s.val.numRef.f
                    refs_val.append(ref)
                    tipos_serie.add(_clasificar_serie(ref))
                else:
                    refs_val.append("N/A")

                # Categorías
                if s.cat:
                    if s.cat.numRef:
                        refs_cat.append(s.cat.numRef.f)
                    elif s.cat.strRef:
                        refs_cat.append(s.cat.strRef.f)
                    else:
                        refs_cat.append("N/A")
                else:
                    refs_cat.append("N/A")

            # Clasificación
            if "rota" in tipos_serie:
                clasif = "rota"
                detalle = "Contiene #REF! — referencias rotas"
                necesita = False
            elif "horizontal" in tipos_serie and "vertical" not in tipos_serie:
                clasif = "ventana_temporal"
                detalle = "Rango horizontal (ventana de meses)"
                necesita = True
            elif "vertical" in tipos_serie and "horizontal" not in tipos_serie:
                clasif = "ultimo_valor"
                detalle = "Rango vertical (snapshot de un período)"
                necesita = True  # Se verificará después si la celda tiene fórmula
            elif "horizontal" in tipos_serie and "vertical" in tipos_serie:
                clasif = "mixta"
                detalle = "Combina rangos horizontales y verticales"
                necesita = True
            else:
                clasif = "estatica"
                detalle = "Rangos fijos sin dependencia temporal"
                necesita = False

            resultado.append(ChartInfo(
                hoja=sheet_name,
                indice=i,
                tipo_chart=tipo,
                titulo=titulo,
                num_series=len(chart.series),
                clasificacion=clasif,
                refs_valores=refs_val,
                refs_categorias=refs_cat,
                necesita_actualizacion=necesita,
                detalle=detalle,
            ))

    wb.close()

    # Imprimir resumen
    print(f"{'='*65}")
    print(f"  DIAGNÓSTICO DE GRÁFICAS: {archivo}")
    print(f"{'='*65}")
    for ci in resultado:
        icono = {"ventana_temporal": "📊", "ultimo_valor": "📍",
                 "estatica": "⚪", "rota": "⚠️", "mixta": "🔄"}.get(ci.clasificacion, "❓")
        titulo_str = ci.titulo or "Sin título"
        print(f"  {icono} [{ci.hoja}] Ch{ci.indice}: {ci.tipo_chart} | {titulo_str}")
        print(f"     {ci.clasificacion} ({ci.num_series} series) — {ci.detalle}")

    conteos = {}
    for ci in resultado:
        conteos[ci.clasificacion] = conteos.get(ci.clasificacion, 0) + 1
    print(f"\n  Resumen: {conteos}")
    print(f"{'='*65}")

    return resultado


# ═══════════════════════════════════════════════════════════════════════════
# 4. Actualización de gráficas
# ═══════════════════════════════════════════════════════════════════════════

def actualizar_graficas(
    archivo: str,
    archivo_datos: str | None = None,
    archivo_salida: str | None = None,
    borrar_rotas: bool = False,
    actualizar_titulos: bool = True,
    ventana_fija: int | None = 13,
    verbose: bool = True,
) -> ChartUpdateResult:
    """
    Actualiza automáticamente los rangos de gráficas en un archivo Excel.

    Comportamiento por tipo:
      - ventana_temporal: detecta la última columna de datos y desliza el
        rango para que termine ahí, manteniendo el ancho de ventana original
        (o ventana_fija si se especifica).
      - ultimo_valor: verifica si las celdas referenciadas contienen fórmulas.
        Si tienen fórmulas → no toca nada (las fórmulas ya se actualizan solas).
        Si son datos directos → mueve la referencia a la nueva columna.
      - rota: si borrar_rotas=True, elimina la gráfica del worksheet.
      - títulos: si actualizar_titulos=True, actualiza nombres de meses.

    Args:
        archivo:         Excel a actualizar (con fórmulas, post-exportación).
        archivo_datos:   Excel con datos actualizados para detectar la última
                         columna temporal. None = mismo archivo.
        archivo_salida:  Dónde guardar. None = sobrescribir archivo.
        borrar_rotas:    Si True, elimina gráficas con #REF!.
        actualizar_titulos: Si True, actualiza nombres de meses en títulos.
        ventana_fija:    Ancho fijo de la ventana temporal (None = mantener
                         el ancho original de cada gráfica).
        verbose:         Si True, imprime progreso.

    Returns:
        ChartUpdateResult con estadísticas de la actualización.
    """
    if archivo_datos is None:
        archivo_datos = archivo

    salida = archivo_salida or archivo
    detalle: list[str] = []
    n_actualizadas = 0
    n_borradas = 0
    n_titulos = 0

    # ── Cargar workbooks ──────────────────────────────────────────────
    wb = openpyxl.load_workbook(archivo)
    wb_datos = openpyxl.load_workbook(archivo_datos, data_only=True)

    # ── Construir mapas de fechas por hoja ────────────────────────────
    mapas_fechas: dict[str, dict[int, Any]] = {}  # hoja → {col_num: datetime}
    mapas_fila: dict[str, int] = {}               # hoja → fila de headers

    for hoja in wb_datos.sheetnames:
        ws_d = wb_datos[hoja]
        fila, mapa = _buscar_fila_fechas(ws_d)
        if fila and mapa:
            mapas_fechas[hoja] = mapa
            mapas_fila[hoja] = fila

    # ── Diagnosticar ──────────────────────────────────────────────────
    diagnostico = diagnosticar_graficas(archivo)
    n_total = len(diagnostico)

    # ── Procesar hoja por hoja ────────────────────────────────────────
    for sheet_name in list(wb.sheetnames):
        ws = wb[sheet_name]
        charts = ws._charts
        if not charts:
            continue

        charts_a_borrar = []

        for i, chart in enumerate(charts):
            # Buscar el diagnóstico correspondiente
            info = next((d for d in diagnostico
                         if d.hoja == sheet_name and d.indice == i), None)
            if info is None:
                continue

            # ── ROTAS ─────────────────────────────────────────────
            if info.clasificacion == "rota":
                if borrar_rotas:
                    charts_a_borrar.append(chart)
                    msg = f"🗑️  [{sheet_name}] Ch{i}: BORRADA ({info.titulo or 'Sin título'})"
                    n_borradas += 1
                else:
                    msg = f"⚠️  [{sheet_name}] Ch{i}: ROTA, ignorada"
                detalle.append(msg)
                if verbose:
                    print(msg)
                continue

            # ── ESTÁTICAS ─────────────────────────────────────────
            if info.clasificacion == "estatica":
                continue

            # ── VENTANA TEMPORAL ──────────────────────────────────
            if info.clasificacion in ("ventana_temporal", "mixta"):
                actualizado = _actualizar_ventana_temporal(
                    chart, info, wb_datos, mapas_fechas, mapas_fila,
                    ventana_fija, verbose,
                )
                if actualizado:
                    n_actualizadas += 1
                    msg = f"📊 [{sheet_name}] Ch{i}: ventana actualizada"
                    detalle.append(msg)
                    if verbose:
                        print(msg)

            # ── ÚLTIMO VALOR ──────────────────────────────────────
            if info.clasificacion == "ultimo_valor":
                # Verificar si las celdas tienen fórmulas
                muestra_ref = info.refs_valores[0] if info.refs_valores else ""
                parsed = _parsear_ref(muestra_ref)
                hoja_ref = parsed["hoja"] or sheet_name

                if parsed["col_start"] and parsed["row_start"]:
                    es_formula = _celda_es_formula(
                        wb, hoja_ref, parsed["col_start"], parsed["row_start"]
                    )
                    if es_formula:
                        msg = f"📍 [{sheet_name}] Ch{i}: fórmulas se auto-actualizan, sin cambios"
                        detalle.append(msg)
                        if verbose:
                            print(msg)
                        # No hacer continue aquí — dejar que llegue a títulos
                    else:
                        # Si son datos directos, mover a la última columna
                        actualizado = _actualizar_ultimo_valor(
                            chart, info, wb_datos, mapas_fechas, mapas_fila, verbose,
                        )
                        if actualizado:
                            n_actualizadas += 1
                            msg = f"📍 [{sheet_name}] Ch{i}: referencia movida a último período"
                            detalle.append(msg)
                            if verbose:
                                print(msg)

            # ── TÍTULOS ───────────────────────────────────────────
            if actualizar_titulos and info.titulo:
                titulo_nuevo = _actualizar_titulo_meses(
                    chart, wb_datos, mapas_fechas, mapas_fila, info,
                )
                if titulo_nuevo:
                    n_titulos += 1
                    msg = f"📝 [{sheet_name}] Ch{i}: título → '{titulo_nuevo}'"
                    detalle.append(msg)
                    if verbose:
                        print(msg)

        # Borrar gráficas marcadas
        for ch in charts_a_borrar:
            if ch in ws._charts:
                ws._charts.remove(ch)

    # ── Guardar ───────────────────────────────────────────────────────
    wb.save(salida)
    wb.close()
    wb_datos.close()

    resultado = ChartUpdateResult(
        archivo_salida=salida,
        graficas_analizadas=n_total,
        graficas_actualizadas=n_actualizadas,
        graficas_borradas=n_borradas,
        titulos_actualizados=n_titulos,
        detalle=detalle,
    )

    if verbose:
        print(f"\n{'='*65}")
        print(f"  ✅ RESULTADO")
        print(f"     Analizadas:   {resultado.graficas_analizadas}")
        print(f"     Actualizadas: {resultado.graficas_actualizadas}")
        print(f"     Borradas:     {resultado.graficas_borradas}")
        print(f"     Títulos:      {resultado.titulos_actualizados}")
        print(f"     Guardado en:  {Path(salida).name}")
        print(f"{'='*65}")

    return resultado


# ═══════════════════════════════════════════════════════════════════════════
# 5. Lógica interna de actualización por tipo
# ═══════════════════════════════════════════════════════════════════════════

def _actualizar_ventana_temporal(
    chart, info: ChartInfo, wb_datos, mapas_fechas, mapas_fila,
    ventana_fija, verbose,
) -> bool:
    """
    Desliza los rangos horizontales de la gráfica para que terminen
    en la última columna temporal disponible.
    """
    # Determinar la hoja de datos referenciada
    muestra_ref = info.refs_valores[0] if info.refs_valores else ""
    parsed = _parsear_ref(muestra_ref)
    hoja_datos = parsed["hoja"] or info.hoja

    if hoja_datos not in mapas_fechas:
        if verbose:
            print(f"     ⚠ No se encontraron fechas en '{hoja_datos}'")
        return False

    mapa = mapas_fechas[hoja_datos]
    if not mapa:
        return False

    # Última columna con fecha en la hoja de datos
    ultima_col_datos = max(mapa.keys())

    # Para cada serie, determinar el bloque correcto de fechas
    # (puede haber múltiples tablas en la misma hoja)
    alguna_actualizada = False

    for s in chart.series:
        for attr_name in ['val', 'cat']:
            attr = getattr(s, attr_name, None)
            if attr is None:
                continue

            num_ref = getattr(attr, 'numRef', None) or getattr(attr, 'strRef', None)
            if num_ref is None or not num_ref.f:
                continue

            ref = num_ref.f
            if '#REF!' in ref:
                continue

            tipo = _clasificar_serie(ref)
            if tipo != "horizontal":
                continue

            p = _parsear_ref(ref)
            if not (p["col_start"] and p["col_end"] and p["row_start"]):
                continue

            col_s = column_index_from_string(p["col_start"])
            col_e = column_index_from_string(p["col_end"])
            ancho_actual = col_e - col_s + 1

            # Buscar la última columna con fecha que esté en el mismo
            # "bloque" que la referencia actual (hay hojas con 2 tablas)
            cols_fechas_en_bloque = sorted(
                c for c in mapa.keys()
                if c >= col_s - ancho_actual and c <= ultima_col_datos + 5
            )

            if not cols_fechas_en_bloque:
                continue

            nueva_col_fin = cols_fechas_en_bloque[-1]

            # Determinar ancho de ventana
            if ventana_fija is not None:
                ancho = ventana_fija
            else:
                ancho = ancho_actual

            nueva_col_inicio = nueva_col_fin - ancho + 1

            # Asegurar que no se salga del rango de datos
            primera_col_datos = min(cols_fechas_en_bloque)
            if nueva_col_inicio < primera_col_datos:
                nueva_col_inicio = primera_col_datos

            # Solo actualizar si cambió algo
            if nueva_col_inicio == col_s and nueva_col_fin == col_e:
                continue

            nueva_ref = _reconstruir_ref(
                p["hoja"],
                get_column_letter(nueva_col_inicio),
                p["row_start"],
                get_column_letter(nueva_col_fin),
                p["row_end"],
            )

            num_ref.f = nueva_ref
            # Limpiar cache para forzar recálculo
            num_ref.numCache = None
            if hasattr(num_ref, 'strCache'):
                num_ref.strCache = None
            alguna_actualizada = True

    return alguna_actualizada


def _actualizar_ultimo_valor(
    chart, info: ChartInfo, wb_datos, mapas_fechas, mapas_fila, verbose,
) -> bool:
    """
    Mueve referencias verticales a la última columna temporal
    cuando apuntan a datos directos (no fórmulas).
    """
    muestra_ref = info.refs_valores[0] if info.refs_valores else ""
    parsed = _parsear_ref(muestra_ref)
    hoja_datos = parsed["hoja"] or info.hoja

    if hoja_datos not in mapas_fechas:
        return False

    mapa = mapas_fechas[hoja_datos]
    if not mapa:
        return False

    ultima_col = max(mapa.keys())
    ultima_letra = get_column_letter(ultima_col)

    alguna_actualizada = False

    for s in chart.series:
        for attr_name in ['val', 'cat']:
            attr = getattr(s, attr_name, None)
            if attr is None:
                continue

            num_ref = getattr(attr, 'numRef', None) or getattr(attr, 'strRef', None)
            if num_ref is None or not num_ref.f:
                continue

            ref = num_ref.f
            tipo = _clasificar_serie(ref)
            if tipo != "vertical":
                continue

            p = _parsear_ref(ref)
            if not p["col_start"]:
                continue

            col_actual = column_index_from_string(p["col_start"])

            # Solo mover si la columna actual es una columna de fechas
            if col_actual not in mapa:
                continue

            if col_actual == ultima_col:
                continue

            nueva_ref = _reconstruir_ref(
                p["hoja"],
                ultima_letra, p["row_start"],
                ultima_letra if p["col_end"] else None,
                p["row_end"],
            )

            num_ref.f = nueva_ref
            num_ref.numCache = None
            if hasattr(num_ref, 'strCache'):
                num_ref.strCache = None
            alguna_actualizada = True

    return alguna_actualizada


def _actualizar_titulo_meses(
    chart, wb_datos, mapas_fechas, mapas_fila, info: ChartInfo,
) -> str | None:
    """
    Actualiza nombres de meses en títulos de gráficas.

    Busca patrones como "Cambio mensual Diciembre - Enero" y los
    actualiza según los dos últimos períodos disponibles.
    """
    if not chart.title:
        return None
    tx = getattr(chart.title, 'tx', None)
    if not tx:
        return None
    rich = getattr(tx, 'rich', None)
    if not rich:
        return None

    titulo_original = _extraer_titulo(chart)
    if not titulo_original:
        return None

    # Buscar patrón "Mes1 - Mes2" o "Mes1 – Mes2" en el título
    patron_meses = re.compile(
        r'(\b(?:' + '|'.join(
            list(_MESES_ES.values()) + list(_MESES_EN.values())
        ) + r')\b)'
        r'\s*[-–]\s*'
        r'(\b(?:' + '|'.join(
            list(_MESES_ES.values()) + list(_MESES_EN.values())
        ) + r')\b)',
        re.IGNORECASE,
    )

    match = patron_meses.search(titulo_original)
    if not match:
        return None

    # Determinar idioma del título
    mes1_orig = match.group(1)
    es_espanol = mes1_orig.lower() in _MES_A_NUM_ES

    # Buscar la hoja de datos de la gráfica
    muestra_ref = info.refs_valores[0] if info.refs_valores else ""
    parsed = _parsear_ref(muestra_ref)
    hoja_datos = parsed["hoja"] or info.hoja

    if hoja_datos not in mapas_fechas:
        return None

    mapa = mapas_fechas[hoja_datos]
    cols_ordenadas = sorted(mapa.keys())
    if len(cols_ordenadas) < 2:
        return None

    # Últimos dos meses
    penultimo = mapa[cols_ordenadas[-2]]
    ultimo = mapa[cols_ordenadas[-1]]

    if es_espanol:
        mes_pen = _MESES_ES.get(penultimo.month, "")
        mes_ult = _MESES_ES.get(ultimo.month, "")
    else:
        mes_pen = _MESES_EN.get(penultimo.month, "")
        mes_ult = _MESES_EN.get(ultimo.month, "")

    nuevo_titulo = titulo_original[:match.start()] + \
                   f"{mes_pen} - {mes_ult}" + \
                   titulo_original[match.end():]

    if nuevo_titulo == titulo_original:
        return None

    # Reescribir en el rich text
    for p in rich.paragraphs:
        for r in p.r:
            if r.t and match.group(0) in (r.t or ""):
                r.t = r.t.replace(match.group(0), f"{mes_pen} - {mes_ult}")
            elif r.t and mes1_orig in (r.t or ""):
                # El título puede estar dividido en múltiples runs
                r.t = r.t.replace(mes1_orig, mes_pen)
        # Segundo mes puede estar en otro run
        for r in p.r:
            mes2_orig = match.group(2)
            if r.t and mes2_orig in (r.t or ""):
                r.t = r.t.replace(mes2_orig, mes_ult)

    return nuevo_titulo