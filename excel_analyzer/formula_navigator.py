"""
formula_navigator.py
====================
Funciones standalone para navegar y mover referencias de fórmulas
en cualquier Excel, mapeando siempre por nombre de métrica y columna,
nunca por coordenadas de celda.

FUNCIONES PRINCIPALES:
  mover_a_ultima_columna(archivo, hoja, ...)
    → Fórmulas que apuntan a la última columna temporal se mueven
      a la columna siguiente (o N columnas adelante).

  mover_a_ultima_fila(archivo, hoja, ...)
    → Fórmulas que apuntan a la última fila de datos se mueven
      a la fila siguiente (o N filas adelante).

  recorrer_columnas(archivo, hoja, n, ...)
    → Mueve TODAS las referencias de fórmulas N columnas a la derecha
      (o izquierda si n es negativo).

  recorrer_filas(archivo, hoja, n, ...)
    → Mueve TODAS las referencias de fórmulas N filas abajo
      (o arriba si n es negativo).

  inspeccionar_formulas(archivo, hoja)
    → Muestra todas las fórmulas con su mapeo semántico:
      celda, fórmula original, métrica, columna referenciada.

FILOSOFÍA:
  Todo funciona por mapeo semántico.  Internamente el código:
    1. Detecta la tabla con detectar_todas_las_tablas
    2. Construye el mapa: letra Excel → nombre columna, fila → métrica
    3. Identifica qué columna/fila es "la última" temporal
    4. Reescribe solo las fórmulas que apuntan a esa posición
    5. Guarda el resultado

  Nunca se mueve "la celda AP30" — se mueve "la referencia a la
  columna ene-26 de la fila Azteca".
"""

import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

sys.path.insert(0, str(Path(__file__).parent))
from _core import detectar_todas_las_tablas, _is_date_string, SheetScanner

# Regex para refs de celda (con soporte $)
_CELL_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

# Regex para refs cruzadas
_CROSS_REF = re.compile(
    r"(?:'([^']+)'|([A-Za-zÁÉÍÓÚáéíóúñÑ0-9_]+))"
    r"!"
    r"(\$?)([A-Z]{1,3})(\$?)(\d+)"
)


# ═══════════════════════════════════════════════════════════════════════════
# 1. Mapa semántico
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class TablaMap:
    """Mapa semántico completo de una tabla detectada."""
    # Mapeo de coordenadas
    col_num_a_nombre: dict[int, str]    # col Excel (base-1) → nombre header
    fila_num_a_metrica: dict[int, str]  # fila Excel (base-1) → nombre métrica
    nombre_a_col_num: dict[str, int]    # nombre header → col Excel (base-1)
    metrica_a_fila_num: dict[str, int]  # nombre métrica → fila Excel (base-1)

    # Estructura
    data_start_row: int                 # primera fila de datos (base-1)
    data_end_row: int                   # última fila de datos (base-1)
    col_start: int                      # primera columna (base-1)
    col_end: int                        # última columna (base-1)

    # Columnas temporales ordenadas
    columnas_temporales: list[str]      # ["jul-24", "ago-24", ..., "ene-26"]
    ultima_col_temporal: str | None     # "ene-26"
    ultima_col_num: int | None          # número de columna Excel de la última temporal


def _encontrar_header_row(ws, cs: int, ce: int) -> int | None:
    """Header real saltando merged titles."""
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(cs, ce + 1)]
        nn = [v for v in vals if v is not None]
        if len(nn) < 2:
            continue
        if all(isinstance(v, str) for v in nn) and len(set(str(v).strip() for v in nn)) == 1:
            continue
        if sum(1 for v in nn if isinstance(v, str)) >= len(nn) * 0.5:
            return r
    return None


def construir_mapa(archivo: str, hoja: str, tabla_idx: int = -1) -> TablaMap | None:
    """
    Construye el mapa semántico completo de una tabla.

    tabla_idx: -1 = auto (prefiere tabla con columnas temporales, sino la más grande).
               0+ = índice específico.
    """
    tablas = detectar_todas_las_tablas(archivo, hoja)
    if not tablas:
        return None

    if tabla_idx >= 0:
        if tabla_idx < len(tablas):
            t = tablas[tabla_idx]
        else:
            return None
    else:
        # Preferir tabla con más columnas temporales (col_headers con fechas)
        mejor_temporal = None
        max_temp = 0
        for candidata in tablas:
            df_c = candidata.get("data")
            if df_c is None:
                continue
            n_temp = sum(1 for c in df_c.columns if _is_date_string(str(c)))
            # También contar col_headers
            for _, nombre in candidata.get("col_headers", {}).items():
                if _is_date_string(str(nombre)):
                    n_temp += 1
            if n_temp > max_temp:
                max_temp = n_temp
                mejor_temporal = candidata

        if mejor_temporal and max_temp >= 2:
            t = mejor_temporal
        else:
            # Fallback: la más grande
            t = max(tablas, key=lambda x: x["data"].shape[0] * x["data"].shape[1] if x.get("data") is not None else 0)
    df = t["data"]
    if df is None:
        return None

    cs = t.get("col_inicio", 1)
    ce = t.get("col_fin", 1)
    fila_header_b1 = t.get("fila_header")

    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb[hoja]

    hr = fila_header_b1 if fila_header_b1 else _encontrar_header_row(ws, cs, ce)
    if hr is None:
        wb.close()
        return None

    data_start = hr + 1

    # Col num → nombre: primero del workbook, luego enriquecer con col_headers de la tabla
    col_num_a_nombre: dict[int, str] = {}
    nombre_a_col_num: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hr, c).value
        if hasattr(v, 'strftime'):
            from _core import SheetScanner
            nombre = SheetScanner.fmt_fecha(v)
        elif v is not None:
            nombre = str(v).strip()
        else:
            nombre = get_column_letter(c)
        col_num_a_nombre[c] = nombre
        nombre_a_col_num[nombre] = c

    # Enriquecer con col_headers de la tabla detectada (más confiable para multinivel)
    col_headers = t.get("col_headers", {})
    for col_b1, nombre in col_headers.items():
        col_num_a_nombre[col_b1] = nombre
        nombre_a_col_num[nombre] = col_b1

    # Fila num → métrica (primera columna de texto como identificador)
    fila_num_a_metrica: dict[int, str] = {}
    metrica_a_fila_num: dict[str, int] = {}

    # Encontrar la columna de métricas (primera col con mayoría strings)
    col_metrica = cs
    for c in range(cs, ce + 1):
        n_str = sum(
            1 for r in range(data_start, ws.max_row + 1)
            if isinstance(ws.cell(r, c).value, str)
        )
        n_total = sum(
            1 for r in range(data_start, ws.max_row + 1)
            if ws.cell(r, c).value is not None
        )
        if n_total > 0 and n_str / n_total >= 0.5:
            col_metrica = c
            break

    for r in range(data_start, ws.max_row + 1):
        v = ws.cell(r, col_metrica).value
        if v is not None:
            nombre = str(v).strip()
            fila_num_a_metrica[r] = nombre
            metrica_a_fila_num[nombre] = r

    # Columnas temporales: usar DataFrame columns como fuente autoritativa
    df_cols = list(df.columns)
    cols_temp: list[tuple[int, str]] = []
    for col_name in df_cols:
        col_str = str(col_name)
        if _is_date_string(col_str):
            col_num = nombre_a_col_num.get(col_str)
            if col_num:
                cols_temp.append((col_num, col_str))

    # Fallback: buscar en col_num_a_nombre directamente
    if not cols_temp:
        for c, nombre in col_num_a_nombre.items():
            if _is_date_string(nombre):
                cols_temp.append((c, nombre))

    cols_temp.sort(key=lambda x: x[0])
    columnas_temporales = [n for _, n in cols_temp]
    ultima_col = cols_temp[-1] if cols_temp else None

    data_end = max(fila_num_a_metrica.keys()) if fila_num_a_metrica else data_start

    wb.close()
    return TablaMap(
        col_num_a_nombre=col_num_a_nombre,
        fila_num_a_metrica=fila_num_a_metrica,
        nombre_a_col_num=nombre_a_col_num,
        metrica_a_fila_num=metrica_a_fila_num,
        data_start_row=data_start,
        data_end_row=data_end,
        col_start=cs,
        col_end=ce,
        columnas_temporales=columnas_temporales,
        ultima_col_temporal=ultima_col[1] if ultima_col else None,
        ultima_col_num=ultima_col[0] if ultima_col else None,
    )


# ═══════════════════════════════════════════════════════════════════════════
# 2. Inspección
# ═══════════════════════════════════════════════════════════════════════════

def inspeccionar_formulas(archivo: str, hoja: str, tabla_idx: int = -1) -> list[dict]:
    """
    Muestra todas las fórmulas con su mapeo semántico.

    Retorna lista de dicts:
      {"celda": "E8", "formula": "=D8/$H$3",
       "metrica": "Azteca", "columna": "AUM USD ene-26",
       "refs": [{"col_nombre": "AUM MXN ene-25", "fila_metrica": "Azteca"}, ...]}
    """
    mapa = construir_mapa(archivo, hoja, tabla_idx)
    if mapa is None:
        print(f"No se pudo construir mapa de '{hoja}'")
        return []

    wb = openpyxl.load_workbook(archivo)
    ws = wb[hoja]

    resultado: list[dict] = []

    for r in range(mapa.data_start_row, ws.max_row + 1):
        for c in range(mapa.col_start, mapa.col_end + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and v.startswith('=')):
                continue

            metrica = mapa.fila_num_a_metrica.get(r, f"fila_{r}")
            columna = mapa.col_num_a_nombre.get(c, f"col_{c}")

            # Parsear refs dentro de la fórmula
            refs = []
            for m in _CELL_REF.finditer(v):
                ref_col = column_index_from_string(m.group(2))
                ref_row = int(m.group(4))
                refs.append({
                    "col_nombre": mapa.col_num_a_nombre.get(ref_col, m.group(2)),
                    "fila_metrica": mapa.fila_num_a_metrica.get(ref_row, f"fila_{ref_row}"),
                    "col_num": ref_col,
                    "fila_num": ref_row,
                })

            resultado.append({
                "celda": f"{get_column_letter(c)}{r}",
                "formula": v,
                "metrica": metrica,
                "columna": columna,
                "refs": refs,
            })

    wb.close()

    # Imprimir resumen
    print(f"{'='*60}")
    print(f"  {hoja}: {len(resultado)} fórmulas")
    print(f"  Última columna temporal: {mapa.ultima_col_temporal}")
    print(f"  Columnas temporales: {mapa.columnas_temporales[:5]}...{mapa.columnas_temporales[-3:]}")
    print(f"{'='*60}")
    for f in resultado[:10]:
        refs_str = ", ".join(f"{r['col_nombre']}[{r['fila_metrica']}]" for r in f["refs"])
        print(f"  {f['metrica']:15s} | {f['columna']:12s} | {f['formula']:25s} → {refs_str}")
    if len(resultado) > 10:
        print(f"  ... y {len(resultado) - 10} más")

    return resultado


# ═══════════════════════════════════════════════════════════════════════════
# 3. Mover a última columna / fila
# ═══════════════════════════════════════════════════════════════════════════

def mover_a_ultima_columna(
    archivo: str,
    hoja: str,
    archivo_salida: str | None = None,
    n: int = 1,
    tabla_idx: int = -1,
) -> dict[str, Any]:
    """
    Mueve las fórmulas que referencian la última columna temporal
    N columnas a la derecha.

    Ejemplo: si la última columna es "ene-26" (col 42) y n=1,
    las fórmulas que apuntan a col 42 ahora apuntan a col 43.

    Mapeo semántico: identifica "última columna" por nombre de header
    temporal, no por número de columna.

    Args:
        archivo:         Excel de entrada.
        hoja:            Hoja a modificar.
        archivo_salida:  Excel de salida. None = sobrescribir.
        n:               Columnas a mover (positivo=derecha, negativo=izquierda).
        tabla_idx:       Índice de la tabla (0 = la más grande).

    Returns:
        {"ultima_col": "ene-26", "nueva_col_num": 43, "formulas_movidas": 15}
    """
    mapa = construir_mapa(archivo, hoja, tabla_idx)
    if mapa is None or mapa.ultima_col_num is None:
        raise ValueError(f"No se detectó columna temporal en '{hoja}'")

    col_objetivo = mapa.ultima_col_num
    col_destino = col_objetivo + n

    return _mover_refs(
        archivo, hoja, archivo_salida,
        col_objetivo=col_objetivo, col_destino=col_destino,
        fila_objetivo=None, fila_destino=None,
        mapa=mapa,
        descripcion=f"última col '{mapa.ultima_col_temporal}' +{n}",
    )


def mover_a_ultima_fila(
    archivo: str,
    hoja: str,
    archivo_salida: str | None = None,
    n: int = 1,
    tabla_idx: int = -1,
) -> dict[str, Any]:
    """
    Mueve las fórmulas que referencian la última fila de datos
    N filas abajo.

    Args:
        n: Filas a mover (positivo=abajo, negativo=arriba).
    """
    mapa = construir_mapa(archivo, hoja, tabla_idx)
    if mapa is None:
        raise ValueError(f"No se pudo construir mapa de '{hoja}'")

    fila_objetivo = mapa.data_end_row
    fila_destino = fila_objetivo + n
    metrica_ultima = mapa.fila_num_a_metrica.get(fila_objetivo, f"fila_{fila_objetivo}")

    return _mover_refs(
        archivo, hoja, archivo_salida,
        col_objetivo=None, col_destino=None,
        fila_objetivo=fila_objetivo, fila_destino=fila_destino,
        mapa=mapa,
        descripcion=f"última fila '{metrica_ultima}' +{n}",
    )


# ═══════════════════════════════════════════════════════════════════════════
# 4. Recorrer N columnas / filas
# ═══════════════════════════════════════════════════════════════════════════

def recorrer_columnas(
    archivo: str,
    hoja: str,
    n: int,
    archivo_salida: str | None = None,
    tabla_idx: int = -1,
) -> dict[str, Any]:
    """
    Mueve TODAS las referencias de fórmulas N columnas.

    n > 0: hacia la derecha.  n < 0: hacia la izquierda.
    """
    mapa = construir_mapa(archivo, hoja, tabla_idx)
    if mapa is None:
        raise ValueError(f"No se pudo construir mapa de '{hoja}'")

    return _recorrer_refs(
        archivo, hoja, archivo_salida,
        delta_col=n, delta_fila=0, mapa=mapa,
    )


def recorrer_filas(
    archivo: str,
    hoja: str,
    n: int,
    archivo_salida: str | None = None,
    tabla_idx: int = -1,
) -> dict[str, Any]:
    """
    Mueve TODAS las referencias de fórmulas N filas.

    n > 0: hacia abajo.  n < 0: hacia arriba.
    """
    mapa = construir_mapa(archivo, hoja, tabla_idx)
    if mapa is None:
        raise ValueError(f"No se pudo construir mapa de '{hoja}'")

    return _recorrer_refs(
        archivo, hoja, archivo_salida,
        delta_col=0, delta_fila=n, mapa=mapa,
    )


# ═══════════════════════════════════════════════════════════════════════════
# 5. Motor interno de reescritura
# ═══════════════════════════════════════════════════════════════════════════

def _mover_refs(
    archivo: str,
    hoja: str,
    archivo_salida: str | None,
    col_objetivo: int | None,
    col_destino: int | None,
    fila_objetivo: int | None,
    fila_destino: int | None,
    mapa: TablaMap,
    descripcion: str = "",
) -> dict[str, Any]:
    """
    Reescribe fórmulas moviendo refs que apuntan a col/fila objetivo
    hacia col/fila destino.  Solo mueve las refs que coinciden exactamente.
    """
    wb = openpyxl.load_workbook(archivo)
    ws = wb[hoja]
    movidas = 0

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and v.startswith('=')):
                continue

            nueva = _reescribir_refs_selectivo(
                v, col_objetivo, col_destino, fila_objetivo, fila_destino
            )
            if nueva != v:
                ws.cell(r, c).value = nueva
                movidas += 1

    salida = archivo_salida or archivo
    wb.save(salida)
    wb.close()

    print(f"✅ {descripcion}: {movidas} fórmulas movidas → {Path(salida).name}")
    return {
        "descripcion": descripcion,
        "formulas_movidas": movidas,
        "archivo_salida": salida,
    }


def _recorrer_refs(
    archivo: str,
    hoja: str,
    archivo_salida: str | None,
    delta_col: int,
    delta_fila: int,
    mapa: TablaMap,
) -> dict[str, Any]:
    """Recorre TODAS las refs en fórmulas por delta_col y delta_fila."""
    wb = openpyxl.load_workbook(archivo)
    ws = wb[hoja]
    movidas = 0

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and v.startswith('=')):
                continue

            nueva = _reescribir_refs_delta(v, delta_col, delta_fila)
            if nueva != v:
                ws.cell(r, c).value = nueva
                movidas += 1

    salida = archivo_salida or archivo
    wb.save(salida)
    wb.close()

    desc = f"recorrer cols={delta_col:+d} filas={delta_fila:+d}"
    print(f"✅ {desc}: {movidas} fórmulas movidas → {Path(salida).name}")
    return {
        "descripcion": desc,
        "formulas_movidas": movidas,
        "archivo_salida": salida,
    }


def _reescribir_refs_selectivo(
    formula: str,
    col_obj: int | None,
    col_dest: int | None,
    fila_obj: int | None,
    fila_dest: int | None,
) -> str:
    """Reescribe solo las refs que coinciden con col/fila objetivo."""
    offset = 0
    nuevo = list(formula)

    for m in _CELL_REF.finditer(formula):
        dol_c = m.group(1)
        col_letter = m.group(2)
        dol_r = m.group(3)
        row_num = int(m.group(4))

        col_num = column_index_from_string(col_letter)
        new_col = col_num
        new_row = row_num

        # Mover columna si coincide
        if col_obj is not None and col_dest is not None and col_num == col_obj:
            new_col = col_dest

        # Mover fila si coincide
        if fila_obj is not None and fila_dest is not None and row_num == fila_obj:
            new_row = fila_dest

        if new_col == col_num and new_row == row_num:
            continue

        new_letter = get_column_letter(new_col)
        reemplazo = f"{dol_c}{new_letter}{dol_r}{new_row}"
        s, e = m.start() + offset, m.end() + offset
        nuevo[s:e] = list(reemplazo)
        offset += len(reemplazo) - (e - s)

    return "".join(nuevo)


def _reescribir_refs_delta(
    formula: str,
    delta_col: int,
    delta_fila: int,
) -> str:
    """Recorre todas las refs por delta."""
    offset = 0
    nuevo = list(formula)

    for m in _CELL_REF.finditer(formula):
        dol_c = m.group(1)
        col_letter = m.group(2)
        dol_r = m.group(3)
        row_num = int(m.group(4))

        col_num = column_index_from_string(col_letter)

        # No mover refs absolutas (con $)
        new_col = col_num if dol_c else col_num + delta_col
        new_row = row_num if dol_r else row_num + delta_fila

        if new_col < 1:
            new_col = 1
        if new_row < 1:
            new_row = 1

        if new_col == col_num and new_row == row_num:
            continue

        new_letter = get_column_letter(new_col)
        reemplazo = f"{dol_c}{new_letter}{dol_r}{new_row}"
        s, e = m.start() + offset, m.end() + offset
        nuevo[s:e] = list(reemplazo)
        offset += len(reemplazo) - (e - s)

    return "".join(nuevo)

# ═══════════════════════════════════════════════════════════════════════════
# 6. Mapeo por período y apuntar a último
# ═══════════════════════════════════════════════════════════════════════════

def _construir_mapa_fechas_hoja(archivo: str, hoja: str) -> dict[str, str]:
    """
    Construye mapa letra_columna → nombre_periodo para una hoja.
    Busca la fila con más fechas y mapea cada columna a su período.
    Retorna {"P": "ene-25", "AA": "dic-25", "AB": "ene-26", ...}
    """
    wb = openpyxl.load_workbook(archivo, data_only=True)
    if hoja not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[hoja]

    mejor_fila = None
    max_fechas = 0
    for r in range(1, min(20, ws.max_row + 1)):
        n = sum(1 for c in range(1, ws.max_column + 1)
                if hasattr(ws.cell(r, c).value, 'strftime')
                or (isinstance(ws.cell(r, c).value, str) and _is_date_string(ws.cell(r, c).value)))
        if n > max_fechas:
            max_fechas = n
            mejor_fila = r

    if mejor_fila is None or max_fechas == 0:
        wb.close()
        return {}

    mapa: dict[str, str] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(mejor_fila, c).value
        if hasattr(v, 'strftime'):
            mapa[get_column_letter(c)] = SheetScanner.fmt_fecha(v)
        elif isinstance(v, str) and _is_date_string(v):
            mapa[get_column_letter(c)] = SheetScanner.fmt_fecha(v)

    wb.close()
    return mapa


def mapear_por_periodo(
    archivo: str,
    hoja: str,
    archivo_datos: str | None = None,
    archivo_salida: str | None = None,
) -> dict[str, Any]:
    """
    Reescribe fórmulas cruzadas para que matcheen por nombre de período.

    Lee las hojas de datos, construye mapa letra → período, y reescribe
    cada referencia para que apunte a la columna con el mismo período
    sin importar si se insertaron o movieron columnas.

    Args:
        archivo:         Excel con las fórmulas.
        hoja:            Hoja con fórmulas a reescribir.
        archivo_datos:   Excel con datos actualizados. None = mismo archivo.
        archivo_salida:  Dónde guardar. None = sobrescribir archivo.
    """
    if archivo_datos is None:
        archivo_datos = archivo

    CROSS = re.compile(
        r"((?:'[^']+?'|[A-Za-zÁÉÍÓÚáéíóúñÑ0-9_]+)!)"
        r"(\$?)([A-Z]{1,3})(\$?)(\d+)"
    )

    wb_formulas = openpyxl.load_workbook(archivo)
    ws = wb_formulas[hoja]

    # Recolectar hojas referenciadas
    hojas_refs: set[str] = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and '!' in v):
                continue
            for m in re.finditer(r"(?:'([^']+)'|([A-Za-z0-9_]+))!", v):
                h = (m.group(1) or m.group(2)).strip()
                if h:
                    hojas_refs.add(h)

    # Mapas original y nuevo
    mapas_orig: dict[str, dict[str, str]] = {}
    mapas_nuevo: dict[str, dict[str, str]] = {}
    for h in hojas_refs:
        mapas_orig[h] = _construir_mapa_fechas_hoja(archivo, h)
        mapas_nuevo[h] = _construir_mapa_fechas_hoja(archivo_datos, h)

    # Invertir mapas nuevos: período → letra nueva
    periodo_a_letra_nueva: dict[str, dict[str, str]] = {}
    for h, mapa in mapas_nuevo.items():
        periodo_a_letra_nueva[h] = {periodo: letra for letra, periodo in mapa.items()}

    # Reescribir
    n_movidas = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and '!' in v):
                continue

            nueva = v
            for m in CROSS.finditer(v):
                hoja_ref = m.group(1).replace("!", "").replace("'", "").strip()
                dol_c = m.group(2)
                col_letter = m.group(3)
                dol_r = m.group(4)
                row_num = m.group(5)

                periodo = mapas_orig.get(hoja_ref, {}).get(col_letter)
                if periodo is None:
                    continue

                letra_nueva = periodo_a_letra_nueva.get(hoja_ref, {}).get(periodo)
                if letra_nueva is None or letra_nueva == col_letter:
                    continue

                original = m.group(0)
                reescrita = f"{m.group(1)}{dol_c}{letra_nueva}{dol_r}{row_num}"
                nueva = nueva.replace(original, reescrita)

            if nueva != v:
                ws.cell(r, c).value = nueva
                n_movidas += 1

    salida = archivo_salida or archivo
    wb_formulas.save(salida)
    wb_formulas.close()

    print(f"✅ mapear_por_periodo({hoja}): {n_movidas} fórmulas remapeadas → {Path(salida).name}")
    return {"hoja": hoja, "formulas_movidas": n_movidas}


def apuntar_a_ultimo(
    archivo: str,
    hoja: str,
    archivo_datos: str | None = None,
    archivo_salida: str | None = None,
) -> dict[str, Any]:
    """
    Reescribe fórmulas cruzadas para que apunten a la última columna
    temporal de la hoja de datos.

    Caso de uso: Resumen_AUM_Siefore tiene ='Activos Netos'!AB34.
    AB = ene-26 (última). Si se agrega feb-26 como AC, reescribe a AC34.

    Args:
        archivo:         Excel con las fórmulas.
        hoja:            Hoja con fórmulas a reescribir.
        archivo_datos:   Excel con datos actualizados. None = mismo archivo.
        archivo_salida:  Dónde guardar. None = sobrescribir archivo.
    """
    if archivo_datos is None:
        archivo_datos = archivo

    CROSS = re.compile(
        r"((?:'[^']+?'|[A-Za-zÁÉÍÓÚáéíóúñÑ0-9_]+)!)"
        r"(\$?)([A-Z]{1,3})(\$?)(\d+)"
    )

    wb_formulas = openpyxl.load_workbook(archivo)
    ws = wb_formulas[hoja]

    # Recolectar hojas y columnas usadas
    hojas_cols: dict[str, set[str]] = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and '!' in v):
                continue
            for m in CROSS.finditer(v):
                hoja_ref = m.group(1).replace("!", "").replace("'", "").strip()
                hojas_cols.setdefault(hoja_ref, set()).add(m.group(3))

    # Para cada hoja: encontrar última col temporal original vs nueva
    col_remap: dict[str, dict[str, str]] = {}

    for hoja_ref, cols_usadas in hojas_cols.items():
        mapa_orig = _construir_mapa_fechas_hoja(archivo, hoja_ref)
        mapa_nuevo = _construir_mapa_fechas_hoja(archivo_datos, hoja_ref)

        if not mapa_orig or not mapa_nuevo:
            continue

        ultima_orig = max(mapa_orig.keys(), key=lambda l: column_index_from_string(l))
        ultima_nueva = max(mapa_nuevo.keys(), key=lambda l: column_index_from_string(l))

        if ultima_orig == ultima_nueva:
            continue

        remap = {col: ultima_nueva for col in cols_usadas if col == ultima_orig}
        if remap:
            col_remap[hoja_ref] = remap

    # Reescribir
    n_movidas = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and '!' in v):
                continue

            nueva = v
            for m in CROSS.finditer(v):
                hoja_ref = m.group(1).replace("!", "").replace("'", "").strip()
                col_letter = m.group(3)
                nueva_letra = col_remap.get(hoja_ref, {}).get(col_letter)
                if nueva_letra is None:
                    continue
                original = m.group(0)
                reescrita = f"{m.group(1)}{m.group(2)}{nueva_letra}{m.group(4)}{m.group(5)}"
                nueva = nueva.replace(original, reescrita)

            if nueva != v:
                ws.cell(r, c).value = nueva
                n_movidas += 1

    salida = archivo_salida or archivo
    wb_formulas.save(salida)
    wb_formulas.close()

    print(f"✅ apuntar_a_ultimo({hoja}): {n_movidas} fórmulas actualizadas → {Path(salida).name}")
    return {"hoja": hoja, "formulas_movidas": n_movidas}


def actualizar_a_ultimo(
    archivo: str,
    hoja: str,
    archivo_salida: str | None = None,
    fila_inicio: int | None = None,
    fila_fin: int | None = None,
) -> dict[str, Any]:
    """
    Detecta fórmulas cruzadas que apuntan a la PENÚLTIMA columna temporal
    de sus hojas de datos y las mueve a la ÚLTIMA. Trabaja con un solo archivo.

    Args:
        archivo:        Excel con fórmulas y datos ya transferidos.
        hoja:           Hoja con fórmulas a actualizar.
        archivo_salida: Dónde guardar. None = sobrescribir.
        fila_inicio:    Solo procesar desde esta fila. None = fila 1.
        fila_fin:       Solo procesar hasta esta fila. None = última fila.
    """
    CROSS = re.compile(
        r"((?:'[^']+?'|[A-Za-zÁÉÍÓÚáéíóúñÑ0-9_ ]+)!)"
        r"(\$?)([A-Z]{1,3})(\$?)(\d+)"
    )

    wb = openpyxl.load_workbook(archivo)
    ws = wb[hoja]

    # Paso 1: recolectar hojas referenciadas y columnas usadas
    ri = fila_inicio or 1
    rf = fila_fin or ws.max_row
    hojas_cols: dict[str, set[str]] = {}
    for r in range(ri, rf + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and '!' in v):
                continue
            for m in CROSS.finditer(v):
                hoja_ref = m.group(1).replace("!", "").replace("'", "").strip()
                hojas_cols.setdefault(hoja_ref, set()).add(m.group(3))

    # Paso 2: para cada hoja de datos, encontrar penúltima y última col temporal
    col_remap: dict[str, dict[str, str]] = {}

    for hoja_ref, cols_usadas in hojas_cols.items():
        if hoja_ref not in wb.sheetnames:
            continue
        mapa = _construir_mapa_fechas_hoja(archivo, hoja_ref)
        if not mapa or len(mapa) < 2:
            continue

        # Ordenar por posición de columna
        cols_ordenadas = sorted(mapa.keys(), key=lambda l: column_index_from_string(l))
        penultima = cols_ordenadas[-2]
        ultima = cols_ordenadas[-1]

        # Mapear: refs que apuntan a la penúltima → última
        remap = {}
        for col in cols_usadas:
            if col == penultima:
                remap[col] = ultima
        if remap:
            col_remap[hoja_ref] = remap

    if not col_remap:
        wb.close()
        print(f"⚠️  actualizar_a_ultimo({hoja}): fórmulas ya apuntan a la última columna")
        return {"hoja": hoja, "formulas_movidas": 0}

    # Paso 3: reescribir fórmulas
    n_movidas = 0
    for r in range(ri, rf + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and '!' in v):
                continue

            nueva = v
            for m in CROSS.finditer(v):
                hoja_ref = m.group(1).replace("!", "").replace("'", "").strip()
                col_letter = m.group(3)
                nueva_letra = col_remap.get(hoja_ref, {}).get(col_letter)
                if nueva_letra is None:
                    continue
                original = m.group(0)
                reescrita = f"{m.group(1)}{m.group(2)}{nueva_letra}{m.group(4)}{m.group(5)}"
                nueva = nueva.replace(original, reescrita)

            if nueva != v:
                ws.cell(r, c).value = nueva
                n_movidas += 1

    salida = archivo_salida or archivo
    wb.save(salida)
    wb.close()

    # Log detallado
    for hr, remap in col_remap.items():
        for old, new in remap.items():
            print(f"  📌 {hr}: {old} → {new}")
    print(f"✅ actualizar_a_ultimo({hoja}): {n_movidas} fórmulas actualizadas → {Path(salida).name}")
    return {"hoja": hoja, "formulas_movidas": n_movidas}


# ═══════════════════════════════════════════════════════════════════════════
# 7. Agregar columna replicando fórmulas de la última
# ═══════════════════════════════════════════════════════════════════════════

def agregar_columna_formulas(
    archivo: str,
    hoja: str,
    bloques: list[tuple[int, int]],
    archivo_salida: str | None = None,
    encabezado_fila: int | None = None,
    nuevo_encabezado: str | None = None,
) -> dict[str, Any]:
    """
    Agrega una columna nueva replicando las fórmulas de la última columna
    de cada bloque, incrementando las referencias de columna en +1.

    Diseñada para hojas tipo Mutual Funds / Mandatos donde cada mes
    se agrega una columna con el mismo patrón de fórmulas pero apuntando
    a la columna siguiente en las hojas de datos.

    Args:
        archivo:          Excel a modificar.
        hoja:             Nombre de la hoja.
        bloques:          Lista de (fila_inicio, fila_fin) inclusivos.
                          Cada bloque se procesa independientemente.
        archivo_salida:   Dónde guardar. None = sobrescribir.
        encabezado_fila:  Fila donde poner el nombre del nuevo período.
        nuevo_encabezado: Texto del encabezado (ej. "mar-26").

    Returns:
        {"hoja": str, "columna_nueva": str, "formulas_creadas": int}
    """
    wb = openpyxl.load_workbook(archivo)
    ws = wb[hoja]

    # Encontrar la última columna con datos en los bloques
    ultima_col = 0
    for fila_ini, fila_fin in bloques:
        for r in range(fila_ini, fila_fin + 1):
            for c in range(ws.max_column, 0, -1):
                v = ws.cell(r, c).value
                if v is not None:
                    ultima_col = max(ultima_col, c)
                    break

    nueva_col = ultima_col + 1
    if ultima_col == 0:
        wb.close()
        print(f"⚠️  agregar_columna_formulas({hoja}): bloques vacíos, nada que hacer")
        return {"hoja": hoja, "columna_nueva": "", "formulas_creadas": 0}
    n_creadas = 0

    for fila_ini, fila_fin in bloques:
        for r in range(fila_ini, fila_fin + 1):
            v = ws.cell(r, ultima_col).value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith('='):
                # Incrementar refs de columna no-absolutas en +1
                nueva_formula = _incrementar_refs_col(v)
                ws.cell(r, nueva_col).value = nueva_formula
                n_creadas += 1
            else:
                # Copiar valor literal (ej. tipo de cambio, labels)
                ws.cell(r, nueva_col).value = v

    if encabezado_fila and nuevo_encabezado:
        ws.cell(encabezado_fila, nueva_col).value = nuevo_encabezado

    salida = archivo_salida or archivo
    wb.save(salida)
    wb.close()

    col_letra = get_column_letter(nueva_col)
    print(f"✅ agregar_columna_formulas({hoja}): col {col_letra}, "
          f"{n_creadas} fórmulas creadas → {Path(salida).name}")
    return {"hoja": hoja, "columna_nueva": col_letra, "formulas_creadas": n_creadas}


def _incrementar_refs_col(formula: str) -> str:
    """
    Incrementa en +1 todas las referencias de columna en una fórmula,
    tanto locales como cross-sheet, respetando refs absolutas ($).

    Ej: =SUM(ESF_BP!AB131,...) → =SUM(ESF_BP!AC131,...)
        =Y4/$Z$27            → =Z4/$Z$27  (Z no se mueve por $)
    """
    CROSS = re.compile(
        r"((?:'[^']+?'|[A-Za-zÁÉÍÓÚáéíóúñÑ0-9_]+)!)"
        r"(\$?)([A-Z]{1,3})(\$?)(\d+)"
    )

    # Recolectar posiciones de cross-refs para excluirlas del paso local
    cross_spans = set()
    for m in CROSS.finditer(formula):
        # El span de la parte col+row (después del !)
        # Necesitamos marcar la posición de la celda ref completa
        cross_spans.add((m.start(), m.end()))

    # Paso 1: incrementar cross-refs
    resultado = formula
    offset = 0
    reemplazos = []
    for m in CROSS.finditer(formula):
        dol_c = m.group(2)
        col_letter = m.group(3)
        dol_r = m.group(4)
        row_num = m.group(5)
        if dol_c:
            continue
        new_col = get_column_letter(column_index_from_string(col_letter) + 1)
        # Reconstruir solo la parte después del !
        viejo = f"{m.group(2)}{m.group(3)}{m.group(4)}{m.group(5)}"
        nuevo = f"{dol_c}{new_col}{dol_r}{row_num}"
        # Posición después del !
        excl_pos = m.group(0).index('!') + 1
        reemplazos.append((m.start() + excl_pos + len(m.group(1)) - len(m.group(1)),
                           m.start(2), m.end(5), viejo, nuevo))

    # Aplicar cross-ref reemplazos de atrás para adelante
    chars = list(formula)
    for m in reversed(list(CROSS.finditer(formula))):
        dol_c = m.group(2)
        if dol_c:
            continue
        col_letter = m.group(3)
        dol_r = m.group(4)
        row_num = m.group(5)
        new_col = get_column_letter(column_index_from_string(col_letter) + 1)
        nuevo_ref = f"{dol_c}{new_col}{dol_r}{row_num}"
        # Reemplazar solo la parte col+row (groups 2-5)
        chars[m.start(2):m.end(5)] = list(nuevo_ref)

    resultado = "".join(chars)

    # Paso 2: incrementar refs locales (sin ! delante)
    # Reconstruir cross-ref spans en el resultado nuevo
    cross_positions = set()
    for m in CROSS.finditer(resultado):
        for i in range(m.start(2), m.end(5)):
            cross_positions.add(i)

    LOCAL = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')
    chars2 = list(resultado)
    for m in reversed(list(LOCAL.finditer(resultado))):
        # Saltar si está dentro de una cross-ref
        if any(i in cross_positions for i in range(m.start(), m.end())):
            continue
        # Saltar si hay ! justo antes
        if m.start() > 0 and resultado[m.start() - 1] == '!':
            continue
        dol_c = m.group(1)
        if dol_c:
            continue
        col_letter = m.group(2)
        dol_r = m.group(3)
        row_num = m.group(4)
        # Saltar si parece nombre de función (SUM, IF, etc)
        if col_letter in ('SUM', 'IF', 'AND', 'OR', 'NOT', 'MAX', 'MIN',
                          'AVERAGE', 'COUNT', 'ROUND', 'ABS', 'INDEX',
                          'MATCH', 'VLOOKUP', 'HLOOKUP', 'OFFSET'):
            continue
        new_col = get_column_letter(column_index_from_string(col_letter) + 1)
        nuevo = f"{dol_c}{new_col}{dol_r}{row_num}"
        chars2[m.start():m.end()] = list(nuevo)

    return "".join(chars2)


# ═══════════════════════════════════════════════════════════════════════════
# 8. Recorrer fórmulas en un rango de filas (no toda la hoja)
# ═══════════════════════════════════════════════════════════════════════════

def recorrer_columnas_rango(
    archivo: str,
    hoja: str,
    n: int,
    fila_inicio: int,
    fila_fin: int,
    archivo_salida: str | None = None,
) -> dict[str, Any]:
    """
    Mueve todas las referencias de fórmulas cruzadas N columnas,
    pero solo en las fórmulas dentro del rango de filas dado.

    Diseñada para Resumen_AUM_Afore bloque superior (filas 2-18)
    donde necesitamos recorrer las fórmulas sin afectar otros bloques.

    Solo mueve refs no-absolutas (sin $) en cross-sheet refs.

    Args:
        archivo:       Excel a modificar.
        hoja:          Nombre de la hoja.
        n:             Columnas a mover (+1 = derecha).
        fila_inicio:   Primera fila del rango (inclusiva).
        fila_fin:      Última fila del rango (inclusiva).
        archivo_salida: Dónde guardar. None = sobrescribir.
    """
    CROSS = re.compile(
        r"((?:'[^']+?'|[A-Za-zÁÉÍÓÚáéíóúñÑ0-9_]+)!)"
        r"(\$?)([A-Z]{1,3})(\$?)(\d+)"
    )

    wb = openpyxl.load_workbook(archivo)
    ws = wb[hoja]
    movidas = 0

    for r in range(fila_inicio, fila_fin + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and v.startswith('=')):
                continue

            def _mover(m):
                prefix = m.group(1)
                dol_c = m.group(2)
                col_letter = m.group(3)
                dol_r = m.group(4)
                row_num = m.group(5)
                if dol_c:
                    return m.group(0)
                new_num = column_index_from_string(col_letter) + n
                if new_num < 1:
                    new_num = 1
                new_letter = get_column_letter(new_num)
                return f"{prefix}{dol_c}{new_letter}{dol_r}{row_num}"

            nueva = CROSS.sub(_mover, v)
            if nueva != v:
                ws.cell(r, c).value = nueva
                movidas += 1

    salida = archivo_salida or archivo
    wb.save(salida)
    wb.close()

    print(f"✅ recorrer_columnas_rango({hoja}, filas {fila_inicio}-{fila_fin}, "
          f"n={n:+d}): {movidas} fórmulas movidas → {Path(salida).name}")
    return {"hoja": hoja, "formulas_movidas": movidas}


# ═══════════════════════════════════════════════════════════════════════════
# 9. Reapuntar fórmulas fijas a la nueva última columna
# ═══════════════════════════════════════════════════════════════════════════

def reapuntar_a_ultima_columna(
    archivo: str,
    hoja: str,
    filas: list[int],
    col_referencia_fila: int | None = None,
    col_inicio: int | None = None,
    col_fin: int | None = None,
    archivo_salida: str | None = None,
) -> dict[str, Any]:
    """
    En fórmulas de celdas fijas (no en serie temporal), reescribe las
    referencias para que apunten a la nueva última columna de la hoja.

    Diseñada para bloques tipo "12 month change" y "cambio mensual"
    donde las fórmulas son estáticas (ej. =(Z28-N28)/1000) y al
    agregar un nuevo mes deben apuntar a la nueva última.

    Args:
        archivo:            Excel a modificar.
        hoja:               Nombre de la hoja.
        filas:              Lista de filas con fórmulas a reescribir.
        col_referencia_fila: Fila donde buscar la última columna temporal.
                            None = auto (busca fila con más fechas).
        col_inicio:         Solo procesar fórmulas a partir de esta col.
                            None = col 1.
        col_fin:            Solo procesar fórmulas hasta esta col.
                            None = última col de la hoja.
        archivo_salida:     Dónde guardar. None = sobrescribir.
    """
    wb = openpyxl.load_workbook(archivo)
    ws = wb[hoja]

    # Encontrar la última y penúltima columna temporal
    # Buscar en las primeras 10 filas la fila con más fechas
    if col_referencia_fila is None:
        max_fechas = 0
        for r in range(1, min(80, ws.max_row + 1)):
            n_f = 0
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and (hasattr(v, 'strftime') or
                          (isinstance(v, str) and _is_date_string(v))):
                    n_f += 1
            if n_f > max_fechas:
                max_fechas = n_f
                col_referencia_fila = r

    if col_referencia_fila is None:
        wb.close()
        raise ValueError(f"No se encontró fila de fechas en '{hoja}'")

    # Obtener las dos últimas columnas con fecha
    cols_con_fecha = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(col_referencia_fila, c).value
        if v and (hasattr(v, 'strftime') or
                  (isinstance(v, str) and _is_date_string(v))):
            cols_con_fecha.append(c)

    if len(cols_con_fecha) < 2:
        wb.close()
        return {"hoja": hoja, "formulas_movidas": 0}

    ultima_col = cols_con_fecha[-1]
    penultima_col = cols_con_fecha[-2]
    ultima_letra = get_column_letter(ultima_col)
    penultima_letra = get_column_letter(penultima_col)

    # Reescribir fórmulas: la que apuntaba a penúltima → última,
    # y la que apuntaba a ante-penúltima → penúltima, etc.
    # Patrón: en =(...Z28-N28...), Z es la última (actual), N es 12 antes
    # Después de agregar col AA: debe ser =(...AA28-O28...)
    movidas = 0
    ci = col_inicio or 1
    cf = col_fin or ws.max_column
    for r in filas:
        for c in range(ci, cf + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and v.startswith('=')):
                continue

            nueva = _reescribir_refs_delta(v, delta_col=1, delta_fila=0)
            if nueva != v:
                ws.cell(r, c).value = nueva
                movidas += 1

    salida = archivo_salida or archivo
    wb.save(salida)
    wb.close()

    print(f"✅ reapuntar_a_ultima_columna({hoja}, filas={filas}): "
          f"{movidas} fórmulas reescritas → {Path(salida).name}")
    return {"hoja": hoja, "formulas_movidas": movidas}


# ═══════════════════════════════════════════════════════════════════════════
# 10. Actualizar referencia absoluta de tipo de cambio
# ═══════════════════════════════════════════════════════════════════════════

def actualizar_ref_absoluta(
    archivo: str,
    hoja: str,
    ref_vieja: str,
    ref_nueva: str,
    fila_inicio: int | None = None,
    fila_fin: int | None = None,
    archivo_salida: str | None = None,
) -> dict[str, Any]:
    """
    Reemplaza una referencia absoluta específica en todas las fórmulas
    de la hoja (o de un rango de filas).

    Diseñada para Mandatos USD donde $Z$20 debe cambiar a $AA$20
    cuando se agrega una nueva columna.

    Args:
        archivo:       Excel a modificar.
        hoja:          Nombre de la hoja.
        ref_vieja:     Referencia a reemplazar (ej. "$Z$20").
        ref_nueva:     Nueva referencia (ej. "$AA$20").
        fila_inicio:   Primera fila del rango. None = toda la hoja.
        fila_fin:      Última fila del rango. None = toda la hoja.
        archivo_salida: Dónde guardar. None = sobrescribir.
    """
    wb = openpyxl.load_workbook(archivo)
    ws = wb[hoja]
    movidas = 0

    ri = fila_inicio or 1
    rf = fila_fin or ws.max_row

    for r in range(ri, rf + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and ref_vieja in v):
                continue
            nueva = v.replace(ref_vieja, ref_nueva)
            if nueva != v:
                ws.cell(r, c).value = nueva
                movidas += 1

    salida = archivo_salida or archivo
    wb.save(salida)
    wb.close()

    print(f"✅ actualizar_ref_absoluta({hoja}): '{ref_vieja}' → '{ref_nueva}', "
          f"{movidas} fórmulas → {Path(salida).name}")
    return {"hoja": hoja, "formulas_movidas": movidas}