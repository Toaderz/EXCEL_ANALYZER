"""
anexos_formulas.py
==================
Copia una hoja de Excel con sus fórmulas funcionales y anexa automáticamente
las hojas referenciadas.

APPROACH:
  En vez de reconstruir la hoja desde tablas detectadas (que pierde layout
  cuando hay múltiples tablas), copia la hoja COMPLETA al archivo de salida
  y reescribe las fórmulas cruzadas para que apunten a las hojas anexas.

  Si una fórmula referencia la misma hoja → no cambia nada.
  Si referencia otra hoja que ya está en el workbook de salida → usa el nombre directo.
  Si referencia una hoja que NO existe → la copia como anexo.

FUNCIONES:
  analizar_y_exportar(archivo, hoja, salida)
    → Copia la hoja con fórmulas + anexa hojas referenciadas.

  exportar_todas_con_formulas(archivo, salida)
    → Detecta TODAS las hojas con fórmulas y las exporta juntas
      en un solo workbook, sin duplicar hojas.
"""

import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy

sys.path.insert(0, str(Path(__file__).parent))
from _core import detectar_todas_las_tablas


# ═══════════════════════════════════════════════════════════════════════════
# 1. Detección de hojas referenciadas
# ═══════════════════════════════════════════════════════════════════════════

_CROSS_SHEET_RE = re.compile(
    r"(?:\[(\d+)\])?"
    r"(?:'([^']+)'|([A-Za-zÁÉÍÓÚáéíóúñÑ0-9_]+))"
    r"!"
    r"(\$?[A-Z]{1,3}\$?\d+)"
)


def _hojas_referenciadas(ws) -> set[str]:
    """Extrae nombres de hojas referenciadas en las fórmulas de un worksheet."""
    hojas: set[str] = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (v and isinstance(v, str) and '!' in v):
                continue
            for m in _CROSS_SHEET_RE.finditer(v):
                h = (m.group(2) or m.group(3) or "").strip()
                if h:
                    hojas.add(h)
    return hojas


def _tiene_formulas(ws) -> bool:
    """True si la hoja tiene al menos una fórmula."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and v.startswith('='):
                return True
    return False


# ═══════════════════════════════════════════════════════════════════════════
# 2. Copiar hoja completa
# ═══════════════════════════════════════════════════════════════════════════

def _copiar_hoja(ws_origen, ws_destino):
    """Copia valores, fórmulas y formato básico de una hoja a otra."""
    for r in range(1, ws_origen.max_row + 1):
        for c in range(1, ws_origen.max_column + 1):
            cell_src = ws_origen.cell(r, c)
            cell_dst = ws_destino.cell(r, c)
            cell_dst.value = cell_src.value
            # Copiar formato básico
            if cell_src.has_style:
                cell_dst.font = copy(cell_src.font)
                cell_dst.fill = copy(cell_src.fill)
                cell_dst.alignment = copy(cell_src.alignment)
                cell_dst.number_format = cell_src.number_format

    # Copiar anchos de columna
    for col_letter, dim in ws_origen.column_dimensions.items():
        ws_destino.column_dimensions[col_letter].width = dim.width

    # Copiar merged cells
    for rng in ws_origen.merged_cells.ranges:
        ws_destino.merge_cells(str(rng))


# ═══════════════════════════════════════════════════════════════════════════
# 3. Exportar hoja con anexos
# ═══════════════════════════════════════════════════════════════════════════

def analizar_y_exportar(
    archivo: str,
    hoja: str,
    archivo_salida: str,
) -> dict[str, Any]:
    """
    Copia una hoja con sus fórmulas funcionales y anexa las hojas referenciadas.

    Las fórmulas se mantienen tal cual — apuntan a las hojas originales.
    Las hojas referenciadas se copian al workbook de salida para que
    las fórmulas funcionen.

    Args:
        archivo:         Excel de entrada.
        hoja:            Hoja a copiar.
        archivo_salida:  Excel de salida.

    Returns:
        Dict con "hoja", "anexos", "n_formulas".
    """
    wb_orig = openpyxl.load_workbook(archivo)

    if hoja not in wb_orig.sheetnames:
        raise ValueError(f"Hoja '{hoja}' no existe. Disponibles: {wb_orig.sheetnames}")

    ws_orig = wb_orig[hoja]

    # Detectar hojas referenciadas
    refs = _hojas_referenciadas(ws_orig)
    # Filtrar: solo hojas que existen en el workbook y no son la misma
    refs_validas = {h for h in refs if h in wb_orig.sheetnames and h != hoja}

    # Crear workbook de salida
    wb_out = openpyxl.Workbook()
    ws_main = wb_out.active
    ws_main.title = hoja[:31]

    # Copiar hoja principal
    _copiar_hoja(ws_orig, ws_main)

    # Copiar hojas referenciadas como anexos
    hojas_copiadas: set[str] = {hoja}
    anexos_creados: list[str] = []

    # Resolver recursivamente: una hoja anexa puede referenciar otras
    pendientes = list(refs_validas)
    while pendientes:
        hoja_ref = pendientes.pop(0)
        if hoja_ref in hojas_copiadas:
            continue

        nombre_salida = hoja_ref[:31]
        # Evitar duplicados
        if nombre_salida in [ws.title for ws in wb_out.worksheets]:
            continue

        ws_ref = wb_orig[hoja_ref]
        ws_anexo = wb_out.create_sheet(nombre_salida)
        _copiar_hoja(ws_ref, ws_anexo)
        hojas_copiadas.add(hoja_ref)
        anexos_creados.append(hoja_ref)

        # ¿Esta hoja referencia otras?
        sub_refs = _hojas_referenciadas(ws_ref)
        for sr in sub_refs:
            if sr in wb_orig.sheetnames and sr not in hojas_copiadas:
                pendientes.append(sr)

    wb_orig.close()
    wb_out.save(archivo_salida)

    # Contar fórmulas
    n_formulas = sum(
        1 for r in range(1, ws_main.max_row + 1)
        for c in range(1, ws_main.max_column + 1)
        if ws_main.cell(r, c).value and isinstance(ws_main.cell(r, c).value, str)
        and ws_main.cell(r, c).value.startswith('=')
    )

    print(f"✅ {Path(archivo_salida).name}: {hoja} ({n_formulas} fórmulas) + {len(anexos_creados)} anexos")
    for a in anexos_creados:
        print(f"   {a}")

    return {
        "hoja": hoja,
        "anexos": anexos_creados,
        "n_formulas": n_formulas,
    }


# ═══════════════════════════════════════════════════════════════════════════
# 4. Exportar TODAS las hojas con fórmulas en un solo workbook
# ═══════════════════════════════════════════════════════════════════════════

def exportar_todas_con_formulas(
    archivo: str,
    archivo_salida: str,
) -> dict[str, Any]:
    """
    Detecta TODAS las hojas con fórmulas y las exporta en un solo workbook.

    No duplica hojas: si una hoja ya fue copiada como anexo de otra,
    no se vuelve a copiar.

    Args:
        archivo:         Excel de entrada.
        archivo_salida:  Excel de salida.

    Returns:
        Dict con "hojas_con_formulas", "hojas_copiadas", "total_formulas".
    """
    wb_orig = openpyxl.load_workbook(archivo)
    wb_out = openpyxl.Workbook()
    # Remover la hoja default vacía
    wb_out.remove(wb_out.active)

    hojas_copiadas: set[str] = set()
    hojas_con_formulas: list[str] = []
    total_formulas = 0

    # Primero: identificar hojas con fórmulas
    for nombre in wb_orig.sheetnames:
        ws = wb_orig[nombre]
        if _tiene_formulas(ws):
            hojas_con_formulas.append(nombre)

    # Copiar hojas con fórmulas y sus dependencias
    pendientes = list(hojas_con_formulas)
    while pendientes:
        hoja = pendientes.pop(0)
        if hoja in hojas_copiadas:
            continue
        if hoja not in wb_orig.sheetnames:
            continue

        nombre_salida = hoja[:31]
        if nombre_salida in [ws.title for ws in wb_out.worksheets]:
            continue

        ws_orig = wb_orig[hoja]
        ws_out = wb_out.create_sheet(nombre_salida)
        _copiar_hoja(ws_orig, ws_out)
        hojas_copiadas.add(hoja)

        # Contar fórmulas
        n_f = sum(
            1 for r in range(1, ws_orig.max_row + 1)
            for c in range(1, ws_orig.max_column + 1)
            if ws_orig.cell(r, c).value and isinstance(ws_orig.cell(r, c).value, str)
            and ws_orig.cell(r, c).value.startswith('=')
        )
        total_formulas += n_f

        # Agregar dependencias
        refs = _hojas_referenciadas(ws_orig)
        for ref in refs:
            if ref in wb_orig.sheetnames and ref not in hojas_copiadas:
                pendientes.append(ref)

    # Copiar hojas de datos que son referenciadas pero no tienen fórmulas
    # (ya se copiaron en el while anterior)

    wb_orig.close()
    wb_out.save(archivo_salida)

    print(f"✅ {Path(archivo_salida).name}: {len(hojas_copiadas)} hojas, {total_formulas} fórmulas")
    print(f"   Hojas con fórmulas: {hojas_con_formulas}")
    print(f"   Hojas copiadas (total): {sorted(hojas_copiadas)}")

    return {
        "hojas_con_formulas": hojas_con_formulas,
        "hojas_copiadas": sorted(hojas_copiadas),
        "total_formulas": total_formulas,
    }