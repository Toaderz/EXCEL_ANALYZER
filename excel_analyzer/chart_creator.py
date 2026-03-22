"""
chart_creator.py
================
Funciones para crear gráficas de Excel desde cero con estilos predefinidos.

En vez de intentar preservar o actualizar gráficas existentes, este módulo
las recrea completamente a partir de los datos del archivo, garantizando
que siempre estén sincronizadas con los datos actuales.

ESTILOS DISPONIBLES:
  bar_stacked       Barras apiladas horizontal (ej: AUM por Afore a lo largo del tiempo)
  bar_clustered     Barras agrupadas (ej: AUM snapshot por Afore)
  bar_stacked_100   Barras apiladas al 100% (ej: composición porcentual)
  pie               Pie chart (ej: market share)
  bar_cambio        Barras de cambio mensual con título dinámico

FUNCIONES PRINCIPALES:
  crear_grafica(ws, estilo, datos_config, posicion, **kwargs)
    → Crea una gráfica en la hoja ws con el estilo y datos indicados.

  crear_graficas_desde_config(archivo, config, archivo_salida=None)
    → Crea múltiples gráficas a partir de una lista de configuraciones.
"""

import re
import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    Paragraph, ParagraphProperties, CharacterProperties,
    Font as DrawingFont, RichTextProperties,
)
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.utils import get_column_letter, column_index_from_string


# ═══════════════════════════════════════════════════════════════════════════
# 1. Utilidades internas
# ═══════════════════════════════════════════════════════════════════════════

_MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _buscar_fila_fechas(ws, rango_filas=range(1, 50)) -> tuple[int | None, dict]:
    """Busca la fila con más fechas datetime."""
    mejor_fila = None
    max_fechas = 0
    mejor_map = {}
    for r in rango_filas:
        fechas = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if hasattr(v, 'strftime'):
                fechas[c] = v
        if len(fechas) > max_fechas:
            max_fechas = len(fechas)
            mejor_fila = r
            mejor_map = fechas
    return mejor_fila, mejor_map


def _ultima_col_temporal(ws) -> int | None:
    """Retorna el número de la última columna con fecha."""
    fila, mapa = _buscar_fila_fechas(ws)
    if not mapa:
        return None
    return max(mapa.keys())


def _encontrar_bloque_fechas(mapa_fechas: dict, col_ref: int) -> list[int]:
    """
    Dado un mapa {col: datetime} y una columna de referencia,
    encuentra el bloque contiguo de fechas que incluye esa columna.
    Útil para hojas con múltiples tablas (ej: MXN y USD lado a lado).
    """
    cols = sorted(mapa_fechas.keys())
    if not cols:
        return []

    # Encontrar el bloque donde cae col_ref o el más cercano
    bloques = []
    bloque_actual = [cols[0]]
    for i in range(1, len(cols)):
        if cols[i] - cols[i-1] <= 2:  # tolerancia de 1 col gap
            bloque_actual.append(cols[i])
        else:
            bloques.append(bloque_actual)
            bloque_actual = [cols[i]]
    bloques.append(bloque_actual)

    # Retornar el bloque que contiene col_ref, o el último
    for bloque in bloques:
        if col_ref in bloque or (bloque[0] <= col_ref <= bloque[-1]):
            return bloque
    return bloques[-1] if bloques else []


# ═══════════════════════════════════════════════════════════════════════════
# 2. Estilos de gráfica
# ═══════════════════════════════════════════════════════════════════════════

def _aplicar_estilo_base(chart, estilo: str, **kwargs):
    """Aplica propiedades de estilo base a una gráfica."""

    # Dimensiones
    chart.width = kwargs.get("width", 15)
    chart.height = kwargs.get("height", 7.5)

    if isinstance(chart, PieChart):
        return

    # Configuración según estilo
    if estilo in ("bar_stacked", "bar_stacked_100"):
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.gapWidth = 150
    elif estilo == "bar_clustered":
        chart.grouping = "clustered"
        chart.overlap = -27
        chart.gapWidth = kwargs.get("gapWidth", 219)
    elif estilo == "bar_cambio":
        chart.grouping = "clustered"
        chart.gapWidth = kwargs.get("gapWidth", 182)

    # Ejes
    y_delete = kwargs.get("y_axis_delete", True)
    chart.y_axis.delete = y_delete
    chart.y_axis.tickLblPos = "nextTo"

    x_tick = kwargs.get("x_tickLblPos", "nextTo")
    chart.x_axis.tickLblPos = x_tick

    # Formato numérico del eje Y
    num_fmt = kwargs.get("num_fmt", None)
    if num_fmt:
        chart.y_axis.numFmt = num_fmt

    # Leyenda
    legend_pos = kwargs.get("legend_pos", None)
    if legend_pos:
        chart.legend.position = legend_pos
    elif legend_pos is False:
        chart.legend = None


# ═══════════════════════════════════════════════════════════════════════════
# 3. Función principal: crear gráfica
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class SerieConfig:
    """Configuración de una serie de datos."""
    hoja: str               # Nombre de la hoja de datos
    val_col_start: str       # Columna inicio de valores (letra)
    val_col_end: str | None  # Columna fin (None para una sola columna)
    val_row_start: int       # Fila inicio
    val_row_end: int | None  # Fila fin (None = misma que start)
    cat_col_start: str | None = None   # Columna de categorías
    cat_col_end: str | None = None
    cat_row_start: int | None = None
    cat_row_end: int | None = None
    label_ref: str | None = None       # Referencia para el nombre de la serie


@dataclass
class GraficaConfig:
    """Configuración completa de una gráfica a crear."""
    hoja_destino: str       # Hoja donde poner la gráfica
    estilo: str             # bar_stacked, bar_clustered, pie, etc.
    posicion: str           # Celda ancla (ej: "P2")
    series: list[SerieConfig] = field(default_factory=list)
    titulo: str | None = None
    width: float = 15
    height: float = 7.5
    legend_pos: str | None = "b"    # "b", "r", "t", "l" o None
    y_axis_delete: bool = True
    x_tickLblPos: str = "nextTo"
    num_fmt: str | None = None
    gapWidth: int | None = None


def crear_grafica(
    wb: openpyxl.Workbook,
    config: GraficaConfig,
) -> None:
    """
    Crea una gráfica en el workbook según la configuración dada.

    Args:
        wb:     Workbook abierto (se modifica in-place).
        config: GraficaConfig con todos los parámetros.
    """
    ws = wb[config.hoja_destino]

    # Crear chart del tipo correcto
    if config.estilo == "pie":
        chart = PieChart()
    else:
        chart = BarChart()

    # Aplicar estilo
    kwargs = {
        "width": config.width,
        "height": config.height,
        "y_axis_delete": config.y_axis_delete,
        "x_tickLblPos": config.x_tickLblPos,
        "num_fmt": config.num_fmt,
        "legend_pos": config.legend_pos,
    }
    if config.gapWidth is not None:
        kwargs["gapWidth"] = config.gapWidth
    _aplicar_estilo_base(chart, config.estilo, **kwargs)

    # Título
    if config.titulo:
        chart.title = config.titulo

    # Agregar series
    for sc in config.series:
        # Construir referencia de valores
        if sc.hoja:
            hoja_prefix = f"'{sc.hoja}'!" if ' ' in sc.hoja else f"{sc.hoja}!"
        else:
            hoja_prefix = ""

        val_end_col = sc.val_col_end or sc.val_col_start
        val_end_row = sc.val_row_end or sc.val_row_start
        val_ref = f"{hoja_prefix}${sc.val_col_start}${sc.val_row_start}:${val_end_col}${val_end_row}"

        # Referencia de categorías
        cat_ref = None
        if sc.cat_col_start and sc.cat_row_start:
            cat_end_col = sc.cat_col_end or sc.cat_col_start
            cat_end_row = sc.cat_row_end or sc.cat_row_start
            cat_ref = f"{hoja_prefix}${sc.cat_col_start}${sc.cat_row_start}:${cat_end_col}${cat_end_row}"

        # Crear serie usando openpyxl objects directamente
        from openpyxl.chart.series import Series
        from openpyxl.chart.data_source import (
            NumDataSource, NumRef, StrRef, AxDataSource,
        )

        vals = NumDataSource(numRef=NumRef(f=val_ref))

        cats = None
        if cat_ref:
            cats = AxDataSource(strRef=StrRef(f=cat_ref))

        serie = Series(val=vals, cat=cats)

        # Label de la serie
        if sc.label_ref:
            serie.title = SeriesLabel(
                strRef=openpyxl.chart.data_source.StrRef(f=sc.label_ref)
            )

        chart.series.append(serie)

    # Poner en la hoja
    ws.add_chart(chart, config.posicion)


def crear_graficas_desde_config(
    archivo: str,
    configs: list[GraficaConfig],
    archivo_salida: str | None = None,
    verbose: bool = True,
) -> dict[str, Any]:
    """
    Crea múltiples gráficas en un archivo Excel.

    Args:
        archivo:         Excel de entrada.
        configs:         Lista de GraficaConfig.
        archivo_salida:  Excel de salida. None = sobrescribir.
        verbose:         Imprimir progreso.

    Returns:
        {"graficas_creadas": N, "archivo_salida": str}
    """
    salida = archivo_salida or archivo
    wb = openpyxl.load_workbook(archivo)

    for i, cfg in enumerate(configs):
        if cfg.hoja_destino not in wb.sheetnames:
            if verbose:
                print(f"  ⚠ Hoja '{cfg.hoja_destino}' no existe, omitiendo Ch{i}")
            continue

        crear_grafica(wb, cfg)

        if verbose:
            titulo_str = cfg.titulo or "Sin título"
            print(f"  ✅ [{cfg.hoja_destino}] {cfg.estilo}: {titulo_str} ({len(cfg.series)} series)")

    wb.save(salida)
    wb.close()

    if verbose:
        print(f"\n  Guardado en: {Path(salida).name}")

    return {"graficas_creadas": len(configs), "archivo_salida": salida}


# ═══════════════════════════════════════════════════════════════════════════
# 4. Helpers para construir configs dinámicamente
# ═══════════════════════════════════════════════════════════════════════════

def ventana_temporal_series(
    hoja_datos: str,
    fila_header: int,
    filas_datos: list[int],
    fila_labels_col: str,
    col_fin: int,
    ventana: int = 13,
) -> tuple[list[SerieConfig], str, str]:
    """
    Construye series para una gráfica de ventana temporal (últimos N meses).

    Args:
        hoja_datos:    Nombre de la hoja con datos.
        fila_header:   Fila donde están las fechas (categorías).
        filas_datos:   Lista de filas con datos (una serie por fila).
        fila_labels_col: Columna con los nombres de cada serie (ej: "B").
        col_fin:       Última columna con datos (número).
        ventana:       Ancho de la ventana (default 13 meses).

    Returns:
        (lista de SerieConfig, col_inicio_letra, col_fin_letra)
    """
    col_inicio = col_fin - ventana + 1
    if col_inicio < 1:
        col_inicio = 1

    col_s = get_column_letter(col_inicio)
    col_e = get_column_letter(col_fin)

    series = []
    for fila in filas_datos:
        series.append(SerieConfig(
            hoja=hoja_datos,
            val_col_start=col_s, val_col_end=col_e,
            val_row_start=fila, val_row_end=fila,
            cat_col_start=col_s, cat_col_end=col_e,
            cat_row_start=fila_header, cat_row_end=fila_header,
            label_ref=f"'{hoja_datos}'!${fila_labels_col}${fila}" if ' ' in hoja_datos
                      else f"{hoja_datos}!${fila_labels_col}${fila}",
        ))

    return series, col_s, col_e


def snapshot_series(
    hoja_datos: str,
    col_valores: str,
    fila_inicio: int,
    fila_fin: int,
    col_categorias: str | None = None,
) -> list[SerieConfig]:
    """
    Construye una serie vertical (snapshot de una columna).

    Args:
        hoja_datos:     Nombre de la hoja.
        col_valores:    Columna con los valores (letra).
        fila_inicio:    Primera fila de datos.
        fila_fin:       Última fila de datos.
        col_categorias: Columna con las etiquetas (letra). None = sin categorías.
    """
    return [SerieConfig(
        hoja=hoja_datos,
        val_col_start=col_valores, val_col_end=col_valores,
        val_row_start=fila_inicio, val_row_end=fila_fin,
        cat_col_start=col_categorias, cat_col_end=col_categorias,
        cat_row_start=fila_inicio if col_categorias else None,
        cat_row_end=fila_fin if col_categorias else None,
    )]