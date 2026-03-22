"""
excel_analyzer.py  v6
=====================
Librería para detectar y extraer tablas de archivos Excel desordenados
con merged cells, filas/columnas vacías y bloques de datos no contiguos.

ARQUITECTURA
============

    WorkbookLoader        — IO puro: abre el workbook, entrega (ws, merged_ranges)
         ↓
    SheetScanner          — matrix numpy + masks cacheadas (una sola lectura del disco)
         ↓
    RegionDetector        — detección de regiones O(n) puro numpy (np.diff, sin sets)
         ↓
    TableParser           — convierte regiones en TableRegion (dataclass tipada)
         ↓
    TableAnalyzer         — orquesta, cachea resultados, expone API interna
         ↓
    API pública           — detectar_tabla(), extraer_columna(), etc. (compatible v3+)

HISTORIAL DE CAMBIOS
====================

  v3 → v4
    • Un solo escaneo de la hoja (ws.values → numpy matrix)
    • Acceso O(1) a celdas vía matrix en lugar de ws.cell()
    • Merged cells expandidas directamente en la matrix
    • WorkbookLoader separado del scanner (IO desacoplado)
    • RegionDetector: _agrupar con np.diff en lugar de set+sort (15x más rápido)
    • RegionDetector.detectar: ndarray de np.where directamente (sin conversión a set,
      elimina 33x overhead)
    • Detección de orientación con umbral absoluto MIN_DATES_FOR_ORIENTATION
    • TableRegion como @dataclass tipada en lugar de dict genérico
    • score = n_filas × n_fechas en lugar de solo n_fechas
    • TableAnalyzer.from_scanner() para reutilizar scanner sin reabrir archivo
    • data: pd.DataFrame | None para lazy parsing futuro
    • ws.iter_rows(values_only=True) en lugar de ws.values (API documentada y explícita)

  v4 → v5  (sin cambio de comportamiento visible, solo performance interno)
    • _build_meaningful_mask y _build_datetime_mask: eliminan np.vectorize sobre
      toda la matrix. Aplican la operación solo sobre celdas no-None (≈25% del total):
        1. (mat != None) → C-level sobre todo el array (rápido)
        2. frompyfunc     → solo sobre el subconjunto no-None
      Resultado medido: 1.3–1.4x más rápido en hojas de 5000×100.

  v5 → v6  (sin cambio de comportamiento visible, solo performance interno)
    • _build_masks(): fusiona las dos máscaras (meaningful + datetime) en UN SOLO
      recorrido de los valores no-None:
        1. np.frompyfunc(type, 1, 1)  → obtiene tipos en batch (una sola llamada)
        2. is_str  = (types == str)   → selecciona strings sin bucle Python
        3. np.char.strip / startswith → operaciones C-level solo sobre strings
        4. is_dt   = (types == datetime.datetime) → máscara datetime directo
      Resultado medido: 2.0x más rápido que v5 en hojas de 5000×100
      (43 ms vs 88 ms para 500k celdas no-None).

NOTAS TÉCNICAS
==============

  • np.frompyfunc(type, 1, 1): aunque frompyfunc sigue siendo un loop Python,
    llamar type() es mucho más barato que str() (no construye un nuevo objeto).
    Extraer los tipos en batch permite derivar AMBAS máscaras con un solo recorrido.

  • np.char.strip / np.char.startswith: operan sobre arrays de dtype str (no object),
    ejecutando las comparaciones en C. Solo se aplican a la fracción de celdas
    que son strings (≈17% en hojas financieras típicas).

  • dtype=object es inevitable para arrays con tipos mixtos (str, datetime, float, None).
    Para hojas normales (5000×100) ≈ 3.8 MB de memoria — aceptable.

  • _MESES_ES vs calendar.month_abbr: calendar da nombres en inglés ('Jan', 'Feb'…).
    Para reportes financieros en español se requiere la lista explícita.
    Rendimiento medido: 0.6 μs vs 2.7 μs por llamada — _MESES_ES también es más rápido.

  • score = n_filas × n_fechas es correcto en todos los casos:
      - fila_header is None → clave_valor → n_fechas=0 → score=0 (correcto)
      - fila_header == fila_inicio → n_filas = fila_fin - fila_header (datos puros)
      - max(0, ...) evita scores negativos si la región es solo el header

  • _bbox_cols no requiere caché: con numpy la operación tarda <0.01ms por región.
    El overhead de gestionar la caché superaría el beneficio.
"""

from __future__ import annotations

import datetime
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

try:
    from ._region_detector import ProjectionRegionDetector
except ImportError:  # pragma: no cover - soporte ejecución directa del módulo
    from _region_detector import ProjectionRegionDetector  # type: ignore


# ══════════════════════════════════════════════════════════════════════════════
# TIPOS Y CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════

TableType   = Literal[
    "manager_metrica", "solo_metrica", "clave_valor",
    "tabla_generica",  "tabla_rotada", "tabla_cruzada",
]
Orientation = Literal["column", "row", "ambiguous"]

# Tipos de valor en un header (eje de filas o columnas)
HeaderTipo  = Literal["fecha", "anio", "periodo", "texto", "numero", "grupo"]

# Etiquetas de meses en español para la salida canónica "mmm-aa".
_MESES_ES = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]

# Mínimo de fechas en una fila/columna para confiar en la detección de orientación.
_MIN_DATES_FOR_ORIENTATION = 3

# ── Diccionarios de meses para detección y normalización ─────────────────
_MESES_ES_SHORT: dict[str, int] = {m: i+1 for i, m in enumerate(_MESES_ES)}
_MESES_EN_SHORT: dict[str, int] = {
    "jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
    "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12,
}
_MESES_ES_FULL: dict[str, int] = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12,
}
_MESES_EN_FULL: dict[str, int] = {
    "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
}
# Variantes adicionales encontradas en reportes financieros mexicanos:
#   "mzo" = marzo, "ago."/"sep." = abreviatura con punto
_MESES_VARIANTES: dict[str, int] = {
    "mzo":3, "ene.":1, "feb.":2, "mar.":3, "abr.":4, "may.":5, "jun.":6,
    "jul.":7, "ago.":8, "sep.":9, "oct.":10, "nov.":11, "dic.":12,
}
_ALL_MONTH_NAMES: dict[str, int] = {
    **_MESES_ES_SHORT, **_MESES_EN_SHORT,
    **_MESES_ES_FULL,  **_MESES_EN_FULL,
    **_MESES_VARIANTES,
}

# ── Regex universal de fechas-string ─────────────────────────────────────
# Aplica DESPUÉS de normalizar: strip() + lower() + colapsar espacios.
#
# Formatos soportados:
#   Numérico   : 2024-01  2024/01  01/2024  01-2024  2024-01-01  01/01/2024  1/1/2024
#   Abreviado  : ene-24  jan-24  ene- 24  ENE-24  ene 24  ene-2024  Jan-2024
#   Completo   : enero 2024  january 2024
#   Trimestre  : Q1-2024  Q1 2024  Q12024
#
# Garantiza NO-match en:
#   "AUM Marzo 2024"  → descartado por longitud > 25 ó guarda semántica
#   "roa-24"          → "roa" no es un mes conocido → rechazado
_DATE_STRING_RE = re.compile(
    r"^\d{4}[\-/\.]\d{1,2}$"              # 2024-01  2024/01  2024.01  2024-1
    r"|^\d{1,2}[\-/\.]\d{4}$"             # 01/2024  01-2024  01.2024  1/2024
    r"|^\d{1,2}[\-/\.]\d{2}$"             # 01-24  01/24  01.24  1-24  (mm-yy)
    r"|^\d{1,2}/\d{1,2}/\d{4}$"           # 01/01/2024  1/1/2024
    r"|^\d{4}-\d{2}-\d{2}$"               # 2024-01-01
    r"|^[a-z]{3}[\s\-/]+\s*\d{2}$"        # ene-24  ene/24  ene 24  ene- 24
    r"|^[a-z]{3}[\s\-/]+\s*\d{4}$"        # ene-2024  jan/2024
    r"|^[a-z]{3,9}[\s\-]+\d{2}$"          # enero-24  ENERO-24 (normalizado)
    r"|^[a-z]{3,9}\s+\d{4}$"              # enero 2024  january 2024
    r"|^q[1-4][\s\-]?\d{4}$"              # Q1-2024  Q1 2024  Q12024
    r"|^[a-z]{3,9}\.?$"                    # ene  Ene  Mzo  Ago.  (mes solo, sin año)
)
_WORD_ONLY_RE = re.compile(r"^([a-z]+)")


def _normalize_date_str(s: str) -> str:
    """Strip, lowercase y colapsa espacios múltiples."""
    return re.sub(r"\s+", " ", s.strip().lower())


def _is_date_string(s: str) -> bool:
    """
    True si la cadena es una etiqueta de período o fecha.

    Tres capas de validación:
      1. Longitud: strings > 25 chars nunca son fechas (evita métricas largas).
      2. Regex estructural: descarta strings sin forma de fecha.
      3. Guardas semánticas:
         a. Si empieza con letras (no dígitos, no Q), la palabra inicial
            debe ser un nombre de mes conocido (rechaza "roa-24", "aum 2024").
         b. Si es formato numérico mm-yy / mm-yyyy, el mes debe ser 1-12
            (rechaza "13-24", "99-24", "45/2024").
    """
    n = _normalize_date_str(s)
    if len(n) > 25:
        return False
    if not _DATE_STRING_RE.match(n):
        return False
    # Guarda 3a: validar nombre de mes si el string empieza con letras
    if n[0].isalpha() and not n.startswith("q"):
        m = _WORD_ONLY_RE.match(n)
        if m and m.group(1) not in _ALL_MONTH_NAMES:
            return False
    # Guarda 3b: si es mm-yy o mm-yyyy numérico, mes debe ser 1-12
    pure_num = re.match(r"^(\d{1,2})[\-/\.](\d{2,4})$", n)
    if pure_num:
        month = int(pure_num.group(1))
        if month < 1 or month > 12:
            return False
    return True


def _make_unique(cols: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for c in cols:
        if c not in seen:
            seen[c] = 0
            out.append(c)
        else:
            seen[c] += 1
            out.append(f"{c}_{seen[c]}")
    return out


# Versión vectorizada para _build_masks() (aplica sobre arrays de strings numpy)
_looks_like_date_vec = np.frompyfunc(_is_date_string, 1, 1)


def _month_name_only(s: str) -> int | None:
    """
    Si `s` es un nombre de mes solo (sin año), devuelve el número de mes (1-12).
    Si no es un nombre de mes solo, devuelve None.

    Ejemplos:
        "Ene"  → 1
        "Mzo"  → 3
        "Ago." → 8
        "ene-24" → None  (tiene año, no es "solo")
        "INGRESOS" → None
    """
    n = _normalize_date_str(s).rstrip(".")
    # Solo letras, longitud 2-9, sin dígitos ni separadores
    if re.match(r"^[a-z]{2,9}$", n) and n in _ALL_MONTH_NAMES:
        return _ALL_MONTH_NAMES[n]
    # Con punto al final: "ago." → "ago"
    n2 = _normalize_date_str(s)
    if re.match(r"^[a-z]{2,9}\.$", n2):
        base = n2[:-1]
        if base in _ALL_MONTH_NAMES:
            return _ALL_MONTH_NAMES[base]
    return None


# ══════════════════════════════════════════════════════════════════════════════
# 1. WorkbookLoader  — IO puro, separado del scanner
# ══════════════════════════════════════════════════════════════════════════════

class WorkbookLoader:
    """
    Responsabilidad única: abrir un .xlsx y entregar (ws, merged_ranges).

    Separar el IO del scanner permite:
      - Testear SheetScanner con datos sintéticos sin tocar disco.
      - Reutilizar un scanner construido para analizar el mismo archivo
        varias veces sin reabrirlo (TableAnalyzer.from_scanner).
      - Cambiar el backend de lectura (xlrd, calamine…) sin tocar el pipeline.
    """

    @staticmethod
    def load(archivo: str | Path, hoja: str) -> tuple[Any, list[Any]]:
        """
        Abre el archivo y devuelve (ws, merged_ranges).

        Args:
            archivo: Ruta al .xlsx.
            hoja   : Nombre de la hoja.

        Returns:
            (worksheet, list[MergedCellRange])

        Raises:
            FileNotFoundError si el archivo no existe.
            ValueError si la hoja no existe en el workbook.
        """
        path = Path(archivo)
        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {path}")
        wb = openpyxl.load_workbook(str(path), data_only=True)
        if hoja not in wb.sheetnames:
            raise ValueError(
                f"Hoja '{hoja}' no existe. Disponibles: {wb.sheetnames}"
            )
        ws            = wb[hoja]
        merged_ranges = list(ws.merged_cells.ranges)
        return ws, merged_ranges


# ══════════════════════════════════════════════════════════════════════════════
# 2. SheetScanner  — matrix numpy + masks cacheadas
# ══════════════════════════════════════════════════════════════════════════════

class SheetScanner:
    """
    Carga el worksheet en memoria UNA sola vez como numpy object array (base-0)
    y expone todas las operaciones de consulta como accesos O(1) o vectorizados.

    Construcción:
        ws, merged_ranges = WorkbookLoader.load(archivo, hoja)
        scanner = SheetScanner(ws, merged_ranges)

    Acceso O(1):
        scanner[r, c]      — base-0
        scanner.get1(r, c) — base-1 (estilo Excel)

    Masks precalculadas (calculadas UNA vez en __init__, luego O(1)):
        scanner._mask_meaningful  — bool array, True = celda con dato válido
        scanner._mask_datetime    — bool array, True = celda con datetime
        scanner.row_density       — int array (nrows,), n celdas con datos por fila
        scanner.col_density       — int array (ncols,), n celdas con datos por columna
        scanner.orientation       — "column" | "row" | "ambiguous"
    """

    def __init__(self, ws: Any, merged_ranges: list[Any]) -> None:
        self.hoja  = getattr(ws, "title", "")
        self.nrows = ws.max_row
        self.ncols = ws.max_column

        # ── Cargar en numpy array (una sola iteración de la hoja) ─────────
        # ws.iter_rows(values_only=True) es la API documentada y explícita
        # de openpyxl y garantiza filas rectangulares (padding con None).
        self._mat: np.ndarray = np.empty((self.nrows, self.ncols), dtype=object)
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            for ci, v in enumerate(row):
                self._mat[ri, ci] = v

        # ── Expandir merged cells directamente en la matrix ───────────────
        # Slicing 2D de numpy sobrescribe el valor en toda la región del merge.
        # Evita el dict auxiliar {(r,c): val} de versiones anteriores.
        #
        # Adicionalmente construimos _merge_map: dict[celda_b0 → bbox_b0]
        # para consultar O(1) si una celda pertenece a un merge y cuáles son
        # sus límites. Usado en TableParser._jerarquia_por_merges() para
        # detectar grupos multi-columna en los headers.
        # _merge_list guarda un registro por rango (no por celda).
        # Para un merge de 100×20 celdas, esto es 1 entrada vs 2000 anterior.
        # La búsqueda O(n_merges) es acceptable: típico < 200 merges por hoja.
        self._merge_list: list[tuple[int,int,int,int]] = []

        for rng in merged_ranges:
            r0m = rng.min_row - 1; r1m = rng.max_row - 1
            c0m = rng.min_col - 1; c1m = rng.max_col - 1
            val = self._mat[r0m, c0m]
            self._mat[r0m : r1m + 1, c0m : c1m + 1] = val
            self._merge_list.append((r0m, r1m, c0m, c1m))

        # Compatibilidad: _merge_map como propiedad lazy (crea el dict al primer uso)
        self.__merge_map_cache: dict[tuple[int,int], tuple[int,int,int,int]] | None = None

        # ── Construir las tres masks en UN solo recorrido de celdas ─────
        self._mask_meaningful, self._mask_datetime, self._mask_numeric = self._build_masks()

        # Densidades: sum() de bool array es C-level sobre numpy
        self.row_density: np.ndarray = self._mask_meaningful.sum(axis=1)
        self.col_density: np.ndarray = self._mask_meaningful.sum(axis=0)

        # Orientación dominante de las fechas
        self.orientation: Orientation = self._detect_orientation()

    # ── Acceso a celdas ────────────────────────────────────────────────────

    def __getitem__(self, idx: tuple[int, int]) -> Any:
        """scanner[r, c] — índice base-0."""
        return self._mat[idx]

    def get(self, r: int, c: int) -> Any:
        """Valor de (r, c), base-0."""
        return self._mat[r, c]

    def get1(self, r: int, c: int) -> Any:
        """Valor de (r, c), base-1 (estilo Excel)."""
        return self._mat[r - 1, c - 1]

    def slice1(self, r0: int, r1: int, c0: int, c1: int) -> np.ndarray:
        """Sub-matrix base-1 [r0..r1]×[c0..c1] (inclusive, estilo Excel)."""
        return self._mat[r0 - 1 : r1, c0 - 1 : c1]

    # ── Masks de sub-rango ────────────────────────────────────────────────

    def get_merge_bbox(self, r: int, c: int) -> tuple[int,int,int,int] | None:
        """
        Retorna el bounding box (r0,r1,c0,c1) base-0 del merge que contiene
        la celda (r,c), o None si la celda no pertenece a ningún merge.
        Búsqueda O(n_merges): típico < 200 merges → negligible.
        """
        for (r0m, r1m, c0m, c1m) in self._merge_list:
            if r0m <= r <= r1m and c0m <= c <= c1m:
                return (r0m, r1m, c0m, c1m)
        return None

    @property
    def _merge_map(self) -> dict[tuple[int,int], tuple[int,int,int,int]]:
        """
        Propiedad de compatibilidad: construye el dict completo lazily.
        Solo se materializa si código externo accede directamente a _merge_map
        (ej. tests). El uso interno debe preferir get_merge_bbox().
        """
        if self.__merge_map_cache is None:
            cache: dict[tuple[int,int], tuple[int,int,int,int]] = {}
            for (r0m, r1m, c0m, c1m) in self._merge_list:
                bbox = (r0m, r1m, c0m, c1m)
                for ri in range(r0m, r1m + 1):
                    for ci in range(c0m, c1m + 1):
                        cache[(ri, ci)] = bbox
            self.__merge_map_cache = cache
        return self.__merge_map_cache

    def meaningful_mask(
        self,
        r0: int = 0, r1: int | None = None,
        c0: int = 0, c1: int | None = None,
    ) -> np.ndarray:
        """Sub-mask de celdas con dato válido, rango base-0 [r0:r1, c0:c1]."""
        return self._mask_meaningful[
            r0 : (r1 or self.nrows),
            c0 : (c1 or self.ncols),
        ]

    def date_mask(
        self,
        r0: int = 0, r1: int | None = None,
        c0: int = 0, c1: int | None = None,
    ) -> np.ndarray:
        """Sub-mask de celdas con datetime, rango base-0 [r0:r1, c0:c1]."""
        return self._mask_datetime[
            r0 : (r1 or self.nrows),
            c0 : (c1 or self.ncols),
        ]

    # ── Core: construir ambas masks en un recorrido ───────────────────────

    def _build_masks(self) -> tuple[np.ndarray, np.ndarray]:
        """
        Construye _mask_meaningful y _mask_datetime en un único recorrido
        de los valores no-None.

        Estrategia (medida en benchmark, 5000×100 = 500k celdas):

          Paso 1: (mat != None)
            Comparación C-level sobre todo el array. Rápida e inevitable.
            Identifica el ≈25% de celdas que tienen algún valor.

          Paso 2: np.frompyfunc(type, 1, 1)(vals)
            Obtiene el tipo Python de CADA celda no-None en UN solo recorrido.
            Llamar type() es mucho más barato que str() porque no construye
            un nuevo objeto — solo devuelve el puntero de tipo ya existente.

          Paso 3a: meaningful  
            - Celdas no-string (datetime, int, float…) → siempre True
            - Strings: np.char.strip + np.char.startswith (C-level, ≈17% del total)

          Paso 3b: datetime
            - (types == datetime.datetime) → comparación de punteros, O(n) C-level

          Resultado medido: 2.0x más rápido que v5 para hojas de 5000×100
          (43 ms vs 88 ms al construir ambas masks).
        """
        # ── Paso 1: celdas no-None ────────────────────────────────────────
        not_none = self._mat != None    # noqa: E711  C-level
        pos      = np.where(not_none)
        vals     = self._mat[pos]

        mask_m  = np.zeros((self.nrows, self.ncols), dtype=bool)
        mask_dt = np.zeros((self.nrows, self.ncols), dtype=bool)

        if len(vals) == 0:
            return mask_m, mask_dt

        # ── Paso 2: tipos en batch (un solo recorrido) ───────────────────
        types = np.frompyfunc(type, 1, 1)(vals)   # object array of Python types

        # ── Paso 3a: meaningful mask ──────────────────────────────────────
        is_str = (types == str)                    # selección de strings

        # Por defecto toda celda no-None es meaningful (datetime, int, float…)
        meaningful = np.ones(len(vals), dtype=bool)

        if is_str.any():
            # np.char requiere array de dtype str (no object)
            # Solo procesamos el subconjunto de strings (≈17% típico)
            str_vals    = vals[is_str].astype(str)
            str_stripped = np.char.strip(str_vals)
            bad          = (str_stripped == "") | np.char.startswith(str_stripped, "#")
            meaningful[is_str] = ~bad

        # ── Paso 3b: datetime mask ────────────────────────────────────────
        # Incluye celdas datetime.datetime Y strings con aspecto de fecha
        # (ej. "ene-24", "Jan-24", "2024-01", "03/2024").
        # Esto permite detectar tablas donde Excel guardó las fechas como texto.
        is_dt = (types == datetime.datetime)       # puntero de tipo, sin isinstance

        if is_str.any():
            # _looks_like_date_vec ya opera sobre str_stripped (dtype=str)
            # Solo se necesita si hay strings — condición del bloque exterior
            str_stripped_for_dt = np.char.strip(vals[is_str].astype(str))
            is_dt_str            = _looks_like_date_vec(str_stripped_for_dt).astype(bool)
            is_dt                = is_dt.copy()
            is_dt[is_str]        = is_dt_str

        # ── Paso 3c: numeric mask ────────────────────────────────────────
        # True = celda con valor numérico real (int o float, excluye bool).
        # np.frompyfunc(type) ya está calculado — reutilizamos types.
        # is_bool: separamos bool ANTES de is_num porque bool es subclase de int.
        is_bool = (types == bool)
        is_num  = ((types == int) | (types == float)) & ~is_bool

        mask_num = np.zeros((self.nrows, self.ncols), dtype=bool)

        # ── Escribir en arrays resultado ──────────────────────────────────
        mask_m[pos]   = meaningful
        mask_dt[pos]  = is_dt.astype(bool)
        mask_num[pos] = is_num

        return mask_m, mask_dt, mask_num

    # ── Orientación ───────────────────────────────────────────────────────

    def _detect_orientation(self) -> Orientation:
        """
        Determina si las fechas son encabezados de columna ("column") o
        etiquetas de fila ("row").

        Lógica:
          - Si ninguna dimensión acumula ≥ _MIN_DATES_FOR_ORIENTATION fechas
            en una sola fila/columna → "ambiguous" (evita falsos positivos con
            hojas sin fechas o con muy pocas).
          - ratio = max_fechas_en_alguna_fila / max_fechas_en_alguna_col
            ratio ≥ 1.5 → "column"
            ratio ≤ 0.67 → "row"
            en medio → "ambiguous"
        """
        row_dates = int(self._mask_datetime.sum(axis=1).max())
        col_dates = int(self._mask_datetime.sum(axis=0).max())

        if max(row_dates, col_dates) < _MIN_DATES_FOR_ORIENTATION:
            return "ambiguous"

        ratio = row_dates / max(col_dates, 1)
        if ratio >= 1.5:
            return "column"
        if ratio <= 0.67:
            return "row"
        return "ambiguous"

    # ── Utilidades ────────────────────────────────────────────────────────

    @staticmethod
    def _is_cell_meaningful(v: Any) -> bool:
        """
        True si v tiene un dato válido (no None, no vacío, no error Excel).
        Llamado en el loop de parseo de filas, no en la construcción de masks.
        Usa type() en lugar de isinstance() para evitar la herencia y ser
        consistente con _build_masks.
        """
        if v is None:
            return False
        t = type(v)
        if t is str:
            s = v.strip()
            return bool(s) and s[0] != "#"
        return True   # datetime, int, float, bool → siempre significativo

    @staticmethod
    def fmt_fecha(v: Any) -> str:
        """
        Convierte CUALQUIER representación de fecha a la etiqueta canónica 'mmm-aa'.

        Tipos y formatos soportados:
          datetime / date            → vía .month/.year
          "2024-01-01"  "01/01/2024" → extrae mes/año
          "2024-01"  "2024.01"  "2024-1"    → año-mes (cualquier sep)
          "01/2024"  "01-2024"  "01.2024"   → mes-año (cualquier sep)
          "01-24"  "01/24"  "1-24"          → mm-yy (sep + 2 dígitos)
          "ene-24"  "ene/24"  "ene- 24"     → abreviatura ES/EN + 2 dígitos
          "ene-2024"  "Jan/2024"             → abreviatura + 4 dígitos
          "ENERO-24"  "enero-24"             → mes completo + guión + 2 dígitos
          "enero 2024"  "january 2024"       → mes completo + año
          "Q1-2024"  "Q1 2024"  "Q12024"    → trimestre → primer mes
          Cualquier otro valor              → str(v).strip() (pass-through)
        """
        # ── datetime / date nativo ───────────────────────────────────────
        if isinstance(v, (datetime.datetime, datetime.date)):
            return f"{_MESES_ES[v.month - 1]}-{str(v.year)[2:]}"
        if v is None:
            return ""

        n = _normalize_date_str(str(v))

        # ── Trimestre Q1–Q4 ─────────────────────────────────────────────
        qm = re.match(r"^q([1-4])[\s\-]?(\d{4})$", n)
        if qm:
            month = (int(qm.group(1)) - 1) * 3 + 1
            return f"{_MESES_ES[month-1]}-{qm.group(2)[2:]}"

        # ── Fecha completa dd/mm/yyyy ────────────────────────────────────
        dm = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$", n)
        if dm:
            mo, yr = int(dm.group(2)), int(dm.group(3))
            if 1 <= mo <= 12:
                return f"{_MESES_ES[mo-1]}-{str(yr)[2:]}"

        # ── ISO 2024-01-01 ───────────────────────────────────────────────
        iso = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", n)
        if iso:
            yr, mo = int(iso.group(1)), int(iso.group(2))
            if 1 <= mo <= 12:
                return f"{_MESES_ES[mo-1]}-{str(yr)[2:]}"

        # ── yyyy-mm  yyyy/mm  yyyy.mm  yyyy-m ───────────────────────────
        ym = re.match(r"^(\d{4})[\-/\.](\d{1,2})$", n)
        if ym:
            yr, mo = int(ym.group(1)), int(ym.group(2))
            if 1 <= mo <= 12:
                return f"{_MESES_ES[mo-1]}-{str(yr)[2:]}"

        # ── mm/yyyy  mm-yyyy  mm.yyyy  m/yyyy ───────────────────────────
        my4 = re.match(r"^(\d{1,2})[\-/\.](\d{4})$", n)
        if my4:
            mo, yr = int(my4.group(1)), int(my4.group(2))
            if 1 <= mo <= 12:
                return f"{_MESES_ES[mo-1]}-{str(yr)[2:]}"

        # ── mm-yy  mm/yy  mm.yy  m-yy  (2-digit year) ──────────────────
        my2 = re.match(r"^(\d{1,2})[\-/\.](\d{2})$", n)
        if my2:
            mo, yr2 = int(my2.group(1)), my2.group(2)
            if 1 <= mo <= 12:
                return f"{_MESES_ES[mo-1]}-{yr2}"

        # ── ene-24  ene/24  ene- 24  (3-letter + 2-digit year) ──────────
        sm2 = re.match(r"^([a-z]{3})[\s\-/]+\s*(\d{2})$", n)
        if sm2:
            w, yr2 = sm2.group(1), sm2.group(2)
            if w in _ALL_MONTH_NAMES:
                return f"{_MESES_ES[_ALL_MONTH_NAMES[w]-1]}-{yr2}"

        # ── ene-2024  jan/2024  (3-letter + 4-digit year) ───────────────
        sm4 = re.match(r"^([a-z]{3})[\s\-/]+\s*(\d{4})$", n)
        if sm4:
            w, yr = sm4.group(1), int(sm4.group(2))
            if w in _ALL_MONTH_NAMES:
                return f"{_MESES_ES[_ALL_MONTH_NAMES[w]-1]}-{str(yr)[2:]}"

        # ── enero-24  ENERO-24  (full month + dash + 2-digit year) ──────
        fm2 = re.match(r"^([a-z]{4,9})[\s\-]+(\d{2})$", n)
        if fm2:
            w, yr2 = fm2.group(1), fm2.group(2)
            if w in _ALL_MONTH_NAMES:
                return f"{_MESES_ES[_ALL_MONTH_NAMES[w]-1]}-{yr2}"

        # ── enero 2024  january 2024  (full month + 4-digit year) ───────
        fm4 = re.match(r"^([a-z]{4,9})\s+(\d{4})$", n)
        if fm4:
            w, yr = fm4.group(1), int(fm4.group(2))
            if w in _ALL_MONTH_NAMES:
                return f"{_MESES_ES[_ALL_MONTH_NAMES[w]-1]}-{str(yr)[2:]}"

        # ── Fallback ─────────────────────────────────────────────────────
        return str(v).strip()


# ══════════════════════════════════════════════════════════════════════════════
# 3. TableRegion  — dataclass tipada para una región detectada
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class ValorHeader:
    """
    Describe un valor individual en el header de un eje (filas o columnas).

    label  : Texto tal como aparece en el Excel (normalizado)
    tipo   : "fecha" | "anio" | "periodo" | "grupo" | "texto" | "numero"
    nivel  : 0 = grupo más externo; nivel máximo = hoja (métrica/fecha)
    anio   : Año inferido del contexto (grupo padre o marcador de año)
    fecha  : date parseada, solo si tipo == "fecha"
    """
    label : str
    tipo  : HeaderTipo
    nivel : int                   = 0
    anio  : int | None            = None
    fecha : datetime.date | None  = None


@dataclass
class EjeInfo:
    """
    Describe el eje de filas o el eje de columnas de una tabla.

    valores          : Lista de ValorHeader en orden de aparición
    n_niveles_grupo  : Cuántos niveles de agrupación hay antes de la hoja
    tipo_hoja        : Tipo dominante de los valores de la hoja (nivel más interno)
    """
    valores         : list[ValorHeader] = field(default_factory=list)
    n_niveles_grupo : int               = 0
    tipo_hoja       : HeaderTipo        = "texto"


@dataclass
class TableRegion:
    """
    Representa una sub-tabla detectada en la hoja.
    Todos los índices son base-1 (estilo Excel) en la API pública.

    Campos heredados (compatibilidad v3-v6):
        id, tipo, orientacion, fila_inicio, fila_fin, col_inicio, col_fin,
        fila_header, col_manager, col_metrica, col_headers, titulo, data

    Campos nuevos (v7):
        eje_filas  : Descripción estructural del eje de filas
        eje_cols   : Descripción estructural del eje de columnas
    """

    id:          int
    tipo:        TableType
    orientacion: Orientation
    fila_inicio: int
    fila_fin:    int
    col_inicio:  int
    col_fin:     int
    fila_header: int | None
    col_manager: int | None
    col_metrica: int | None
    col_headers: dict[int, str]
    titulo:      str | None
    data:        pd.DataFrame | None = field(repr=False)
    eje_filas:   EjeInfo             = field(default_factory=EjeInfo)
    eje_cols:    EjeInfo             = field(default_factory=EjeInfo)

    @property
    def score(self) -> float:
        """
        Puntuación para selección de tabla principal:
            score = n_filas_datos × n_columnas_fecha

        n_filas_datos = fila_fin - fila_header  (excluye la fila de header).
        Si fila_header es None (clave_valor) se usa fila_inicio, pero en ese
        caso n_columnas_fecha = 0 así que score = 0 de todas formas.
        max(0, …) evita scores negativos en regiones degeneradas (solo header).
        """
        ref     = self.fila_header if self.fila_header is not None else self.fila_inicio
        n_filas = max(0, self.fila_fin - ref)
        n_fechas = len(self.col_headers)

        # Densidad = celdas de datos esperadas / celdas totales del bloque.
        # Penaliza tablas grandes con mucho espacio vacío que podrían ganar
        # solo por tamaño físico.
        total     = (self.fila_fin - self.fila_inicio + 1) * (self.col_fin - self.col_inicio + 1)
        meaningful = n_filas * max(n_fechas, 1)
        densidad  = meaningful / max(total, 1)

        return float(n_filas * n_fechas * densidad)

    def as_dict(self) -> dict[str, Any]:
        """Devuelve un dict compatible con la API pública v3+ y campos nuevos v7."""
        return {
            # ── campos heredados (v3-v6) ──────────────────────────────────
            "id":          self.id,
            "tipo":        self.tipo,
            "orientacion": self.orientacion,
            "fila_inicio": self.fila_inicio,
            "fila_fin":    self.fila_fin,
            "col_inicio":  self.col_inicio,
            "col_fin":     self.col_fin,
            "fila_header": self.fila_header,
            "col_manager": self.col_manager,
            "col_metrica": self.col_metrica,
            "col_headers": self.col_headers,
            "titulo":      self.titulo,
            "score":       self.score,
            "data":        self.data,
            # ── campos nuevos (v7) ────────────────────────────────────────
            "eje_filas":   self.eje_filas,
            "eje_cols":    self.eje_cols,
        }


# RAMA ELIMINADA: RegionDetector (densidad de filas + subdivisión por ancla)
# Reemplazado por HeaderFirstDetector en _region_detector.py que implementa
# el método de referencia: header-first (firma de encabezado → gaps → columnas).

class TableParser:
    """
    Convierte un rango de filas (r0, r1) base-0 en uno o varios TableRegion.

    Todo acceso a celdas va contra scanner._mat (numpy, O(1)).
    Sin llamadas a openpyxl después de la construcción del scanner.
    """

    # Umbral de densidad numérica para identificar el data block
    _UMBRAL_NUMERICO   : float = 0.50
    _UMBRAL_FALLBACK   : float = 0.30
    _MIN_DATA_COLS     : int   = 1
    _MIN_DATA_ROWS     : int   = 1

    def __init__(self, scanner: SheetScanner) -> None:
        self.sc = scanner

    # ── API pública ────────────────────────────────────────────────────────

    def parse(
        self,
        r0: int,
        r1: int,
        region_id: int,
        c0: int | None = None,
        c1: int | None = None,
    ) -> TableRegion | None:
        """Retorna el primer TableRegion detectado (compatibilidad v3-v6)."""
        results = self.parse_multi(r0, r1, region_id, c0=c0, c1=c1)
        return results[0] if results else None

    def parse_multi(
        self,
        r0: int,
        r1: int,
        region_id: int,
        c0: int | None = None,
        c1: int | None = None,
    ) -> list[TableRegion]:
        """
        Parsea el rango y retorna todos los TableRegion detectados.
        Puede retornar múltiples si hay bloques horizontales (grupos de columnas).
        """
        c0, c1 = self._bbox_cols(r0, r1, c0_hint=c0, c1_hint=c1)
        if c0 > c1:
            return []

        # ── Paso 1: detectar data block ───────────────────────────────────
        data_cols, data_rows = self._detectar_data_block(r0, r1, c0, c1)

        if not data_cols or not data_rows:
            # Sin data block → intentar clave-valor como fallback
            kv = self._parse_clave_valor(r0, r1, c0, c1, region_id)
            if kv is not None:
                return [kv]
            t = self._parse_generic_table(r0, r1, c0, c1, region_id)
            return [t] if t is not None else []

        # Header top: filas antes de data_rows
        header_top_rows  = [r for r in range(r0, r1 + 1) if r not in set(data_rows) and r < min(data_rows)]
        # Header left: columnas antes de data_cols que tienen contenido de texto
        # en al menos una fila de data_rows. Excluye columnas vacías o numéricas.
        _data_rows_set = set(data_rows)
        _min_data_col  = min(data_cols)
        def es_columna_grupo(scanner: SheetScanner, rr0: int, rr1: int, c: int) -> bool:
            """
            Determina si una columna es de grupo/etiqueta (header izquierdo).

            Criterios:
              - Al menos 60% de celdas son strings (señal de etiqueta, no datos).
              - Algún grado de repetición (>= 10%) O alta densidad de strings
                (>= 80%). Columnas de métrica tienen valores únicos pero alta
                densidad de texto; columnas de grupo tienen valores repetidos.
                El umbral bajo de repetición (10%) admite ambos patrones.
            """
            vals = [scanner[r, c] for r in range(rr0, rr1 + 1)]

            strings = [v for v in vals if isinstance(v, str) and v.strip()]
            if not strings:
                return False

            string_ratio = len(strings) / len(vals)
            if string_ratio < 0.6:
                return False

            from collections import Counter
            counts = Counter(strings)
            repetition_score = max(counts.values()) / len(strings)

            # Columnas con alta densidad de texto (>=80%) son header izquierdo
            # aunque cada valor sea único (ej: columna de métricas).
            # Columnas con menor densidad necesitan algo de repetición para
            # distinguirse de columnas de datos con texto disperso.
            return string_ratio >= 0.8 or repetition_score >= 0.15

        header_left_cols = []
        for c in range(c0, c1 + 1):
            if c in set(data_cols) or c >= _min_data_col:
                continue
            # Verificar que tenga al menos un valor de texto en data_rows
            tiene_texto = any(
                isinstance(self.sc[r, c], str) and str(self.sc[r, c]).strip()
                for r in data_rows
            )
            if tiene_texto:
                header_left_cols.append(c)

        if data_rows:
            r_data_min = min(data_rows)
            r_data_max = max(data_rows)
            header_left_cols = [
                c for c in header_left_cols
                if es_columna_grupo(self.sc, r_data_min, r_data_max, c)
            ]

        # ── Paso 2 y 3: clasificar headers y detectar orientación ─────────
        celdas_top  = self._clasificar_header_top(header_top_rows, data_cols, r0, r1, c0, c1)
        celdas_left = self._clasificar_header_left(header_left_cols, data_rows, r0, r1, c0, c1)

        orientacion = self._detectar_orientacion(celdas_top, celdas_left, data_cols, data_rows)

        # ── Paso 4: construir jerarquía ───────────────────────────────────
        eje_cols = self._construir_eje(celdas_top,  data_cols,  orientacion, "cols")
        eje_filas = self._construir_eje(celdas_left, data_rows, orientacion, "filas")

        # ── Paso 5: inferir años ──────────────────────────────────────────
        self._inferir_anios(eje_cols)
        self._inferir_anios(eje_filas)

        # ── Paso 6: construir DataFrame único ─────────────────────────────
        df = self._construir_dataframe(
            data_rows, data_cols,
            celdas_left, header_left_cols,
            celdas_top, header_top_rows,
            orientacion
        )
        if df is None:
            t = self._parse_generic_table(r0, r1, c0, c1, region_id)
            return [t] if t is not None else []

        # Calidad: si las columnas parecen valores de datos (números como strings)
        # o son nombres duplicados del título merged, la clasificación jerárquica falló.
        # Fallback a tabla genérica que lee headers de la fila de encabezado.
        if df.columns.tolist():
            cols_str = [str(c) for c in df.columns]
            n_numeric_names = sum(1 for c in cols_str if self._looks_like_number(c))
            if n_numeric_names > len(cols_str) * 0.5 and len(cols_str) >= 2:
                t = self._parse_generic_table(r0, r1, c0, c1, region_id)
                return [t] if t is not None else []
            # Columnas con nombres repetidos del título (merged cell mal interpretado)
            unique_base = set(c.rstrip('_0123456789') for c in cols_str)
            if len(unique_base) == 1 and len(cols_str) >= 2:
                t = self._parse_generic_table(r0, r1, c0, c1, region_id)
                return [t] if t is not None else []

        titulo = self._detectar_titulo(r0, c0, c1, header_top_rows, header_left_cols)
        tipo   = self._inferir_tipo(eje_filas, eje_cols, header_left_cols)

        # col_manager y col_metrica para compatibilidad
        col_manager_b1, col_metrica_b1 = self._cols_compat(header_left_cols)

        return [TableRegion(
            id          = region_id,
            tipo        = tipo,
            orientacion = orientacion,
            fila_inicio = r0 + 1,
            fila_fin    = r1 + 1,
            col_inicio  = c0 + 1,
            col_fin     = c1 + 1,
            fila_header = (min(header_top_rows) + 1) if header_top_rows else None,
            col_manager = col_manager_b1,
            col_metrica = col_metrica_b1,
            col_headers = {c + 1: vh.label for c, vh in zip(data_cols, eje_cols.valores)} if eje_cols.valores else {},
            titulo      = titulo,
            data        = df,
            eje_filas   = eje_filas,
            eje_cols    = eje_cols,
        )]

    def _numeric_ratio_cols(self, r0: int, r1: int, c0: int, c1: int) -> np.ndarray:
        """
        Para cada columna en [c0..c1]: ratio = celdas numéricas / celdas no-vacías.
        Excluye bools. Retorna array de float de longitud (c1-c0+1).

        Vectorizado con _mask_numeric y _mask_meaningful (precalculadas en SheetScanner).
        Evita el loop Python de versiones anteriores — C-level puro para hojas grandes.
        División segura: columnas completamente vacías → ratio 0.0 (no NaN).
        """
        num  = self.sc._mask_numeric   [r0 : r1 + 1, c0 : c1 + 1].sum(axis=0).astype(float)
        deno = self.sc._mask_meaningful[r0 : r1 + 1, c0 : c1 + 1].sum(axis=0).astype(float)
        return np.divide(num, deno, out=np.zeros_like(num), where=deno > 0)

    def _numeric_ratio_rows(self, r0: int, r1: int, c0: int, c1: int) -> np.ndarray:
        """
        Para cada fila en [r0..r1]: ratio = celdas numéricas / celdas no-vacías.
        Retorna array de float de longitud (r1-r0+1).

        Vectorizado con _mask_numeric y _mask_meaningful (precalculadas en SheetScanner).
        División segura: filas completamente vacías → ratio 0.0 (no NaN).
        """
        num  = self.sc._mask_numeric   [r0 : r1 + 1, c0 : c1 + 1].sum(axis=1).astype(float)
        deno = self.sc._mask_meaningful[r0 : r1 + 1, c0 : c1 + 1].sum(axis=1).astype(float)
        return np.divide(num, deno, out=np.zeros_like(num), where=deno > 0)

    def _columnas_alineadas(
        self,
        r0: int, r1: int, c0: int, c1: int,
        min_hits: int | None = None,
    ) -> np.ndarray:
        """
        Column alignment detection (inspirado en Camelot/Tabula).

        Devuelve un array bool de longitud (c1-c0+1) donde True indica que
        la columna tiene ≥ min_hits celdas numéricas alineadas verticalmente.
        Es señal fuerte de columna de datos incluso en tablas sparse donde
        numeric_ratio cae bajo el umbral (pocas celdas no-None en total).

        Diferencia vs _numeric_ratio_cols:
          - numeric_ratio normaliza por celdas no-vacías → penaliza columnas
            sparse aunque tengan muchos números.
          - _columnas_alineadas usa conteo absoluto → detecta columnas de
            datos aunque solo tengan valores en el 15% de las filas.

        Args:
            min_hits: mínimo de celdas numéricas para considerar la columna
                      alineada. Default = max(3, 15% de filas del rango).
        """
        n_filas = r1 - r0 + 1
        if min_hits is None:
            min_hits = max(3, int(n_filas * 0.15))

        hits = self.sc._mask_numeric[r0 : r1 + 1, c0 : c1 + 1].sum(axis=0)
        return hits >= min_hits

    def _detectar_fila_header_mixta(
        self,
        r0: int, r1: int, c0: int, c1: int,
        max_filas: int = 10,
    ) -> int | None:
        """
        Header type transition detection (inspirado en Power Query / Tabula).

        Busca la primera fila que contiene una mezcla de texto y valores
        numéricos/fechas — patrón típico de una fila de encabezado:

            Métrica  | Enero | Febrero | Marzo    ← str + str-fecha + str-fecha
            Manager  | 2024  | 2025    | 2026     ← str + num + num
            Producto | Q1    | Q2      | Q3       ← str + str-periodo + str-periodo

        Criterio: al menos 10% de celdas son texto (etiqueta izquierda) Y
        al menos 30% son fecha/número/fecha-string (encabezados de datos).

        Más robusto que solo contar datetimes porque detecta:
          - Nombres de mes como strings ("Enero", "Q1") que no son datetime.
          - Años como enteros (2024, 2025) que tampoco son datetime.
          - Trimestres y períodos textuales.

        Returns:
            Índice base-0 de la fila header, o None si no se detecta.
        """
        for r in range(r0, min(r0 + max_filas, r1 + 1)):
            row     = self.sc._mat[r, c0 : c1 + 1]
            no_none = [v for v in row if v is not None]
            if len(no_none) < 3:
                continue

            n_str      = sum(1 for v in no_none if isinstance(v, str))
            n_num      = sum(1 for v in no_none
                             if isinstance(v, (int, float)) and not isinstance(v, bool))
            n_dt       = sum(1 for v in no_none
                             if isinstance(v, (datetime.datetime, datetime.date)))
            n_date_str = sum(1 for v in no_none
                             if isinstance(v, str) and _is_date_string(v))
            total      = len(no_none)

            str_ratio      = n_str      / total
            temporal_ratio = (n_num + n_dt + n_date_str) / total

            # Patrón header: algo de texto (etiqueta) + bastante temporal
            if str_ratio > 0.1 and temporal_ratio > 0.3:
                return r

        return None

    # ── Paso 1a: detectar headers ─────────────────────────────────────────

    def _detectar_filas_header(
        self, r0: int, r1: int, c0: int, c1: int
    ) -> tuple[set[int], int, np.ndarray]:
        """
        Detecta filas de header (fechas, headers mixtos, filas de años).

        Returns:
            (filas_header, n_fechas_fila_max, dt_counts_rows)
        """
        dt_counts_rows = self.sc._mask_datetime[r0 : r1 + 1, c0 : c1 + 1].sum(axis=1)
        n_fechas_fila  = int(dt_counts_rows.max()) if len(dt_counts_rows) > 0 else 0

        filas_header: set[int] = set()
        for i in range(r1 - r0 + 1):
            if dt_counts_rows[i] >= max(2, n_fechas_fila * 0.3):
                filas_header.add(r0 + i)

        # Header type transition: fila con mezcla texto + temporal
        fila_mixta = self._detectar_fila_header_mixta(r0, r1, c0, c1)
        if fila_mixta is not None:
            filas_header.add(fila_mixta)

        # Filas de años: filas adyacentes a una fila de fechas donde todos
        # los valores numéricos parecen años (enteros en rango 20-99 o 2000-2100).
        # Patrón típico en hojas financieras:
        #   Fila N:   22    22    22    23    23    24    25    2026   ← años
        #   Fila N+1: Ago.  Sep.  Oct   Ene   Feb   Ene   Ene  Ene   ← meses
        if filas_header:
            filas_anio = self._detectar_filas_anio(r0, r1, c0, c1, filas_header)
            filas_header.update(filas_anio)

        return filas_header, n_fechas_fila, dt_counts_rows

    def _detectar_filas_anio(
        self, r0: int, r1: int, c0: int, c1: int,
        filas_fecha: set[int],
    ) -> set[int]:
        """
        Detecta filas de años adyacentes a filas de header de fechas.

        Una fila es "de años" si:
          - Tiene >= 3 valores numéricos en el rango de datos.
          - Todos los valores numéricos son enteros que parecen años
            (rango 20-99 para años de 2 dígitos, o 2000-2100 para 4 dígitos).
          - Está inmediatamente antes o después de una fila de fechas conocida.

        Esta heurística es general para hojas financieras con headers
        multinivel (año + mes).
        """
        resultado: set[int] = set()
        # Solo buscar en las filas adyacentes a filas de fecha
        candidatas: set[int] = set()
        for f in filas_fecha:
            if f - 1 >= r0:
                candidatas.add(f - 1)
            if f + 1 <= r1:
                candidatas.add(f + 1)
        candidatas -= filas_fecha  # no re-evaluar filas ya marcadas

        for r in candidatas:
            # Extraer valores numéricos de la fila
            nums = []
            for c in range(c0, c1 + 1):
                v = self.sc[r, c]
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    nums.append(v)

            if len(nums) < 3:
                continue

            # Verificar que todos parecen años
            todos_anio = True
            for v in nums:
                vi = int(v)
                if vi != v:  # no es entero
                    todos_anio = False
                    break
                # Año de 2 dígitos (20-99) o 4 dígitos (2000-2100)
                if not ((20 <= vi <= 99) or (2000 <= vi <= 2100)):
                    todos_anio = False
                    break

            if todos_anio:
                resultado.add(r)

        return resultado

    def _detectar_cols_header(
        self, r0: int, r1: int, c0: int, c1: int,
        dt_counts_cols: np.ndarray,
    ) -> set[int]:
        """
        Detecta columnas de header (etiquetas de texto como Manager, Métrica).
        Una columna es header si ≥60% de sus celdas son strings y tiene <2 fechas.
        Vectorizado con _mask_meaningful y _mask_numeric.
        """
        sub_meaningful = self.sc._mask_meaningful[r0 : r1 + 1, c0 : c1 + 1]
        sub_numeric    = self.sc._mask_numeric   [r0 : r1 + 1, c0 : c1 + 1]

        n_meaningful = sub_meaningful.sum(axis=0)          # (ncols,)
        n_numeric    = sub_numeric.sum(axis=0)             # (ncols,)
        n_str        = n_meaningful - n_numeric            # aprox strings + dates

        cols_header: set[int] = set()
        for j in range(c1 - c0 + 1):
            if n_meaningful[j] == 0:
                continue
            str_ratio = int(n_str[j]) / int(n_meaningful[j])
            if str_ratio >= 0.6 and dt_counts_cols[j] < 2:
                cols_header.add(c0 + j)
        return cols_header

    # ── Paso 1b: scoring continuo ─────────────────────────────────────────

    def _score_cols(
        self,
        filas_candidatas: list[int],
        cols_candidatas: list[int],
        dt_counts_cols: np.ndarray,
        c0: int,
    ) -> np.ndarray:
        """
        Score continuo [0,1] por columna candidata.
        Combina: α·numeric_ratio + β·alignment_score + γ·date_signal.

        Args:
            filas_candidatas: filas de datos (excluye filas de header).
            cols_candidatas:  columnas candidatas (excluye header izquierdo).
            dt_counts_cols:   número de fechas por columna, relativo a c0.
            c0:               columna inicial del bbox de la región.

        Returns:
            np.ndarray float de longitud len(cols_candidatas).
        """
        if not filas_candidatas or not cols_candidatas:
            return np.array([], dtype=float)

        r0c, r1c = min(filas_candidatas), max(filas_candidatas)
        c0c, c1c = min(cols_candidatas),  max(cols_candidatas)
        n_r = len(filas_candidatas)

        _α, _β, _γ = 0.55, 0.25, 0.20

        # Extraer solo las columnas en cols_candidatas (pueden ser no-contiguas)
        cols_rel  = np.array([cabs - c0c for cabs in cols_candidatas], dtype=int)
        ratio_all = self._numeric_ratio_cols(r0c, r1c, c0c, c1c)
        ratio     = ratio_all[cols_rel]
        hits_all  = self.sc._mask_numeric[r0c : r1c + 1, c0c : c1c + 1].sum(axis=0).astype(float)
        align     = np.clip(hits_all[cols_rel] / max(n_r * 0.15, 1), 0, 1)
        date_s    = np.array([
            1.0 if dt_counts_cols[cabs - c0] >= 1 else 0.0
            for cabs in cols_candidatas
        ])
        return _α * ratio + _β * align + _γ * date_s

    def _score_rows(
        self,
        filas_candidatas: list[int],
        cols_candidatas: list[int],
        dt_counts_rows: np.ndarray,
        r0: int,
    ) -> np.ndarray:
        """
        Score continuo [0,1] por fila candidata.
        Combina: α·numeric_ratio + β·alignment_score + γ·date_signal.

        Args:
            dt_counts_rows: número de fechas por fila, relativo a r0.

        Returns:
            np.ndarray float de longitud len(filas_candidatas).
        """
        if not filas_candidatas or not cols_candidatas:
            return np.array([], dtype=float)

        r0c, r1c = min(filas_candidatas), max(filas_candidatas)
        c0c, c1c = min(cols_candidatas),  max(cols_candidatas)
        n_c = len(cols_candidatas)

        _α, _β = 0.55, 0.25

        # Extraer solo las filas en filas_candidatas (pueden ser no-contiguas)
        rows_rel  = np.array([rabs - r0c for rabs in filas_candidatas], dtype=int)
        ratio_all = self._numeric_ratio_rows(r0c, r1c, c0c, c1c)
        ratio     = ratio_all[rows_rel]
        hits_all  = self.sc._mask_numeric[r0c : r1c + 1, c0c : c1c + 1].sum(axis=1).astype(float)
        align     = np.clip(hits_all[rows_rel] / max(n_c * 0.15, 1), 0, 1)
        return _α * ratio + _β * align

    def _choose_cut(
        self,
        scores: np.ndarray,
        indices_base: list[int],
    ) -> list[int]:
        """
        Elige el corte en el gap natural entre scores ordenados.
        Si no hay gap claro (≥0.15), usa umbral estático _UMBRAL_FALLBACK.

        Returns:
            Lista de índices base cuyo score supera el corte.
        """
        if len(scores) == 0:
            return []
        if len(scores) == 1:
            return [indices_base[0]] if scores[0] >= self._UMBRAL_FALLBACK else []

        sorted_s  = np.sort(scores)
        gaps      = np.diff(sorted_s)
        max_gap   = float(gaps.max())

        if max_gap >= 0.15:
            corte = float(sorted_s[int(gaps.argmax())])
            return [idx for idx, s in zip(indices_base, scores) if s > corte]
        else:
            return [idx for idx, s in zip(indices_base, scores)
                    if s >= self._UMBRAL_FALLBACK]

    # ── Paso 1 (orquestador) ──────────────────────────────────────────────

    def _detectar_data_block(
        self, r0: int, r1: int, c0: int, c1: int
    ) -> tuple[list[int], list[int]]:
        """
        Detecta el bloque denso de datos numéricos.
        Retorna (data_cols, data_rows) como listas de índices base-0 absolutos.

        Pipeline separado en responsabilidades únicas:
          1a. _detectar_filas_header  → filas de header (fechas, mixtos)
          1b. _detectar_cols_header   → columnas de header (etiquetas texto)
          1c. _score_cols / _score_rows → scoring continuo
          1d. _choose_cut             → corte por gap natural
          1e. Fallback sparse         → tablas con pocos datos numéricos
        """
        # 1a — Detectar headers
        filas_header, n_fechas_fila, dt_counts_rows = self._detectar_filas_header(r0, r1, c0, c1)
        dt_counts_cols = self.sc._mask_datetime[r0 : r1 + 1, c0 : c1 + 1].sum(axis=0)
        n_fechas_col   = int(dt_counts_cols.max()) if len(dt_counts_cols) > 0 else 0
        cols_header    = self._detectar_cols_header(r0, r1, c0, c1, dt_counts_cols)

        filas_candidatas = [r for r in range(r0, r1 + 1) if r not in filas_header]
        cols_candidatas  = [c for c in range(c0, c1 + 1) if c not in cols_header]

        if not filas_candidatas or not cols_candidatas:
            return [], []

        # 1b — Scoring continuo por eje
        sc_c = self._score_cols(filas_candidatas, cols_candidatas, dt_counts_cols, c0)
        sc_r = self._score_rows(filas_candidatas, cols_candidatas, dt_counts_rows, r0)

        # 1c — Corte por gap natural
        dc_score = self._choose_cut(sc_c, cols_candidatas)
        dr_score = self._choose_cut(sc_r, filas_candidatas)

        # Expandir dc con columnas de fecha
        if n_fechas_fila >= 2:
            for j in range(c1 - c0 + 1):
                if dt_counts_cols[j] >= 1:
                    cf = c0 + j
                    if cf not in cols_header and cf not in dc_score:
                        dc_score.append(cf)
            dc_score.sort()

        # Expandir dr con filas de fecha (tabla rotada)
        if n_fechas_col >= 2 and n_fechas_fila < 2:
            for i in range(r1 - r0 + 1):
                if dt_counts_rows[i] >= 1:
                    rf = r0 + i
                    if rf not in filas_header and rf not in dr_score:
                        dr_score.append(rf)
            dr_score.sort()

        # 1d — Validar y retornar si suficiente
        if len(dc_score) >= self._MIN_DATA_COLS and len(dr_score) >= self._MIN_DATA_ROWS:
            # ordenar columnas primero
            dc_score = sorted(dc_score)

            # Expandir dc_score para incluir columnas intercaladas entre columnas de
            # datos conocidas (p.ej. montos de presupuesto entre fechas).
            # Si una columna dentro del rango [min, max] de dc_score no esta en
            # cols_header y tiene datos numericos en filas de datos, pertenece a la tabla.
            dc = self._expandir_cols_con_datos_intercalados(
                dc_score, cols_header, filas_candidatas
            )

            # Separar tablas distintas: si hay un gap donde todas las columnas
            # intermedias estan vacias (sin datos numericos), truncar en ese punto.
            if len(dc) > 3 and filas_candidatas:
                diffs = np.diff(dc)
                max_gap_j = int(np.argmax(diffs))
                if diffs[max_gap_j] > 1:
                    gap_c0 = dc[max_gap_j] + 1
                    gap_c1 = dc[max_gap_j + 1] - 1
                    if gap_c0 <= gap_c1:
                        r0c = min(filas_candidatas)
                        r1c = max(filas_candidatas)
                        gap_numeric = self.sc._mask_numeric[r0c : r1c + 1, gap_c0 : gap_c1 + 1]
                        if not gap_numeric.any():
                            # Gap vacio -> tablas distintas, truncar
                            dc = dc[:max_gap_j + 1]

            dr = self._rango_contiguo_mayor(dr_score)
            if len(dr) >= len(filas_candidatas) * 0.4:
                return dc, dr

        # 1e — Fallback sparse: columnas de fecha × filas con dato en header
        if n_fechas_fila >= 2 and filas_header:
            cols_fecha = sorted(c0 + j for j in range(c1 - c0 + 1) if dt_counts_cols[j] >= 1)
            filas_con_dato = [
                r for r in filas_candidatas
                if self.sc._mask_meaningful[r, list(cols_header)].any()
            ] if cols_header else filas_candidatas
            if cols_fecha and filas_con_dato:
                # Expandir cols_fecha para incluir columnas de datos intercaladas entre
                # fechas (p.ej. montos de presupuesto/FN usados como cabeceras de período).
                # Si una columna no-fecha cae dentro del rango [min_fecha, max_fecha] y
                # tiene datos numéricos en las filas de datos, pertenece a la tabla.
                if len(cols_fecha) >= 2 and filas_con_dato:
                    r0c = min(filas_con_dato)
                    r1c = max(filas_con_dato)
                    c_min, c_max = cols_fecha[0], cols_fecha[-1]
                    cols_rango_completo = []
                    for c in range(c_min, c_max + 1):
                        if c in cols_fecha:
                            cols_rango_completo.append(c)
                        elif c not in cols_header:
                            # Incluir si tiene datos numéricos en filas de datos
                            if self.sc._mask_numeric[r0c : r1c + 1, c].any():
                                cols_rango_completo.append(c)
                    if cols_rango_completo:
                        return cols_rango_completo, filas_con_dato
                return self._rango_contiguo_mayor(cols_fecha), filas_con_dato

        if n_fechas_col >= 2 and cols_header:
            filas_fecha = sorted(r0 + i for i in range(r1 - r0 + 1) if dt_counts_rows[i] >= 1)
            cols_con_dato = [
                c for c in cols_candidatas
                if self.sc._mask_meaningful[list(filas_header), c].any()
            ] if filas_header else cols_candidatas
            if filas_fecha and cols_con_dato:
                return cols_con_dato, self._rango_contiguo_mayor(filas_fecha)

        return [], []

    @staticmethod
    def _rango_contiguo_mayor(indices: list[int]) -> list[int]:
        """Dado una lista de índices, retorna el sub-rango contiguo más largo."""
        if not indices:
            return []
        mejor_inicio = 0
        mejor_largo  = 1
        inicio_actual = 0
        for i in range(1, len(indices)):
            if indices[i] == indices[i - 1] + 1:
                largo = i - inicio_actual + 1
                if largo > mejor_largo:
                    mejor_largo  = largo
                    mejor_inicio = inicio_actual
            else:
                inicio_actual = i
        return indices[mejor_inicio : mejor_inicio + mejor_largo]

    def _expandir_cols_con_datos_intercalados(
        self,
        dc_score: list[int],
        cols_header: set[int],
        filas_candidatas: list[int],
    ) -> list[int]:
        """
        Expande una lista de columnas candidatas para incluir columnas intercaladas
        que tienen datos numericos en las filas de datos, aunque no aparezcan en
        dc_score (por no haber pasado el umbral de scoring o de fechas).

        Caso tipico: tablas donde las columnas de datos alternan entre fechas y
        montos de presupuesto/FN:
            col:  fecha  fecha  monto  fecha  fecha  monto  ...
        El scoring detecta los montos pero _rango_contiguo_mayor descarta los grupos
        separados por gap=2. Esta funcion los recupera verificando si las columnas
        del gap tienen datos numericos reales en las filas de datos.

        Args:
            dc_score:        columnas candidatas ordenadas (base-0).
            cols_header:     columnas de etiquetas texto a excluir.
            filas_candidatas: filas de datos (excluye header rows).

        Returns:
            Lista ordenada de columnas incluyendo las intercaladas con datos.
        """
        if len(dc_score) < 2 or not filas_candidatas:
            return dc_score

        r0c = min(filas_candidatas)
        r1c = max(filas_candidatas)
        c_min, c_max = dc_score[0], dc_score[-1]

        # Detectar el paso periodico del patron intercalado (si existe).
        # Si las brechas entre elementos de dc_score son uniformes, hay un patron
        # periodico: p.ej. [2,3,4,5,6,7, 9,10, 12,13, 15,16, 18,19] tiene
        # brechas [1,1,1,1,1, 2, 1, 2, 1, 2, 1, 2, 1].
        # El tamano del grupo periodico = min brecha entre bloques de fechas.
        diffs_dc = np.diff(dc_score) if len(dc_score) > 1 else np.array([1])
        # Paso periodico: el gap entre grupos consecutivos de fechas
        paso_periodico = int(diffs_dc.max())  # gap maximo entre cols de dc_score

        result = []
        dc_set = set(dc_score)
        for c in range(c_min, c_max + 1):
            if c in dc_set:
                result.append(c)
            elif c not in cols_header:
                # Incluir si tiene al menos un dato numerico en las filas de datos
                if self.sc._mask_numeric[r0c : r1c + 1, c].any():
                    result.append(c)

        # Extender hacia la derecha si el patron periodico continua mas alla de
        # max(dc_score). Avanzar de a paso_periodico desde la ultima col conocida
        # mientras la siguiente columna no sea de header y tenga datos numericos.
        if paso_periodico >= 2:
            c_cursor = c_max + (paso_periodico - 1)
            # Limitar a columnas existentes (self.sc.ncols - 1, base-0)
            while c_cursor < self.sc.ncols and c_cursor not in cols_header:
                if self.sc._mask_numeric[r0c : r1c + 1, c_cursor].any():
                    result.append(c_cursor)
                    c_cursor += paso_periodico
                else:
                    break

        return result

    # ── Paso 2: clasificar headers ─────────────────────────────────────────

    def _clasificar_celda(self, v: Any) -> tuple[HeaderTipo, str, datetime.date | None, int | None]:
        """
        Clasifica un valor de celda como tipo de header.
        Retorna (tipo, label, fecha, anio).
        """
        if not self.sc._is_cell_meaningful(v):
            return "texto", "", None, None

        # Fecha nativa
        if isinstance(v, (datetime.datetime, datetime.date)):
            label = self.sc.fmt_fecha(v)
            fecha = v.date() if isinstance(v, datetime.datetime) else v
            return "fecha", label, fecha, fecha.year

        # Numérico
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            vi = int(v) if float(v) == int(v) else None
            if vi is not None and 2000 <= vi <= 2100:
                return "anio", str(vi), None, vi
            label = str(int(v)) if float(v) == int(v) else str(v)
            return "numero", label, None, None

        s = str(v).strip()
        sn = _normalize_date_str(s)

        # Año como string
        if re.match(r"^\d{4}$", sn) and 2000 <= int(sn) <= 2100:
            return "anio", s, None, int(sn)

        # Fecha como string
        if _is_date_string(s):
            label = self.sc.fmt_fecha(s)
            return "fecha", label, None, None

        # Periodo: Q1-Q4, Trim, Sem, Ago-Jun, etc.
        if re.match(r"^(q[1-4]|trim|sem|bim|cuatrim|[a-z]{3,9}-[a-z]{3,9})", sn):
            return "periodo", s, None, None

        return "texto", s, None, None

    def _clasificar_celdas_batch(
        self,
        coords: list[tuple[int, int]],
        contexto_anio: bool = False,
    ) -> dict[tuple[int,int], ValorHeader]:
        """
        Clasifica un conjunto de coordenadas (r, c) base-0 en batch.

        Estrategia vectorizada en dos fases:
          Fase 1 — detección rápida por masks (sin llamar _clasificar_celda):
            - _mask_datetime → "fecha"
            - int/float en rango 2000-2100 → "anio"
          Fase 2 — residuos (strings, números fuera de rango, períodos) →
            llamada individual a _clasificar_celda() solo donde necesario.

        Para hojas 20 000×200 esto reduce ≈10x las llamadas Python comparado
        con el loop original.

        Args:
            contexto_anio: si True, suprime clasificación "anio" para celdas
                           que están en columnas de header left (los años
                           como métricas — 2024, 2025 — no son marcadores
                           de año en ese contexto).
        """
        result: dict[tuple[int, int], ValorHeader] = {}
        if not coords:
            return result

        rs = np.array([r for r, c in coords], dtype=int)
        cs = np.array([c for r, c in coords], dtype=int)

        # Fase 1a — fechas nativas via _mask_datetime (C-level)
        is_dt = self.sc._mask_datetime[rs, cs]

        # Fase 1b — enteros en rango año (2000-2100) via _mask_numeric
        is_num = self.sc._mask_numeric[rs, cs]

        for i, (r, c) in enumerate(coords):
            v = self.sc[r, c]
            if not self.sc._is_cell_meaningful(v):
                continue

            if is_dt[i]:
                # Fecha nativa: clasificación directa sin _clasificar_celda
                if isinstance(v, (datetime.datetime, datetime.date)):
                    fecha = v.date() if isinstance(v, datetime.datetime) else v
                    label = self.sc.fmt_fecha(v)
                    result[(r, c)] = ValorHeader(label=label, tipo="fecha",
                                                  fecha=fecha, anio=fecha.year)
                else:
                    # String fecha-like: delegar (necesita fmt_fecha con regex)
                    tipo, label, fecha, anio = self._clasificar_celda(v)
                    result[(r, c)] = ValorHeader(label=label, tipo=tipo, anio=anio, fecha=fecha)
                continue

            if is_num[i] and not contexto_anio:
                vi = int(v) if isinstance(v, float) and float(v) == int(v) else (v if isinstance(v, int) else None)
                if vi is not None and 2000 <= vi <= 2100:
                    result[(r, c)] = ValorHeader(label=str(vi), tipo="anio", anio=vi)
                    continue

            # Fase 2 — residuos: strings, números fuera de rango, períodos
            tipo, label, fecha, anio = self._clasificar_celda(v)
            # P5: suprimir "anio" en contexto_anio (años como métricas)
            if contexto_anio and tipo == "anio":
                tipo = "texto"
                anio = None
            result[(r, c)] = ValorHeader(label=label, tipo=tipo, anio=anio, fecha=fecha)

        return result

    def _clasificar_header_top(
        self,
        header_rows: list[int],
        data_cols: list[int],
        r0: int, r1: int, c0: int, c1: int
    ) -> dict[tuple[int,int], ValorHeader]:
        """
        Clasifica cada celda en las filas de header superior (vectorizado).
        Solo considera columnas dentro de data_cols.
        Retorna {(fila_b0, col_b0): ValorHeader}.
        """
        coords = [(r, c) for r in header_rows for c in data_cols]
        return self._clasificar_celdas_batch(coords, contexto_anio=False)

    def _clasificar_header_left(
        self,
        header_cols: list[int],
        data_rows: list[int],
        r0: int, r1: int, c0: int, c1: int
    ) -> dict[tuple[int,int], ValorHeader]:
        """
        Clasifica cada celda en las columnas de header izquierdo (vectorizado).
        Solo considera filas dentro de data_rows.
        Retorna {(fila_b0, col_b0): ValorHeader}.

        Usa contexto_anio=True para no confundir años como métricas
        (tablas financieras con columnas 2024, 2025, 2026 como índices).
        """
        coords = [(r, c) for c in header_cols for r in data_rows]
        return self._clasificar_celdas_batch(coords, contexto_anio=True)

    # ── Paso 3: detectar orientación ─────────────────────────────────────

    def _detectar_orientacion(
        self,
        celdas_top: dict[tuple[int,int], ValorHeader],
        celdas_left: dict[tuple[int,int], ValorHeader],
        data_cols: list[int],
        data_rows: list[int],
    ) -> Orientation:
        """
        Decide si las fechas están en columnas (header top) o en filas (header left).
        Usa scoring normalizado por tamaño del eje.
        """
        n_fechas_top  = sum(1 for vh in celdas_top.values()  if vh.tipo in ("fecha", "anio", "periodo"))
        n_fechas_left = sum(1 for vh in celdas_left.values() if vh.tipo in ("fecha", "anio", "periodo"))

        score_top  = n_fechas_top  / max(len(data_cols), 1)
        score_left = n_fechas_left / max(len(data_rows), 1)

        if score_top == 0 and score_left == 0:
            return "ambiguous"
        if score_top >= score_left:
            return "column"
        return "row"

    # ── Paso 4: construir jerarquía de eje ─────────────────────────────────

    def _construir_eje(
        self,
        celdas: dict[tuple[int,int], ValorHeader],
        indices_datos: list[int],
        orientacion: Orientation,
        cual: str,  # "cols" o "filas"
    ) -> EjeInfo:
        """
        Construye EjeInfo a partir de las celdas clasificadas de un eje.

        Para eje_cols (cual=="cols"): los índices_datos son columnas,
            y las celdas están en header_top_rows × data_cols.
            El valor de cada columna de datos es la celda de la fila de header
            más cercana al data block (la hoja). Las filas de header más
            lejanas son grupos.

        Para eje_filas (cual=="filas"): los índices_datos son filas,
            y las celdas están en data_rows × header_left_cols.
            La columna más a la derecha antes del data block es la hoja.
        """
        if not celdas or not indices_datos:
            return EjeInfo()

        if cual == "cols":
            # Agrupar por columna (índice_datos = columna)
            por_col: dict[int, list[tuple[int, ValorHeader]]] = {}
            for (r, c), vh in celdas.items():
                if c in indices_datos:
                    por_col.setdefault(c, []).append((r, vh))

            # Para cada columna de datos: la fila más baja (mayor r) = hoja
            # las filas más altas = grupos
            valores_hoja: list[ValorHeader] = []
            for c in indices_datos:
                entradas = sorted(por_col.get(c, []), key=lambda x: x[0])
                if not entradas:
                    valores_hoja.append(ValorHeader(label="", tipo="texto"))
                    continue
                # Última fila = hoja
                _, vh_hoja = entradas[-1]
                # Primeras filas = grupos (nivel 0, 1, ...)
                for nivel, (_, vh_grupo) in enumerate(entradas[:-1]):
                    vh_grupo.nivel = nivel
                    vh_grupo.tipo  = "grupo"
                vh_hoja.nivel = len(entradas) - 1
                valores_hoja.append(vh_hoja)

            n_grupos  = max((vh.nivel for vh in valores_hoja), default=0)
            tipo_hoja = self._tipo_dominante([vh for vh in valores_hoja if vh.label])
            return EjeInfo(valores=valores_hoja, n_niveles_grupo=n_grupos, tipo_hoja=tipo_hoja)

        else:  # cual == "filas"
            # Agrupar por fila (índice_datos = fila)
            por_fila: dict[int, list[tuple[int, ValorHeader]]] = {}
            for (r, c), vh in celdas.items():
                if r in indices_datos:
                    por_fila.setdefault(r, []).append((c, vh))

            # Para cada fila de datos: la columna más a la derecha = hoja (métrica)
            # las columnas más a la izquierda = grupos
            valores_hoja_filas: list[ValorHeader] = []
            for r in indices_datos:
                entradas = sorted(por_fila.get(r, []), key=lambda x: x[0])
                if not entradas:
                    valores_hoja_filas.append(ValorHeader(label="", tipo="texto"))
                    continue
                # Última columna = hoja (métrica o fecha si tabla rotada)
                _, vh_hoja = entradas[-1]
                for nivel, (_, vh_grupo) in enumerate(entradas[:-1]):
                    vh_grupo.nivel = nivel
                    vh_grupo.tipo  = "grupo"
                vh_hoja.nivel = len(entradas) - 1
                valores_hoja_filas.append(vh_hoja)

            n_grupos  = max((vh.nivel for vh in valores_hoja_filas), default=0)
            tipo_hoja = self._tipo_dominante([vh for vh in valores_hoja_filas if vh.label])
            return EjeInfo(valores=valores_hoja_filas, n_niveles_grupo=n_grupos, tipo_hoja=tipo_hoja)

    def _jerarquia_por_merges(
        self,
        header_top_rows: list[int],
        data_cols: list[int],
        eje_cols: EjeInfo,
    ) -> EjeInfo:
        """
        Enriquece EjeInfo usando merged ranges como señal de jerarquía.

        En Excel, un merge horizontal en el header significa un grupo que
        abarca varias columnas:

            [──────── 2024 ────────][──────── 2025 ────────]   ← merge (nivel grupo)
            [ ene ][ feb ][ mar ][ ene ][ feb ][ mar ]         ← celdas (nivel hoja)

        El algoritmo:
          1. Para cada fila de header_top (de arriba hacia abajo), detecta
             merges horizontales que abarcan ≥2 columnas de data_cols.
          2. Cada merge se convierte en un ValorHeader de tipo "grupo"
             que se asigna a todas las columnas que abarca.
          3. El nivel del grupo = índice de la fila de header (0 = más alto).

        Si no hay merges relevantes, retorna eje_cols sin cambios.
        Modifica eje_cols.valores in-place y actualiza n_niveles_grupo.

        Nota: esta señal es adicional a la heurística de posición de fila
        que ya usa _construir_eje(). Si ambas coinciden, el resultado es
        el mismo. Si divergen (merge inconsistente), prevalece el merge.
        """
        if not self.sc._merge_list or not header_top_rows or not data_cols:
            return eje_cols

        data_cols_set = set(data_cols)

        # Ordenar filas de header: la más alta (menor r) = nivel grupo más externo
        header_top_rows_sorted = sorted(header_top_rows)

        # Para cada columna de datos, construir lista de grupos por nivel de fila
        # grupos_por_col[c] = list[ValorHeader] en orden de fila (externo → interno)
        grupos_por_col: dict[int, list[ValorHeader]] = {c: [] for c in data_cols}

        for nivel, r in enumerate(header_top_rows_sorted[:-1]):
            # Última fila de header = hojas (ya capturada en eje_cols.valores)
            # Solo procesamos filas superiores como posibles grupos-merge

            # Detectar merges únicos en esta fila que toquen data_cols
            merges_vistos: set[tuple[int,int,int,int]] = set()
            for c in data_cols:
                bbox = self.sc.get_merge_bbox(r, c)
                if bbox is None:
                    continue  # celda no mergeada en esta fila
                if bbox in merges_vistos:
                    continue  # ya procesado
                merges_vistos.add(bbox)

                mr0, mr1, mc0, mc1 = bbox
                # Columnas de data_cols que abarca este merge
                cols_en_merge = [col for col in data_cols if mc0 <= col <= mc1]
                if len(cols_en_merge) < 2:
                    continue  # merge de 1 columna → no aporta jerarquía

                # Valor del merge = el label del grupo
                v = self.sc[r, c]
                if not self.sc._is_cell_meaningful(v):
                    continue
                tipo, label, fecha, anio = self._clasificar_celda(v)
                vh_grupo = ValorHeader(label=label, tipo="grupo", nivel=nivel, anio=anio, fecha=fecha)

                for col in cols_en_merge:
                    grupos_por_col[col].append(vh_grupo)

        # Comprobar si encontramos algún grupo por merge
        hay_grupos = any(len(gs) > 0 for gs in grupos_por_col.values())
        if not hay_grupos:
            return eje_cols  # sin merges relevantes → sin cambios

        # Reconstruir eje_cols.valores incorporando los grupos de merge
        # ANTES del valor de hoja existente
        nuevos_valores: list[ValorHeader] = []
        for c, vh_hoja in zip(data_cols, eje_cols.valores):
            grupos = grupos_por_col.get(c, [])
            # Asignar niveles consecutivos a los grupos de este merge
            for nv, vhg in enumerate(grupos):
                vhg.nivel = nv
            # La hoja queda en el nivel más profundo
            vh_hoja.nivel = len(grupos)
            nuevos_valores.append(vh_hoja)
            # Pero en EjeInfo los grupos se representan en los valores via nivel
            # Solo añadimos el vh_hoja — los grupos se leen a través de nivel
            # en _construir_dataframe al crear el MultiIndex.
            # Añadir los grupos a "pre-valores" usando el mecanismo existente:
            # guardarlos en vh_hoja como atributo extra no rompe nada downstream.
            vh_hoja._grupos_merge = grupos  # type: ignore[attr-defined]

        eje_cols.valores = nuevos_valores
        max_nivel = max(
            (vh.nivel for vh in eje_cols.valores),
            default=0,
        )
        eje_cols.n_niveles_grupo = max(eje_cols.n_niveles_grupo, max_nivel)
        return eje_cols

    @staticmethod
    def _tipo_dominante(vhs: list[ValorHeader]) -> HeaderTipo:
        """Retorna el tipo más frecuente en una lista de ValorHeader."""
        if not vhs:
            return "texto"
        conteo: dict[str, int] = {}
        for vh in vhs:
            conteo[vh.tipo] = conteo.get(vh.tipo, 0) + 1
        return max(conteo, key=lambda t: conteo[t])  # type: ignore

    # ── Paso 5: inferir años ──────────────────────────────────────────────

    def _inferir_anios(self, eje: EjeInfo) -> None:
        """
        Para valores de tipo "fecha" o "periodo" sin año asignado,
        buscar hacia la izquierda/arriba el marcador de año más cercano
        en la jerarquía (valores de tipo "anio" o "grupo" con anio).
        Modifica eje.valores in-place.
        """
        ultimo_anio: int | None = None
        for vh in eje.valores:
            if vh.tipo == "anio":
                ultimo_anio = vh.anio
            elif vh.tipo in ("fecha", "periodo") and vh.anio is None and ultimo_anio is not None:
                vh.anio = ultimo_anio

    # ── Detectar bloques horizontales (grupos de columnas) ────────────────

    def _titulo_desde_separadora(
        self, col_sep_b0: int, header_top_rows: list[int], r0: int, r1: int
    ) -> str | None:
        """
        Busca el título del bloque: valor de la columna separadora en header_top
        o en la fila justo arriba de r0.
        """
        for r in sorted(header_top_rows, reverse=True):
            v = self.sc[r, col_sep_b0]
            if self.sc._is_cell_meaningful(v):
                s = str(v).strip()
                if s:
                    return s
        if r0 > 0:
            v = self.sc[r0 - 1, col_sep_b0]
            if self.sc._is_cell_meaningful(v):
                return str(v).strip()
        return None

    # ── Paso 6: construir DataFrame ───────────────────────────────────────

    def _construir_dataframe(
        self,
        data_rows: list[int],
        data_cols: list[int],
        celdas_left: dict[tuple[int,int], ValorHeader],
        header_left_cols: list[int],
        celdas_top: dict[tuple[int,int], ValorHeader],
        header_top_rows: list[int],
        orientacion: Orientation,
    ) -> pd.DataFrame | None:
        """
        Construye el DataFrame con índice reflejo de la jerarquía de header_left
        y columnas reflejo de la jerarquía de header_top.
        """
        if not data_rows or not data_cols:
            return None

        # Etiquetas de columnas (de header_top, fila con más fechas = hoja)
        col_labels: list[str] = []
        if header_top_rows:
            # Usar la fila con más fechas como fuente de etiquetas de columna.
            # max() no es correcto: la fila de fechas puede no ser la última del header.
            dt_por_fila = {
                r: int(self.sc._mask_datetime[r, :].sum())
                for r in header_top_rows
            }
            fila_hoja = max(header_top_rows, key=lambda r: dt_por_fila[r])
            for c in data_cols:
                v = self.sc[fila_hoja, c]
                if self.sc._is_cell_meaningful(v):
                    _, label, _, _ = self._clasificar_celda(v)
                    col_labels.append(label)
                else:
                    # Buscar en otras filas de header_top si esta tiene vacío
                    found = False
                    for r_alt in sorted(header_top_rows, key=lambda r: dt_por_fila[r], reverse=True):
                        v2 = self.sc[r_alt, c]
                        if self.sc._is_cell_meaningful(v2):
                            _, label2, _, _ = self._clasificar_celda(v2)
                            col_labels.append(label2)
                            found = True
                            break
                    if not found:
                        col_labels.append(f"col{c}")
        else:
            col_labels = [f"col{c}" for c in data_cols]
        col_labels = _make_unique(col_labels)

        # Etiquetas de índice (de header_left)
        registros: list[dict[str, Any]] = []
        ultimo_grupo: dict[int, str] = {}   # nivel → último valor visto (para merges)

        for r in data_rows:
            fila: dict[str, Any] = {}

            # Niveles de grupo (columnas izquierda excepto la última = métrica)
            for nivel, c in enumerate(header_left_cols[:-1] if len(header_left_cols) > 1 else []):
                key_grupo = (r, c)
                vh = celdas_left.get(key_grupo)
                if vh and vh.label:
                    ultimo_grupo[nivel] = vh.label
                fila[f"Grupo_{nivel}"] = ultimo_grupo.get(nivel)

            # Hoja del eje de filas (métrica, fecha, texto)
            if header_left_cols:
                col_hoja = header_left_cols[-1]
                vh_hoja  = celdas_left.get((r, col_hoja))
                hoja_label = vh_hoja.label if vh_hoja and vh_hoja.label else None
                if hoja_label is None:
                    # Leer directo de la matriz
                    v = self.sc[r, col_hoja]
                    hoja_label = str(v).strip() if self.sc._is_cell_meaningful(v) else None
                if hoja_label is None:
                    continue
                fila["Métrica"] = hoja_label
            else:
                fila["Métrica"] = f"fila{r}"

            # Valores del data block
            for c, lbl in zip(data_cols, col_labels):
                v = self.sc[r, c]
                fila[lbl] = v if self.sc._is_cell_meaningful(v) else None

            registros.append(fila)

        if not registros:
            return None

        df = pd.DataFrame(registros)

        # Construir índice
        idx_cols = [k for k in df.columns if k.startswith("Grupo_") or k == "Métrica"]
        val_cols = [k for k in df.columns if k not in idx_cols]

        if not val_cols:
            return None

        df = df[idx_cols + val_cols]
        if idx_cols:
            df = df.set_index(idx_cols)
            if len(idx_cols) == 1:
                df.index.name = idx_cols[0]

        return df

    # ── Parseo clave-valor ─────────────────────────────────────────────────

    def _parse_clave_valor(
        self,
        r0: int, r1: int, c0: int, c1: int, region_id: int,
    ) -> TableRegion | None:
        """
        Parsea una región como nombre → valor.
        Funciona para tablas sin fechas de 2+ columnas donde la primera es texto.
        """
        if c1 - c0 < 1:
            return None

        # Verificar que la primera columna tiene mayoria de strings
        col_izq = [self.sc[r, c0] for r in range(r0, r1 + 1)]
        no_none = [v for v in col_izq if self.sc._is_cell_meaningful(v)]
        if not no_none:
            return None
        n_str = sum(1 for v in no_none if isinstance(v, str) and str(v).strip())
        if n_str / len(no_none) < 0.6:
            return None

        # Leer como clave → valor (o clave → multiples valores)
        # La primera fila con dato en col A y sin dato en col B = título
        registros: list[dict[str, Any]] = []
        titulo: str | None = None
        c_b = c0 + 1

        # Detectar headers de columnas (primera fila con valor en col B)
        col_headers: list[str] = []
        primera_data = r0
        for r in range(r0, r1 + 1):
            va = self.sc[r, c0]
            vb = self.sc[r, c_b] if c_b <= c1 else None
            a_ok = self.sc._is_cell_meaningful(va)
            b_ok = self.sc._is_cell_meaningful(vb)
            if a_ok and not b_ok and not col_headers and titulo is None:
                titulo = str(va).strip()
                primera_data = r + 1
                continue
            if a_ok and b_ok and not col_headers:
                # Posible fila de headers
                todos_str = all(
                    isinstance(self.sc[r, c], str)
                    for c in range(c0, c1 + 1)
                    if self.sc._is_cell_meaningful(self.sc[r, c])
                )
                if todos_str and r == r0:
                    col_headers = [
                        str(self.sc[r, c]).strip() if self.sc._is_cell_meaningful(self.sc[r, c]) else f"col{c}"
                        for c in range(c0, c1 + 1)
                    ]
                    primera_data = r + 1
                    continue
            break

        col_labels = col_headers if col_headers else [
            f"col{i}" for i in range(c1 - c0 + 1)
        ]

        for r in range(primera_data, r1 + 1):
            va = self.sc[r, c0]
            if not self.sc._is_cell_meaningful(va):
                continue
            fila: dict[str, Any] = {"Nombre": str(va).strip()}
            for i, c in enumerate(range(c_b, c1 + 1)):
                lbl = col_labels[i + 1] if i + 1 < len(col_labels) else f"col{c}"
                v   = self.sc[r, c]
                fila[lbl] = v if self.sc._is_cell_meaningful(v) else None
            registros.append(fila)

        if not registros:
            return None

        df = pd.DataFrame(registros).set_index("Nombre")

        return TableRegion(
            id          = region_id,
            tipo        = "clave_valor",
            orientacion = "ambiguous",
            fila_inicio = r0 + 1,
            fila_fin    = r1 + 1,
            col_inicio  = c0 + 1,
            col_fin     = c1 + 1,
            fila_header = None,
            col_manager = None,
            col_metrica = None,
            col_headers = {},
            titulo      = titulo,
            data        = df,
        )

    # ── Helpers ─────────────────────────────────────────────────────────────

    @staticmethod
    def _looks_like_number(s: str) -> bool:
        """True si el string parece un valor numérico, no un nombre de columna."""
        s = s.strip()
        if not s:
            return False
        try:
            float(s)
            return True
        except ValueError:
            return False

    def _bbox_cols(
        self,
        r0: int,
        r1: int,
        c0_hint: int | None = None,
        c1_hint: int | None = None,
    ) -> tuple[int, int]:
        """
        Columnas mínima y máxima con datos en [r0, r1] base-0.
        Si se proveen hints c0/c1, restringe el cálculo al rango sugerido.
        """
        n_cols = self.sc.ncols
        if c0_hint is not None and c1_hint is not None:
            c0 = max(0, int(c0_hint))
            c1 = min(n_cols - 1, int(c1_hint))
            if c0 > c1:
                return (1, 0)
            sub = self.sc._mask_meaningful[r0 : r1 + 1, c0 : c1 + 1]
            cols = np.where(sub.any(axis=0))[0]
            if len(cols) == 0:
                return (1, 0)
            return (c0 + int(cols[0]), c0 + int(cols[-1]))

        sub  = self.sc._mask_meaningful[r0 : r1 + 1, :]
        cols = np.where(sub.any(axis=0))[0]
        return (int(cols[0]), int(cols[-1])) if len(cols) > 0 else (0, 0)

    def _detectar_titulo(
        self,
        r0: int, c0: int, c1: int,
        header_top_rows: list[int],
        header_left_cols: list[int],
    ) -> str | None:
        """Busca un título en la fila justo encima de r0 o en header_top."""
        if r0 > 0:
            for c in range(c0, c1 + 1):
                v = self.sc[r0 - 1, c]
                if self.sc._is_cell_meaningful(v):
                    s = str(v).strip()
                    if s and not _is_date_string(s):
                        return s
        if header_top_rows:
            r_titulo = min(header_top_rows)
            if header_left_cols:
                v = self.sc[r_titulo, min(header_left_cols)]
                if self.sc._is_cell_meaningful(v):
                    s = str(v).strip()
                    if s and not _is_date_string(s):
                        return s
        return None

    def _inferir_tipo(
        self,
        eje_filas: EjeInfo,
        eje_cols: EjeInfo,
        header_left_cols: list[int],
    ) -> TableType:
        """Infiere el TableType para compatibilidad con código existente."""
        n_grupos = eje_filas.n_niveles_grupo
        if n_grupos >= 1:
            return "manager_metrica"
        if eje_cols.tipo_hoja in ("fecha", "anio", "periodo"):
            return "solo_metrica"
        if not header_left_cols:
            return "tabla_generica"
        return "solo_metrica"

    def _cols_compat(
        self, header_left_cols: list[int]
    ) -> tuple[int | None, int | None]:
        """Retorna (col_manager_b1, col_metrica_b1) para compatibilidad."""
        if not header_left_cols:
            return None, None
        if len(header_left_cols) >= 2:
            return header_left_cols[0] + 1, header_left_cols[-1] + 1
        return None, header_left_cols[0] + 1

    # ── Mantener helpers de detección de fecha/columna para RegionDetector ─

    def _fila_con_mas_fechas(
        self, r0: int, r1: int, c0: int, c1: int,
    ) -> tuple[int, int]:
        """(fila_base0, n_fechas) de la fila con más datetimes/fecha-strings en el rango."""
        sub    = self.sc._mask_datetime[r0 : r1 + 1, c0 : c1 + 1]
        counts = sub.sum(axis=1)
        best   = int(counts.argmax())
        if best > 0 and counts[best] > 0:
            prev = int(counts[best - 1])
            if prev >= counts[best] * 0.30:
                best = best - 1
        return r0 + best, int(counts[best])

    def _columna_con_mas_fechas(
        self, r0: int, r1: int, c0: int, c1: int,
    ) -> tuple[int, int]:
        """(col_base0, n_fechas) de la columna con más fechas en el rango."""
        sub    = self.sc._mask_datetime[r0 : r1 + 1, c0 : c1 + 1]
        counts = sub.sum(axis=0)
        best   = int(counts.argmax())
        return c0 + best, int(counts[best])

    def _parse_generic_table(
        self,
        r0: int, r1: int, c0: int, c1: int, region_id: int,
    ) -> TableRegion:
        """Fallback universal: convierte cualquier bloque en tabla genérica."""
        mat = self.sc._mat[r0 : r1 + 1, c0 : c1 + 1]
        df  = pd.DataFrame(mat)

        # Saltar filas de título merged (todos valores no-None iguales)
        while df.shape[0] > 1:
            first_row = df.iloc[0].tolist()
            non_none = [v for v in first_row if v is not None]
            if non_none and all(isinstance(v, str) for v in non_none):
                unique = set(str(v).strip() for v in non_none)
                if len(unique) == 1:
                    # Fila de título merged → saltar
                    df = df.iloc[1:].reset_index(drop=True)
                    continue
            break

        if df.empty:
            df.columns = [f"col{i}" for i in range(df.shape[1])]
        else:
            first_row       = df.iloc[0].tolist()
            n_strings       = sum(isinstance(v, str) and v.strip() != "" for v in first_row)
            has_text_header = n_strings >= len(first_row) * 0.5

            if has_text_header:
                df.columns = [
                    str(v).strip() if (isinstance(v, str) and v.strip()) else f"col{i}"
                    for i, v in enumerate(first_row)
                ]
                df = df.iloc[1:].reset_index(drop=True)
            else:
                df.columns = [f"col{i}" for i in range(df.shape[1])]

        tipo: TableType = "tabla_generica"

        return TableRegion(
            id          = region_id,
            tipo        = tipo,
            orientacion = "ambiguous",
            fila_inicio = r0 + 1,
            fila_fin    = r1 + 1,
            col_inicio  = c0 + 1,
            col_fin     = c1 + 1,
            fila_header = None,
            col_manager = None,
            col_metrica = None,
            col_headers = {},
            titulo      = None,
            data        = df,
        )

    @staticmethod
    def _is_rotated_table(df: pd.DataFrame) -> bool:
        if df.shape[0] < 3 or df.shape[1] < 2:
            return False
        first_row  = df.iloc[0]
        frs = first_row.map(lambda x: isinstance(x, str) and x.strip() != "").mean()
        if frs < 0.8:
            return False
        data_rows = df.iloc[1:]
        data_num  = data_rows.map(
            lambda x: x is None or (isinstance(x, (int, float)) and not isinstance(x, bool))
        ).mean().mean()
        return bool(data_num >= 0.7)

    @staticmethod
    def _is_cross_table(df: pd.DataFrame) -> bool:
        if df.shape[0] < 2 or df.shape[1] < 3:
            return False
        first_col = df.iloc[:, 0].dropna()
        if len(first_col) == 0:
            return False
        fcs = first_col.map(lambda x: isinstance(x, str) and x.strip() != "").mean()
        if fcs < 0.8:
            return False
        other_cols = df.iloc[:, 1:]
        other_num  = other_cols.map(
            lambda x: x is None or (isinstance(x, (int, float)) and not isinstance(x, bool))
        ).mean().mean()
        return bool(other_num >= 0.7)

    @staticmethod
    def _normalize_cross_table(df: pd.DataFrame) -> pd.DataFrame:
        metric_col = df.columns[0]
        df_long    = df.melt(id_vars=[metric_col], var_name="Header", value_name="Valor")
        if str(metric_col) != "Métrica":
            df_long.rename(columns={metric_col: "Métrica"}, inplace=True)
        return df_long.reset_index(drop=True)

    def _parse_con_fechas_en_columna(
        self,
        r0: int, r1: int, c0: int, c1: int,
        col_h0: int, region_id: int,
    ) -> TableRegion | None:
        """Compatibilidad: parsea tabla con fechas en columna (delegado al nuevo pipeline)."""
        results = self.parse_multi(r0, r1, region_id)
        return results[0] if results else None


# ══════════════════════════════════════════════════════════════════════════════
# 6. TableAnalyzer  — orquesta y cachea resultados
# ══════════════════════════════════════════════════════════════════════════════

class TableAnalyzer:
    """
    Orquesta WorkbookLoader → SheetScanner → HeaderFirstDetector → TableParser.

    La hoja se escanea UNA sola vez en __init__.
    Todas las operaciones posteriores usan la matrix y masks cacheadas.

    Construcción estándar (desde archivo):
        analyzer = TableAnalyzer("archivo.xlsx", "Hoja1")

    Reutilizando scanner existente (no reabre el archivo):
        ws, merged = WorkbookLoader.load("archivo.xlsx", "Hoja1")
        scanner    = SheetScanner(ws, merged)
        analyzer   = TableAnalyzer.from_scanner(scanner)
    """

    def __init__(
        self,
        archivo: str | Path,
        hoja: str,
        **kwargs,  # compatibilidad con parámetros antiguos (ignorados)
    ) -> None:
        ws, merged = WorkbookLoader.load(archivo, hoja)
        self._setup(SheetScanner(ws, merged))

    @classmethod
    def from_scanner(
        cls,
        scanner: SheetScanner,
        **kwargs,  # compatibilidad
    ) -> "TableAnalyzer":
        """Construye desde un SheetScanner ya construido (sin reabrir archivo)."""
        instance = object.__new__(cls)
        instance._setup(scanner)
        return instance

    def _setup(self, scanner: SheetScanner) -> None:
        self.scanner  = scanner
        self.detector = ProjectionRegionDetector(scanner)  # REEMPLAZA: RegionDetector + ProjectionRegionDetector
        self.parser   = TableParser(scanner)
        self._tablas: list[TableRegion] | None = None

    # ── API interna ────────────────────────────────────────────────────────

    def todas_las_tablas(self) -> list[TableRegion]:
        """
        Detecta y parsea todas las sub-tablas. Resultado cacheado.

        Usa HeaderFirstDetector (método de referencia): detecta headers primero,
        luego delimita bloques de datos por gaps de filas vacías.
        Sin fusión ni expansión post-hoc — el detector ya entrega regiones limpias.
        """
        if self._tablas is not None:
            return self._tablas

        regiones = self.detector.detectar_regiones()

        tablas: list[TableRegion] = []
        id_counter = 1

        for r0, r1, c0, c1 in regiones:
            for region in self.parser.parse_multi(r0, r1, id_counter, c0=c0, c1=c1):
                region = region.__class__(
                    id          = id_counter,
                    tipo        = region.tipo,
                    orientacion = region.orientacion,
                    fila_inicio = region.fila_inicio,
                    fila_fin    = region.fila_fin,
                    col_inicio  = region.col_inicio,
                    col_fin     = region.col_fin,
                    fila_header = region.fila_header,
                    col_manager = region.col_manager,
                    col_metrica = region.col_metrica,
                    col_headers = region.col_headers,
                    titulo      = region.titulo,
                    data        = region.data,
                )
                tablas.append(region)
                id_counter += 1

        self._tablas = tablas
        return tablas

    def tabla_principal(self) -> TableRegion:
        """
        Tabla con mayor score = n_filas_datos × n_columnas_fecha.
        Más robusto que solo contar fechas: favorece tablas con muchas filas
        de datos Y muchos periodos temporales.
        """
        tablas = self.todas_las_tablas()
        if not tablas:
            # Fallback: envolver la hoja completa como tabla genérica.
            # Nunca falla, permite que el pipeline continúe aunque la detección
            # no haya encontrado ningún bloque estructurado.
            mat = self.scanner._mat
            df  = pd.DataFrame(mat)
            return TableRegion(
                id          = 1,
                tipo        = "tabla_generica",
                orientacion = "ambiguous",
                fila_inicio = 1,
                fila_fin    = self.scanner.nrows,
                col_inicio  = 1,
                col_fin     = self.scanner.ncols,
                fila_header = None,
                col_manager = None,
                col_metrica = None,
                col_headers = {},
                titulo      = "fallback_sheet",
                data        = df,
            )
        return max(tablas, key=lambda t: t.score)

    def fechas(self) -> list[dict[str, Any]]:
        """
        Fechas únicas de la hoja en orden cronológico.
        Reutiliza _mask_datetime ya cacheada: sin escaneo adicional.
        """
        posiciones = np.argwhere(self.scanner._mask_datetime)
        vistas: set[datetime.datetime] = set()
        result: list[dict[str, Any]]   = []

        for r, c in posiciones:
            v = self.scanner[int(r), int(c)]
            if isinstance(v, datetime.datetime) and v not in vistas:
                vistas.add(v)
                result.append({
                    "datetime": v,
                    "label":    self.scanner.fmt_fecha(v),
                    "row":      int(r) + 1,
                    "col":      int(c) + 1,
                })

        result.sort(key=lambda x: x["datetime"])
        return result


# ══════════════════════════════════════════════════════════════════════════════
# 7. API PÚBLICA  — 100% compatible con v3, v4 y v5
# ══════════════════════════════════════════════════════════════════════════════

# RAMA ELIMINADA: detectar_fechas standalone
# Duplicaba TableAnalyzer.fechas() con loop celda-por-celda O(n²).
# Usar TableAnalyzer.fechas() que opera sobre masks numpy precalculadas.


def detectar_todas_las_tablas(archivo: str, hoja: str) -> list[dict[str, Any]]:
    """
    Detecta y parsea TODAS las sub-tablas de la hoja en un solo escaneo.

    Args:
        archivo: Ruta al .xlsx
        hoja   : Nombre de la hoja

    Returns:
        Lista de dicts, uno por sub-tabla:
        {
          "id", "tipo", "orientacion", "score",
          "fila_inicio", "fila_fin", "col_inicio", "col_fin",
          "fila_header", "col_manager", "col_metrica",
          "col_headers", "titulo",
          "data": pd.DataFrame
        }

    Ejemplo:
        tablas = detectar_todas_las_tablas("archivo.xlsx", "Hoja1")
        for t in tablas:
            print(t["id"], t["tipo"], f"score={t['score']:.0f}")
    """
    return [t.as_dict() for t in TableAnalyzer(archivo, hoja).todas_las_tablas()]


def detectar_tabla(archivo: str, hoja: str) -> dict[str, Any]:
    """
    Devuelve la tabla PRINCIPAL de la hoja (mayor score).

    score = n_filas_datos × n_columnas_fecha

    Args:
        archivo: Ruta al .xlsx
        hoja   : Nombre de la hoja

    Returns:
        Dict con claves:
          "id", "tipo", "orientacion", "score",
          "fila_inicio", "fila_fin", "col_inicio", "col_fin",
          "fila_header", "col_manager", "col_metrica",
          "col_headers", "titulo",
          "data" → pd.DataFrame con MultiIndex (Manager, Métrica)

    Ejemplo:
        tabla = detectar_tabla("archivo.xlsx", "Hoja1")
        print(tabla["score"])        # 1620.0 = 45 filas × 36 fechas
        print(tabla["orientacion"])  # "column"
        print(tabla["data"].head())
    """
    return TableAnalyzer(archivo, hoja).tabla_principal().as_dict()


def extraer_columna(tabla: dict[str, Any], fecha: str) -> pd.Series:
    """
    Extrae todos los valores de una columna (fecha/periodo).

    Args:
        tabla : Resultado de detectar_tabla() o elemento de detectar_todas_las_tablas().
        fecha : Etiqueta de columna, ej. 'ene-26', 'TOTAL'.

    Returns:
        pd.Series con MultiIndex (Manager, Métrica) o (Métrica,).

    Raises:
        KeyError si la columna no existe (muestra disponibles).
    """
    df = tabla["data"]
    if fecha not in df.columns:
        raise KeyError(
            f"Columna '{fecha}' no encontrada. Disponibles: {list(df.columns)}"
        )
    return df[fecha]


def extraer_fila(tabla: dict[str, Any], metrica: str) -> pd.DataFrame:
    """
    Extrae todos los valores de una métrica para todos los managers.
    Búsqueda parcial e insensible a mayúsculas.

    Args:
        tabla   : Resultado de detectar_tabla().
        metrica : Nombre de la métrica, ej. 'AUM', 'ROA', 'FN'.

    Returns:
        pd.DataFrame con índice Manager y columnas de fecha.

    Raises:
        KeyError si la métrica no existe (muestra disponibles).
    """
    df    = tabla["data"]
    names = df.index.names or []
    nivel = "Métrica" if "Métrica" in names else (names[-1] if names else None)
    if nivel is None:
        raise ValueError("El DataFrame no tiene índice nombrado.")
    mask = df.index.get_level_values(nivel).str.contains(
        metrica, case=False, na=False
    )
    res = df[mask]
    if res.empty:
        disp = df.index.get_level_values(nivel).unique().tolist()
        raise KeyError(
            f"Métrica '{metrica}' no encontrada. Disponibles: {disp}"
        )
    return res.droplevel(nivel) if "Manager" in names else res


def crear_tabla(tabla: dict[str, Any]) -> pd.DataFrame:
    """
    Convierte la tabla al formato LARGO con columna "Fecha" en orden cronológico.

    Estructura del resultado:
        Manager | Métrica | Fecha   | Valor
        --------|---------|---------|------
        First T | AUM     | ene-24  | 12.0
        First T | AUM     | feb-24  | None

    Args:
        tabla: Resultado de detectar_tabla() o elemento de detectar_todas_las_tablas().

    Returns:
        pd.DataFrame con columnas [Manager (si existe), Métrica, Fecha, Valor].
    """
    df = tabla["data"]
    if df is None or df.empty:
        return pd.DataFrame()

    col_headers = tabla.get("col_headers", {})
    if col_headers:
        cols_ord    = [col_headers[c] for c in sorted(col_headers.keys())]
        cols_fechas = [c for c in cols_ord if c in df.columns]
    else:
        cols_fechas = list(df.columns)

    df_reset = df.reset_index()
    id_vars  = [c for c in df_reset.columns if c not in cols_fechas]

    df_largo = df_reset.melt(
        id_vars=id_vars,
        value_vars=cols_fechas,
        var_name="Fecha",
        value_name="Valor",
    )

    orden_map        = {lbl: i for i, lbl in enumerate(cols_fechas)}
    df_largo["_ord"] = df_largo["Fecha"].map(orden_map)
    df_largo = (
        df_largo
        .sort_values(id_vars + ["_ord"])
        .drop(columns="_ord")
        .reset_index(drop=True)
    )
    return df_largo


def replicar_tabla(
    tabla: dict[str, Any],
    archivo_salida: str,
    hoja_salida: str = "Tabla_Limpia",
    formato: str = "ancho",
) -> None:
    """
    Exporta la tabla a un Excel limpio sin filas vacías ni separaciones.

    Args:
        tabla          : Resultado de detectar_tabla().
        archivo_salida : Ruta del .xlsx de salida.
        hoja_salida    : Nombre de la hoja destino.
        formato        : "ancho" (una col por fecha) | "largo" (col Fecha + Valor).
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    df_export = crear_tabla(tabla) if formato == "largo" else tabla["data"].reset_index()
    if df_export.empty:
        raise ValueError("La tabla detectada está vacía.")

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = hoja_salida

    hfont  = Font(bold=True, color="FFFFFF")
    hfill  = PatternFill(fill_type="solid", fgColor="2F5496")
    halign = Alignment(horizontal="center")

    for ci, h in enumerate(df_export.columns, 1):
        cell = ws_out.cell(row=1, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign

    for ri, row in enumerate(df_export.itertuples(index=False), 2):
        for ci, v in enumerate(row, 1):
            ws_out.cell(row=ri, column=ci, value=v)

    for col in ws_out.columns:
        w = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws_out.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    wb_out.save(archivo_salida)
    print(f"✅ Exportado ({formato}) → {Path(archivo_salida).resolve()}")


# ── Compatibilidad hacia atrás ─────────────────────────────────────────────

# RAMA ELIMINADA: expandir_merged_cells
# Redundante: SheetScanner expande merges directamente en la matrix numpy.
# Si código externo necesita esta función, usar scanner._mat directamente.


def diagnosticar(archivo: str, hoja: str) -> None:
    """
    Imprime un diagnóstico detallado de lo que el código ve en una hoja.

    Úsalo cuando el código no detecta las fechas correctamente en un Excel nuevo.
    Muestra: hojas disponibles, filas con fechas, regiones detectadas, tablas
    parseadas y una muestra de cada tabla.

    Args:
        archivo : Ruta al .xlsx
        hoja    : Nombre de la hoja a diagnosticar

    Ejemplo:
        import excel_analyzer as ea
        ea.diagnosticar("mi_archivo.xlsx", "Hoja1")
    """
    import openpyxl as _opx

    SEP = "=" * 65

    # ── 0. Hojas disponibles ─────────────────────────────────────────
    wb = _opx.load_workbook(archivo, data_only=True, read_only=True)
    print(f"\n{SEP}")
    print(f"  ARCHIVO: {Path(archivo).name}")
    print(f"  Hojas disponibles: {wb.sheetnames}")
    print(SEP)
    wb.close()

    # ── 1. Scanner ────────────────────────────────────────────────────
    ws, merged = WorkbookLoader.load(archivo, hoja)
    sc = SheetScanner(ws, merged)
    print(f"\n📐 Hoja: {hoja!r}  ({sc.nrows} filas × {sc.ncols} cols)")
    print(f"   Celdas con datos : {sc._mask_meaningful.sum():,}")
    print(f"   Celdas con fechas: {sc._mask_datetime.sum()}")
    print(f"   Orientación      : {sc.orientation!r}")

    # ── 2. Filas con más fechas ───────────────────────────────────────
    print(f"\n📅 Filas con fechas (top 5):")
    counts = sc._mask_datetime.sum(axis=1)
    top_rows = sorted(enumerate(counts), key=lambda x: -x[1])[:5]
    for r_idx, n in top_rows:
        if n == 0:
            break
        # Mostrar los primeros valores de fecha en esa fila
        date_vals = []
        for c in range(sc.ncols):
            v = sc._mat[r_idx, c]
            if sc._mask_datetime[r_idx, c] and v is not None:
                date_vals.append(sc.fmt_fecha(v))
                if len(date_vals) >= 4:
                    date_vals.append("…")
                    break
        print(f"   fila {r_idx+1:3d}: {n} fechas → {date_vals}")

    # ── 3. Regiones detectadas ─────────────────────────────────────────
    detector = ProjectionRegionDetector(sc)
    parser   = TableParser(sc)
    analyzer = TableAnalyzer.from_scanner(sc)
    raw      = detector.detectar_regiones()

    print(f"\n🔍 Regiones detectadas: {len(raw)}")
    for i, (r0, r1, c0, c1) in enumerate(raw):
        if c0 is None:
            c0_d, c1_d = parser._bbox_cols(r0, r1)
        else:
            c0_d, c1_d = c0, c1
        first_val = next((sc._mat[r0, c] for c in range(sc.ncols) if sc._mat[r0, c] is not None), None)
        n_dt = int(sc._mask_datetime[r0: r1 + 1, :].sum(axis=1).max()) if r1 >= r0 else 0
        mark = " ← tiene fechas" if n_dt >= 2 else ""
        print(f"   [{i+1}] filas {r0+1}–{r1+1}  cols {c0_d+1}–{c1_d+1}  "
              f"fechas={n_dt}  primera={repr(str(first_val))[:30]}{mark}")

    # ── 4. Tablas parseadas ───────────────────────────────────────────
    tablas = analyzer.todas_las_tablas()
    print(f"\n🗂  {len(tablas)} tabla(s) parseada(s):")
    for t in tablas:
        df = t.data
        if df is None:
            print(f"   [{t.id}] {t.tipo}  (sin datos)")
            continue
        print(f"\n   [{t.id}] tipo={t.tipo!r}  shape={df.shape}  score={t.score:.0f}")
        if df.columns.tolist():
            print(f"        columnas ({len(df.columns)}): "
                  f"{list(df.columns)[:6]}{'…' if len(df.columns) > 6 else ''}")
        if df.index is not None and len(df.index) > 0:
            vals = df.index.get_level_values(-1).unique().tolist()
            print(f"        métricas ({len(vals)}): "
                  f"{vals[:5]}{'…' if len(vals) > 5 else ''}")
        print(f"        muestra:")
        print("        " + df.head(2).to_string(max_cols=5).replace("\n", "\n        "))

    # ── 5. Posibles problemas ─────────────────────────────────────────
    print(f"\n⚠️  Diagnóstico rápido:")
    if sc._mask_datetime.sum() == 0:
        print("   ❌ NO SE ENCONTRARON FECHAS en esta hoja.")
        print("      Causa probable: las fechas están en un formato no reconocido.")
        print("      Muestra de valores en fila 1:")
        row0 = [sc._mat[0, c] for c in range(min(sc.ncols, 15)) if sc._mat[0, c] is not None]
        for v in row0[:8]:
            print(f"        {repr(v)}")
    elif all(t.tipo in ('clave_valor', 'tabla_generica') for t in tablas):
        print("   ⚠️  Solo se detectaron tablas sin fechas (clave_valor / tabla_generica).")
        print("      Las fechas existen pero puede que estén en una fila con baja densidad.")
        best_row = int(counts.argmax())
        print(f"      La fila con más fechas es la fila {best_row+1} ({int(counts[best_row])} fechas).")
    else:
        print("   ✅ Detección normal. Si faltan métricas, revisa los shapes arriba.")

    print()


# ══════════════════════════════════════════════════════════════════════════════
# DEMO
# ══════════════════════════════════════════════════════════════════════════════


def subdividir_por_anio(
    tabla: dict[str, Any],
    col_anio: str | None = None,
) -> list[dict[str, Any]]:
    """
    Divide una tabla detectada en sub-tablas por año.

    Opera sobre el resultado de ``detectar_tabla()`` o un elemento de
    ``detectar_todas_las_tablas()``.  No modifica la detección original —
    es un post-procesamiento independiente que puedes aplicar selectivamente.

    Estrategia
    ----------
    Busca columnas cuyo **header** sea un año válido (entero o string "20XX"
    en rango 2000-2100) en el DataFrame ya construido.  Esas columnas actúan
    como marcadores de año: todas las columnas desde el marcador anterior
    (exclusive) hasta el marcador actual (inclusive) pertenecen a ese año.

    Si no se encuentran marcadores de año, retorna una lista con la tabla
    original sin modificar.

    Parámetros
    ----------
    tabla    : dict devuelto por ``detectar_tabla()`` o elemento de
               ``detectar_todas_las_tablas()``.
    col_anio : Nombre de columna a usar como marcador (opcional).
               Si se pasa, solo esa columna se usa como separador de año.
               Si es None, se detectan automáticamente todas las columnas
               cuyo nombre sea un año 20XX.

    Retorna
    -------
    Lista de dicts con la misma estructura que ``tabla``.  Cada sub-tabla
    tiene un campo ``"titulo"`` con el año (str) y solo las columnas de
    ese año.  El campo ``"data"`` es un nuevo DataFrame.

    Ejemplo
    -------
    ::

        import excel_analyzer as ea

        tablas = ea.detectar_todas_las_tablas("CAPELLAWM.xlsx", "Tablero Proyección")
        # tablas[0] tiene cols: ['2022', 'Ago22-Jun23', 'Q1', 'Q2', 'Q3', 'Q4', '2023']
        por_anio = ea.subdividir_por_anio(tablas[0])
        # → [{'titulo': '2022', ...}, {'titulo': '2023', ...}]
    """
    df: pd.DataFrame = tabla["data"]
    cols = list(df.columns)

    def _es_anio_col(nombre: str) -> int | None:
        s = str(nombre).strip()
        if re.match(r"^\d{4}$", s):
            yr = int(s)
            if 2000 <= yr <= 2100:
                return yr
        return None

    if col_anio is not None:
        # Usar solo la columna especificada como marcador
        if col_anio not in cols:
            return [tabla]
        marcadores = {col_anio: int(str(col_anio).strip())}
    else:
        marcadores = {
            c: _es_anio_col(c)
            for c in cols
            if _es_anio_col(c) is not None
        }

    if not marcadores:
        return [tabla]

    # Ordenar marcadores por posición en cols
    marcadores_ordenados = sorted(marcadores.keys(), key=lambda c: cols.index(c))

    resultado: list[dict[str, Any]] = []
    prev_idx = 0  # inicio del sub-bloque actual (inclusivo)

    for marc in marcadores_ordenados:
        marc_idx = cols.index(marc)
        # Sub-bloque: cols[prev_idx .. marc_idx] inclusive
        sub_cols = cols[prev_idx : marc_idx + 1]
        if sub_cols:
            sub_df = df[sub_cols]
            sub_tabla = dict(tabla)
            sub_tabla["data"] = sub_df
            sub_tabla["titulo"] = str(marcadores[marc])
            sub_tabla["col_headers"] = {
                k: v for k, v in tabla.get("col_headers", {}).items()
                if v in sub_cols
            }
            resultado.append(sub_tabla)
        prev_idx = marc_idx + 1

    # Columnas después del último marcador
    rest_cols = cols[prev_idx:]
    if rest_cols:
        sub_df = df[rest_cols]
        sub_tabla = dict(tabla)
        sub_tabla["data"] = sub_df
        sub_tabla["titulo"] = tabla.get("titulo")
        sub_tabla["col_headers"] = {
            k: v for k, v in tabla.get("col_headers", {}).items()
            if v in rest_cols
        }
        resultado.append(sub_tabla)

    return resultado if resultado else [tabla]


# ══════════════════════════════════════════════════════════════════════════════
# reemplazar_valores — inyectar fórmulas en una columna/métrica de la tabla
# ══════════════════════════════════════════════════════════════════════════════

def reemplazar_valores(
    tabla: dict[str, Any],
    metrica: str,
    formula: str,
    columna: str | None = None,
    inplace: bool = False,
) -> dict[str, Any]:
    """
    Reemplaza los valores de una métrica (fila) con el resultado de una
    fórmula que puede referenciar otras métricas por nombre.

    Diseñado para el caso de uso donde el parser detectó correctamente todas
    las columnas y métricas de la tabla, y se quiere calcular una columna
    derivada — por ejemplo, "Utilidades = Ingresos T - Costos T".

    Parámetros
    ----------
    tabla   : dict resultado de detectar_tabla() o detectar_todas_las_tablas()
    metrica : Nombre de la métrica (fila del índice) que se va a reemplazar.
              Acepta coincidencia fuzzy si el nombre exacto no existe.
    formula : Expresión en Python/pandas que define el nuevo valor.
              Las referencias a otras métricas van en el formato: T['nombre'].
              Columnas individuales van como C['nombre'].
              Ejemplos:
                "T['Ingresos'] - T['Costos']"
                "T['Ingresos'] * 0.15"
                "T['Ingresos T'] - T['Costos T']"
                "C['dic-26'] * 12"   ← solo modifica esa columna
    columna : Si se especifica, la fórmula solo se aplica a esa columna.
              Si es None (default), se aplica a todas las columnas de datos.
    inplace : Si True, modifica tabla['data'] in-place. Default False.

    Retorna
    -------
    dict con la misma estructura que ``tabla`` pero con ``data`` actualizado.
    El DataFrame tiene los valores originales salvo la fila de la métrica
    indicada, que ahora tiene los valores calculados.

    Errores
    -------
    ValueError si la métrica o columna no se encuentran (ni con fuzzy).
    ValueError si la fórmula tiene un error de sintaxis o referencia.

    Ejemplos
    --------
    ::

        import excel_analyzer as ea

        tablas = ea.detectar_todas_las_tablas("PRESUPUESTO 2026.xlsx", "Generación EV-2026")
        tabla  = tablas[0]

        # Calcular Utilidades como diferencia de dos métricas detectadas
        tabla2 = ea.reemplazar_valores(
            tabla,
            metrica = "Utilidades",
            formula = "T['Ingresos T'] - T['Costos T']",
        )

        # Sobrescribir solo un mes específico
        tabla3 = ea.reemplazar_valores(
            tabla,
            metrica = "Proyección",
            formula = "T['AUM'] * 1.05",
            columna = "dic-26",
        )
    """
    import difflib

    df_orig: pd.DataFrame = tabla["data"]
    df = df_orig if inplace else df_orig.copy()

    # ── Resolver nombre de métrica (fuzzy si no existe exacto) ─────────────
    def _resolver_fila(nombre: str) -> Any:
        """Retorna el label real del índice más cercano a `nombre`."""
        # MultiIndex: buscar en el nivel más interno (Métrica)
        if isinstance(df.index, pd.MultiIndex):
            # Nivel más profundo = métricas
            nivel_metrica = df.index.nlevels - 1
            valores_nivel = df.index.get_level_values(nivel_metrica).astype(str).tolist()
        else:
            valores_nivel = df.index.astype(str).tolist()

        # Coincidencia exacta (case-insensitive)
        nombre_low = nombre.strip().lower()
        for v in valores_nivel:
            if v.lower() == nombre_low:
                return v

        # Fuzzy matching
        matches = difflib.get_close_matches(nombre, valores_nivel, n=1, cutoff=0.6)
        if matches:
            return matches[0]

        raise ValueError(
            f"reemplazar_valores: métrica '{nombre}' no encontrada en la tabla.\n"
            f"Métricas disponibles: {valores_nivel[:20]}"
        )

    def _resolver_col(nombre: str) -> Any:
        """Retorna el label real de columna más cercano a `nombre`."""
        cols_str = [str(c) for c in df.columns]
        nombre_low = nombre.strip().lower()
        for i, c in enumerate(cols_str):
            if c.lower() == nombre_low:
                return df.columns[i]
        matches = difflib.get_close_matches(nombre, cols_str, n=1, cutoff=0.6)
        if matches:
            return df.columns[cols_str.index(matches[0])]
        raise ValueError(
            f"reemplazar_valores: columna '{nombre}' no encontrada.\n"
            f"Columnas disponibles: {cols_str[:20]}"
        )

    label_metrica = _resolver_fila(metrica)

    # ── Construir namespace de evaluación ──────────────────────────────────
    # T['nombre'] = Series de la métrica nombre (todas las columnas)
    # C['nombre'] = columna como Series (todas las filas)
    # Operaciones aritméticas de pandas funcionan naturalmente.

    class _MetricaAccessor:
        """
        T['Ingresos'] → Series con los valores de esa métrica.

        Para tablas con MultiIndex (Grupo × Métrica), xs() puede devolver
        un DataFrame si la misma métrica aparece en múltiples grupos.
        En ese caso se alinea con el contexto:
          - Si se está asignando a la misma métrica (mismo patrón de filas),
            se retorna el DataFrame tal cual para que la asignación fila-a-
            fila funcione.
          - Para operaciones aritméticas en la fórmula (T[m1] - T[m2]),
            pandas alinea DataFrames automáticamente por índice.
        """
        def __getitem__(self, nombre: str) -> "pd.Series | pd.DataFrame":
            lbl = _resolver_fila(nombre)
            if isinstance(df.index, pd.MultiIndex):
                nivel = df.index.nlevels - 1
                mask  = df.index.get_level_values(nivel) == lbl
                sub   = df[mask]
                # Una sola fila → Series; múltiples → DataFrame
                return sub.iloc[0] if len(sub) == 1 else sub
            return df.loc[lbl]

    class _ColAccessor:
        """C['ene-24'] → Series con los valores de esa columna."""
        def __getitem__(self, nombre: str) -> pd.Series:
            lbl = _resolver_col(nombre)
            return df[lbl]

    T = _MetricaAccessor()
    C = _ColAccessor()
    import numpy as _np

    namespace = {
        "T": T,
        "C": C,
        "np": _np,
        "pd": pd,
    }

    # ── Evaluar fórmula ─────────────────────────────────────────────────────
    try:
        resultado: pd.Series = eval(formula, {"__builtins__": {}}, namespace)  # noqa: S307
    except Exception as exc:
        raise ValueError(
            f"reemplazar_valores: error al evaluar fórmula '{formula}'.\n"
            f"Error: {exc}\n"
            f"Ejemplo válido: \"T['Ingresos T'] - T['Costos T']\""
        ) from exc

    # ── Aplicar resultado al DataFrame ──────────────────────────────────────
    if isinstance(df.index, pd.MultiIndex):
        nivel = df.index.nlevels - 1
        mask  = df.index.get_level_values(nivel) == label_metrica
        filas_a_modificar = df.index[mask]
    else:
        mask  = df.index == label_metrica
        filas_a_modificar = df.index[mask]

    if len(filas_a_modificar) == 0:
        raise ValueError(f"reemplazar_valores: no se encontraron filas para '{label_metrica}'")

    if columna is not None:
        # Solo modificar una columna específica
        label_col = _resolver_col(columna)
        for fila_idx in filas_a_modificar:
            if isinstance(resultado, pd.DataFrame):
                # MultiIndex: fila_idx identifica el grupo+métrica
                try:
                    val = resultado.loc[fila_idx, label_col]
                except KeyError:
                    val = float(resultado[label_col].iloc[0])
            elif isinstance(resultado, pd.Series):
                val = resultado[label_col] if label_col in resultado.index else float(resultado.iloc[0])
            else:
                val = float(resultado)
            df.at[fila_idx, label_col] = val
    else:
        # Modificar todas las columnas
        if isinstance(resultado, pd.DataFrame):
            # DataFrame: las filas del resultado deben corresponder a filas_a_modificar
            # Iterar en paralelo (mismo orden de grupos)
            for fila_idx, (_, row_res) in zip(filas_a_modificar, resultado.iterrows()):
                for col in df.columns:
                    if col in row_res.index:
                        df.at[fila_idx, col] = row_res[col]
        elif isinstance(resultado, pd.Series):
            for fila_idx in filas_a_modificar:
                for col in df.columns:
                    if col in resultado.index:
                        df.at[fila_idx, col] = resultado[col]
        else:
            for fila_idx in filas_a_modificar:
                df.loc[fila_idx] = float(resultado)

    resultado_tabla = dict(tabla)
    resultado_tabla["data"] = df
    return resultado_tabla


# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import time, warnings
    warnings.filterwarnings("ignore")
    pd.set_option("display.max_columns", 7)
    pd.set_option("display.width", 120)

    ARCHIVO = "PRESUPUESTO_2026.xlsx"
    HOJA    = "Generación EV-2026"

    print("=" * 70)
    print(f"  {ARCHIVO}  /  {HOJA}")
    print("=" * 70)

    # ── Benchmark de construcción ──────────────────────────────────────────
    t0 = time.perf_counter()
    analyzer = TableAnalyzer(ARCHIVO, HOJA)
    t1 = time.perf_counter()
    sc = analyzer.scanner
    print(f"\n⚡ SheetScanner v6: {(t1-t0)*1000:.1f} ms  "
          f"({sc.nrows}×{sc.ncols} = {sc.nrows*sc.ncols:,} celdas)")
    print(f"   Orientación     : {sc.orientation!r}")
    print(f"   Celdas con datos: {sc._mask_meaningful.sum():,} "
          f"({sc._mask_meaningful.mean()*100:.0f}%)")
    print(f"   Datetimes       : {sc._mask_datetime.sum()}")

    # 0. Fechas
    fch = analyzer.fechas()
    print(f"\n📅 {len(fch)} fechas: "
          f"{[f['label'] for f in fch[:3]]} … {[f['label'] for f in fch[-3:]]}")

    # 1. Todas las tablas
    todas = analyzer.todas_las_tablas()
    print(f"\n🗂  {len(todas)} sub-tablas:")
    for t in todas:
        print(f"   [{t.id}] {t.tipo:20s}  score={t.score:>8.0f}  "
              f"filas {t.fila_inicio:3d}→{t.fila_fin:3d}  shape={t.data.shape}"
              + (f"  título={t.titulo!r}" if t.titulo else ""))

    # 2. Tabla principal
    principal = analyzer.tabla_principal()
    df        = principal.data
    print(f"\n🏆 Tabla principal → score={principal.score:.0f}  shape={df.shape}")
    if "Manager" in df.index.names:
        print(f"   Managers : {df.index.get_level_values('Manager').unique().tolist()}")
        print(f"   Métricas : {df.index.get_level_values('Métrica').unique().tolist()}")
    print()
    print(df.head(8).to_string(max_cols=7))

    # Benchmark API pública
    t2 = time.perf_counter()
    tabla = detectar_tabla(ARCHIVO, HOJA)
    t3 = time.perf_counter()
    print(f"\n⚡ detectar_tabla() API: {(t3-t2)*1000:.1f} ms")

    # from_scanner
    t4 = time.perf_counter()
    ws_r, mr = WorkbookLoader.load(ARCHIVO, HOJA)
    sc2  = SheetScanner(ws_r, mr)
    a2   = TableAnalyzer.from_scanner(sc2)
    _ = a2.tabla_principal()
    t5 = time.perf_counter()
    print(f"⚡ from_scanner:         {(t5-t4)*1000:.1f} ms  (scanner reutilizable)")

    # extraer_columna
    primera = list(tabla["col_headers"].values())[0]
    print(f"\n── extraer_columna(tabla, '{primera}') ──")
    print(extraer_columna(tabla, primera).dropna().to_string())

    # extraer_fila
    print(f"\n── extraer_fila(tabla, 'AUM') ──")
    print(extraer_fila(tabla, "AUM").to_string(max_cols=7))

    # crear_tabla
    df_l = crear_tabla(tabla)
    print(f"\n── crear_tabla → shape={df_l.shape} ──")
    print(df_l.dropna(subset=["Valor"]).head(10).to_string(index=False))

    # Exportar
    replicar_tabla(tabla, "tabla_limpia_ancho.xlsx", formato="ancho")
    replicar_tabla(tabla, "tabla_limpia_largo.xlsx",  formato="largo")