"""
excel_analyzer/query_engine.py
================================
Motor de consulta de alto nivel para Excel Analyzer.

Construido sobre las funciones de detección existentes sin modificarlas.
Sólo importa de ._core — nunca escribe en él.

DIAGRAMA DE FLUJO
=================

    excel_query(archivo, metric, sheet, table, where, mode)
          │
          ▼
    _resolve_sheets()        ← sheet=N (1-based) → lista de nombres de hoja
          │
          ▼ por cada hoja
    detectar_todas_las_tablas()   (de ._core)
          │
          ▼
    _select_table()          ← table=N filtra por id; None → todas
          │
          ▼
    _find_best_metric()      ← fuzzy matching + grupos de sinónimos
          │
          ▼
    _filter_metric()         ← df[mask] sobre el índice Métrica
          │
          ▼
    _apply_where()           ← reset_index() + df.query(where)
          │
          ┌──────────────────┴──────────────────┐
          ▼ mode="best"                          ▼ mode="merge"
    max(score) → un solo df             pd.concat + columna Hoja

SISTEMA DE SINÓNIMOS
====================
Se usan grupos bidireccionales en lugar de un dict query→expansión.
El motivo: expandir "assets" → "assets under management" produce scores
~0.08 contra métricas cortas como "AUM". Con grupos, el candidato "aum"
del mismo grupo compara directamente y obtiene score 1.0.

    query "assets"  → candidatos ["assets", "aum", "assets under management"]
                     → max_score("aum", "AUM") = 1.0  ✓

    query "return on assets" → candidatos ["return on assets", "roa"]
                             → max_score("roa", "ROA") = 1.0  ✓

UMBRAL EN MODE=MERGE
====================
En mode="merge" se usa merge_min_score = max(min_score, 0.6) para evitar
falsos positivos. Ejemplo: "AUM" puntúa 0.54 en CA_Resultados ("Gasto de
venta" — match incorrecto). Con threshold 0.6 esa hoja se excluye.
El usuario puede forzar la inclusión bajando min_score explícitamente.
"""

from __future__ import annotations

from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from ._core import detectar_todas_las_tablas

# ══════════════════════════════════════════════════════════════════════════════
# MOTOR SEMÁNTICO — sinónimos y fuzzy matching
# ══════════════════════════════════════════════════════════════════════════════

# Grupos de sinónimos: términos del mismo grupo deben encontrar la misma métrica.
# Se usan solo alias genéricos, no nombres de métricas específicos del archivo.
_SYNONYM_GROUPS: list[set[str]] = [
    {"aum", "assets", "assets under management"},
    {"roa", "return on assets"},
    {"roi", "return on investment"},
    {"rev", "revenue", "ingreso", "ingresos"},
    {"profit", "net profit", "utilidad", "ganancia"},
    {"mgmt fee", "management fee"},
    {"fn", "fn/renta", "flujo neto", "flujo"},
    {"placement", "placement fee"},
]

# Índice invertido: término_normalizado → grupo (construido una vez al importar)
_SYN_IDX: dict[str, set[str]] = {
    term.lower(): group
    for group in _SYNONYM_GROUPS
    for term in group
}


def _normalize_text(text: str) -> str:
    """Normaliza texto para comparación: minúsculas y strip."""
    return str(text).lower().strip()


def _similarity(a: str, b: str) -> float:
    """
    Similitud entre dos strings [0.0, 1.0] con SequenceMatcher.
    Insensible a mayúsculas y espacios extra.
    """
    return SequenceMatcher(
        None, _normalize_text(a), _normalize_text(b)
    ).ratio()


def _get_synonym_candidates(query: str) -> list[str]:
    """
    Devuelve todos los candidatos de búsqueda para el query.

    Si el query pertenece a un grupo de sinónimos (directamente o como
    subconjunto de algún término del grupo), devuelve todos los términos
    del grupo. Si no hay grupo, devuelve [query].

    Ejemplos:
        "assets"           → ["assets", "aum", "assets under management"]
        "return on assets" → ["return on assets", "roa"]
        "cobro"            → ["cobro"]   (sin grupo: búsqueda directa)
    """
    q = _normalize_text(query)

    # Búsqueda exacta en el índice
    if q in _SYN_IDX:
        return list(_SYN_IDX[q])

    # Búsqueda por contención: "assets mgmt" ⊆ "assets under management"
    for term, group in _SYN_IDX.items():
        if q in term or term in q:
            return list(group)

    return [q]


def _find_best_metric(
    tablas: list[dict[str, Any]],
    query: str,
) -> tuple[dict[str, Any] | None, str | None, float]:
    """
    Encuentra la métrica más similar al query en una lista de tablas.

    Para cada métrica en cada tabla calcula:
        score = max(similarity(candidato, métrica) for candidato in candidatos)
    donde candidatos viene de _get_synonym_candidates(query).

    Args:
        tablas: Lista de dicts de detectar_todas_las_tablas().
        query:  Texto libre, ej. "AUM", "assets", "return on assets".

    Returns:
        (tabla_dict, nombre_métrica, score) del mejor match,
        o (None, None, 0.0) si tablas está vacía o no hay índice Métrica.
    """
    candidates = _get_synonym_candidates(query)

    best_table:  dict[str, Any] | None = None
    best_metric: str | None            = None
    best_score:  float                 = 0.0

    for t in tablas:
        df = t.get("data")
        if df is None:
            continue
        names = df.index.names or []
        nivel = "Métrica" if "Métrica" in names else None
        if nivel is None:
            continue
        for m in df.index.get_level_values(nivel).unique():
            score = max(_similarity(c, str(m)) for c in candidates)
            if score > best_score:
                best_score  = score
                best_metric = str(m)
                best_table  = t

    return best_table, best_metric, best_score


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS INTERNOS DEL QUERY ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def _resolve_sheets(archivo: str, sheet: int | None) -> list[str]:
    """
    Convierte sheet (1-based int | None) en lista de nombres de hoja.

    sheet=None → todas las hojas del workbook.
    sheet=N    → [hoja N] validando que N esté en rango.

    Abre el workbook en read_only solo para leer sheetnames y lo cierra
    inmediatamente para no bloquear el archivo.

    Raises:
        ValueError con lista de hojas disponibles si sheet está fuera de rango.
    """
    wb = openpyxl.load_workbook(str(archivo), read_only=True, data_only=True)
    all_sheets = wb.sheetnames
    wb.close()

    if sheet is None:
        return list(all_sheets)

    if not (1 <= sheet <= len(all_sheets)):
        available = ", ".join(
            f"{i+1}={s!r}" for i, s in enumerate(all_sheets)
        )
        raise ValueError(
            f"sheet={sheet} fuera de rango. "
            f"El archivo tiene {len(all_sheets)} hoja(s): {available}."
        )

    return [all_sheets[sheet - 1]]


def _select_table(
    tablas: list[dict[str, Any]],
    table: int | None,
    hoja: str,
) -> list[dict[str, Any]]:
    """
    Filtra la lista de tablas por id (1-based) si se especifica table.

    table=None → devuelve todas las tablas sin filtrar.
    table=N    → devuelve [tabla con id==N].

    Raises:
        ValueError con IDs disponibles si table no existe en la hoja.
    """
    if table is None:
        return tablas

    matching = [t for t in tablas if t["id"] == table]
    if not matching:
        ids = [t["id"] for t in tablas]
        raise ValueError(
            f"table={table} no existe en la hoja '{hoja}'. "
            f"IDs disponibles: {ids}."
        )
    return matching


def _filter_metric(
    tabla: dict[str, Any],
    metric_name: str,
) -> pd.DataFrame:
    """
    Filtra el DataFrame de una tabla a las filas cuyo índice Métrica
    coincide exactamente con metric_name.

    Returns:
        Copia del DataFrame filtrado con el índice original preservado.
    """
    df    = tabla["data"]
    nivel = "Métrica" if "Métrica" in df.index.names else df.index.names[-1]
    mask  = df.index.get_level_values(nivel) == metric_name
    return df[mask].copy()


def _apply_where(df: pd.DataFrame, where: str | None) -> pd.DataFrame:
    """
    Aplica un filtro SQL-like al DataFrame usando pandas df.query().

    Convierte el índice a columnas con reset_index() antes del filtro para
    que columnas como Manager, Métrica sean accesibles en la condición.

    Args:
        df:    DataFrame con MultiIndex (Manager, Métrica) o (Métrica,).
        where: Condición estilo pandas query.
               Ejemplos: "Manager == 'First Trust'"
                         "Manager != 'BlackRock'"
                         "Manager in ['First Trust', 'LSV']"
               None → devuelve df sin modificar.

    Returns:
        DataFrame filtrado (con índice aplanado si where fue aplicado).

    Raises:
        ValueError con las columnas disponibles si la condición es inválida.
    """
    if where is None:
        return df

    flat = df.reset_index()
    try:
        return flat.query(where).reset_index(drop=True)
    except Exception as exc:
        cols = list(flat.columns)
        raise ValueError(
            f"Condición where inválida: {where!r}\n"
            f"Columnas disponibles para filtrar: {cols}\n"
            f"Error de pandas: {exc}"
        ) from exc


# ══════════════════════════════════════════════════════════════════════════════
# API PÚBLICA — excel_query
# ══════════════════════════════════════════════════════════════════════════════

def excel_query(
    archivo: str,
    metric: str,
    sheet: int | None = None,
    table: int | None = None,
    where: str | None = None,
    mode: str = "best",
    min_score: float = 0.4,
) -> pd.DataFrame:
    """
    Motor de consulta semántica para archivos Excel desordenados.

    Detecta tablas automáticamente, localiza la métrica por nombre aproximado
    o sinónimo, aplica filtros SQL-like y combina resultados de múltiples hojas.

    Args:
        archivo   : Ruta al archivo .xlsx.
        metric    : Métrica a buscar. Acepta nombre exacto, abreviación o sinónimo.
                    Ejemplos: "AUM", "aum", "assets", "return on assets".
        sheet     : Número de hoja (1-based, como en Excel UI).
                    None (default) → escanea todas las hojas del workbook.
        table     : ID de tabla dentro de la hoja (1-based).
                    None (default) → elige la tabla con mejor score semántico.
        where     : Condición de filtro estilo pandas query, aplicada sobre el
                    resultado con el índice aplanado (reset_index).
                    Ejemplos: "Manager == 'First Trust'"
                              "Manager != 'BlackRock'"
                    None (default) → sin filtro.
        mode      : "best"  (default) → devuelve el resultado de la hoja con
                                        mayor score de similitud.
                    "merge"           → devuelve resultados de TODAS las hojas
                                        donde score >= merge_threshold, concatenados
                                        con una columna "Hoja" indicando la fuente.
        min_score : Umbral mínimo de similitud para aceptar un match [0.0, 1.0].
                    Default: 0.4.
                    En mode="merge" se aplica max(min_score, 0.6) automáticamente
                    para evitar falsos positivos en hojas donde la métrica no existe.

    Returns:
        pd.DataFrame con los datos de la métrica encontrada.

        Metadatos accesibles en df.attrs:
            df.attrs["matched_metric"]  → str   nombre real encontrado
            df.attrs["match_score"]     → float score de similitud
            df.attrs["sheet"]           → str   nombre de hoja fuente
                                                (solo en mode="best")
            df.attrs["sheets_found"]    → list  hojas incluidas
                                                (solo en mode="merge")

    Raises:
        FileNotFoundError si el archivo no existe.
        ValueError si sheet o table están fuera de rango.
        ValueError si no se encuentra ninguna métrica con score >= umbral.
        ValueError si la condición where es inválida (muestra columnas disponibles).

    Ejemplos:
        # Buscar AUM en todas las hojas (escoge la mejor)
        df = excel_query("presupuesto.xlsx", metric="AUM")

        # Usando sinónimo
        df = excel_query("presupuesto.xlsx", metric="assets")

        # Solo hoja 2, tabla 1
        df = excel_query("presupuesto.xlsx", metric="AUM", sheet=2, table=1)

        # Con filtro por Manager
        df = excel_query("presupuesto.xlsx", metric="AUM",
                         where="Manager == 'First Trust'")

        # Combinar resultados de todas las hojas
        df = excel_query("presupuesto.xlsx", metric="AUM", mode="merge")
        print(df["Hoja"].unique())
    """
    archivo = str(Path(archivo))

    if mode not in ("best", "merge"):
        raise ValueError(
            f"mode debe ser 'best' o 'merge', recibido: {mode!r}."
        )

    hojas = _resolve_sheets(archivo, sheet)

    # En mode="merge" subimos el umbral para evitar falsos positivos
    # en hojas donde la métrica realmente no existe.
    merge_threshold = max(min_score, 0.6)

    # ── Recopilar resultado de cada hoja ─────────────────────────────────
    # Almacenamos (hoja, df_raw, metric_found, score) para poder elegir
    # el mejor en mode="best" o concatenar todos en mode="merge".
    hits: list[tuple[str, pd.DataFrame, str, float]] = []

    for hoja in hojas:
        tablas = detectar_todas_las_tablas(archivo, hoja)
        if not tablas:
            continue

        tablas = _select_table(tablas, table, hoja)

        threshold = merge_threshold if mode == "merge" else min_score
        found_tabla, metric_found, score = _find_best_metric(tablas, metric)

        if found_tabla is None or score < threshold:
            continue

        df_raw = _filter_metric(found_tabla, metric_found)
        hits.append((hoja, df_raw, metric_found, score))

    # ── Nada encontrado — error descriptivo ─────────────────────────────
    if not hits:
        # Diagnóstico: mostrar el mejor score por hoja aunque no alcanzó el umbral
        diag_lines: list[str] = []
        for hoja in hojas:
            tablas = detectar_todas_las_tablas(archivo, hoja)
            if tablas:
                tablas = _select_table(tablas, table, hoja)
                _, mf, sc = _find_best_metric(tablas, metric)
                diag_lines.append(
                    f"  '{hoja}': mejor match={mf!r} score={sc:.2f}"
                )
        detail = "\n".join(diag_lines) if diag_lines else "  (sin tablas detectadas)"
        umbral_usado = merge_threshold if mode == "merge" else min_score
        raise ValueError(
            f"No se encontró métrica similar a '{metric}' "
            f"(mode={mode!r}, umbral={umbral_usado:.2f}).\n"
            f"Scores por hoja:\n{detail}\n"
            f"Sugerencias: prueba con min_score más bajo, "
            f"o usa describir_tablas() para ver las métricas disponibles."
        )

    # ── Ensamblar resultado ───────────────────────────────────────────────
    if mode == "best":
        hoja, df_raw, metric_found, score = max(hits, key=lambda x: x[3])
        result = _apply_where(df_raw, where)
        result.attrs["matched_metric"] = metric_found
        result.attrs["match_score"]    = round(score, 4)
        result.attrs["sheet"]          = hoja
        return result

    # mode == "merge": concatenar, añadir columna Hoja, luego aplicar where
    parts: list[pd.DataFrame] = []
    for hoja, df_raw, metric_found, score in hits:
        part = df_raw.reset_index()
        part.insert(0, "Hoja", hoja)
        parts.append(part)

    merged = pd.concat(parts, ignore_index=True)

    if where is not None:
        try:
            merged = merged.query(where).reset_index(drop=True)
        except Exception as exc:
            raise ValueError(
                f"Condición where inválida: {where!r}\n"
                f"Columnas disponibles: {list(merged.columns)}\n"
                f"Error de pandas: {exc}"
            ) from exc

    merged.attrs["sheets_found"] = [h for h, *_ in hits]
    merged.attrs["mode"]         = "merge"
    return merged


# ══════════════════════════════════════════════════════════════════════════════
# API PÚBLICA — utilidades de exploración
# ══════════════════════════════════════════════════════════════════════════════

def buscar_tabla(
    tablas: list[dict[str, Any]],
    metrica: str,
) -> dict[str, Any]:
    """
    Devuelve la primera tabla que contiene una métrica con el texto dado.

    Búsqueda por substring, insensible a mayúsculas. Útil cuando ya tienes
    la lista de tablas y quieres filtrar sin reabrir el archivo.

    Args:
        tablas  : Lista de dicts de detectar_todas_las_tablas().
        metrica : Substring del nombre de la métrica a buscar.

    Returns:
        Primera tabla que contiene la métrica.

    Raises:
        ValueError si no hay ninguna tabla con esa métrica.

    Ejemplo:
        tablas = detectar_todas_las_tablas("archivo.xlsx", "Hoja1")
        tabla  = buscar_tabla(tablas, "AUM")
        df     = extraer_fila(tabla, "AUM")
    """
    metrica_low = _normalize_text(metrica)

    for t in tablas:
        df = t.get("data")
        if df is None:
            continue
        names = df.index.names or []
        nivel = "Métrica" if "Métrica" in names else None
        if nivel is None:
            continue
        if (
            df.index.get_level_values(nivel)
            .str.lower()
            .str.contains(metrica_low, regex=False)
            .any()
        ):
            return t

    raise ValueError(
        f"No se encontró tabla con métrica '{metrica}'. "
        f"Usa describir_tablas(tablas) para ver las métricas disponibles."
    )


def tabla_mas_grande(tablas: list[dict[str, Any]]) -> dict[str, Any]:
    """
    Devuelve la tabla con más filas de datos.

    Args:
        tablas: Lista de dicts de detectar_todas_las_tablas().

    Returns:
        La tabla con mayor shape[0] en su DataFrame.

    Raises:
        ValueError si la lista está vacía.
    """
    if not tablas:
        raise ValueError("La lista de tablas está vacía.")
    return max(
        tablas,
        key=lambda t: t["data"].shape[0] if t.get("data") is not None else 0,
    )


def describir_tablas(tablas: list[dict[str, Any]]) -> None:
    """
    Imprime un resumen legible de todas las tablas detectadas.

    Muestra por cada tabla: id, tipo, score, dimensiones, rango de filas
    y preview de las métricas disponibles.

    Args:
        tablas: Lista de dicts de detectar_todas_las_tablas().

    Ejemplo:
        tablas = detectar_todas_las_tablas("archivo.xlsx", "Hoja1")
        describir_tablas(tablas)

    Salida:
        [1] manager_metrica   score=  1620  shape=(39,36)  filas 2→47
             Métricas: FN/RENTA, AUM, AUM Marzo 2024, ROA, ... (+9 más)
        [2] clave_valor       score=     0  shape=(10,1)   filas 51→63
    """
    if not tablas:
        print("No se detectaron tablas.")
        return

    for t in tablas:
        df     = t.get("data")
        filas  = df.shape[0] if df is not None else 0
        cols   = df.shape[1] if df is not None else 0
        titulo = f"  título={t['titulo']!r}" if t.get("titulo") else ""

        print(
            f"[{t['id']}] {t['tipo']:20s}  "
            f"score={t['score']:>6.0f}  "
            f"shape=({filas},{cols})  "
            f"filas {t['fila_inicio']}→{t['fila_fin']}"
            + titulo
        )

        if df is not None and "Métrica" in (df.index.names or []):
            metricas = df.index.get_level_values("Métrica").unique().tolist()
            preview  = ", ".join(metricas[:8])
            suffix   = f", … (+{len(metricas) - 8} más)" if len(metricas) > 8 else ""
            print(f"     Métricas: {preview}{suffix}")