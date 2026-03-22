"""
test_basico.py
==============
Pruebas mínimas para excel_analyzer.
Crea Excels sintéticos en memoria y verifica que la detección sea correcta.
Correr después de CADA cambio al código.

Uso:
    python test_basico.py
"""

import io
import sys
import datetime
import openpyxl
import pandas as pd

# Importar desde el directorio actual
sys.path.insert(0, ".")
from _core import detectar_todas_las_tablas, detectar_tabla, SheetScanner, WorkbookLoader, TableAnalyzer


# ── Helpers ────────────────────────────────────────────────────────────────

def wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def tabla_desde_bytes(data: bytes, hoja: str = "Hoja1") -> dict:
    import tempfile, os
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(data)
        tmp = f.name
    try:
        return detectar_tabla(tmp, hoja)
    finally:
        os.unlink(tmp)


def todas_desde_bytes(data: bytes, hoja: str = "Hoja1") -> list:
    import tempfile, os
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(data)
        tmp = f.name
    try:
        return detectar_todas_las_tablas(tmp, hoja)
    finally:
        os.unlink(tmp)


def check(nombre: str, condicion: bool, detalle: str = "") -> bool:
    if condicion:
        print(f"  ✅ {nombre}")
    else:
        print(f"  ❌ {nombre}" + (f"  →  {detalle}" if detalle else ""))
    return condicion


# ── CASO 1: Tabla simple con fechas como datetime ──────────────────────────

def caso_1_tabla_simple():
    """
    Tabla: Manager | Métrica | ene-24 | feb-24 | mar-24
    2 managers × 3 métricas × 3 fechas = 18 valores
    """
    print("\nCASO 1 — Tabla simple con fechas datetime")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    fechas = [datetime.date(2024, 1, 1), datetime.date(2024, 2, 1), datetime.date(2024, 3, 1)]

    # Header
    ws.cell(1, 1, "Manager")
    ws.cell(1, 2, "Métrica")
    for j, f in enumerate(fechas, 3):
        ws.cell(1, j, f)

    # Datos
    managers = ["Alpha", "Beta"]
    metricas = ["AUM", "ROA", "FN"]
    row = 2
    for m in managers:
        for met in metricas:
            ws.cell(row, 1, m)
            ws.cell(row, 2, met)
            for j, _ in enumerate(fechas, 3):
                ws.cell(row, j, float(row * j))
            row += 1

    tabla = tabla_desde_bytes(wb_to_bytes(wb))
    df = tabla["data"]

    ok = True
    ok &= check("DataFrame no es None", df is not None)
    if df is None:
        return ok
    ok &= check("Tiene columnas de fecha", len(df.columns) >= 3,
                f"columnas={list(df.columns)}")
    ok &= check("Índice tiene nivel Métrica",
                "Métrica" in (df.index.names if hasattr(df.index, 'names') else [df.index.name]),
                f"index.names={df.index.names if hasattr(df.index, 'names') else df.index.name}")
    ok &= check("Shape correcto (6 filas × 3 cols)", df.shape == (6, 3),
                f"shape={df.shape}")
    return ok


# ── CASO 2: Tabla con fechas como strings (ene-24, feb-24) ────────────────

def caso_2_fechas_como_strings():
    """
    Tabla donde las fechas son strings tipo 'ene-24', 'feb-24'
    """
    print("\nCASO 2 — Fechas como strings (ene-24)")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    fechas_str = ["ene-24", "feb-24", "mar-24", "abr-24"]

    ws.cell(1, 1, "Métrica")
    for j, f in enumerate(fechas_str, 2):
        ws.cell(1, j, f)

    metricas = ["AUM", "ROA", "Ingresos", "Costos"]
    for i, met in enumerate(metricas, 2):
        ws.cell(i, 1, met)
        for j in range(2, 6):
            ws.cell(i, j, float(i * j * 100))

    tabla = tabla_desde_bytes(wb_to_bytes(wb))
    df = tabla["data"]

    ok = True
    ok &= check("DataFrame no es None", df is not None)
    if df is None:
        return ok
    ok &= check("Columnas son fechas normalizadas",
                all(str(c).count("-") == 1 for c in df.columns),
                f"columnas={list(df.columns)}")
    ok &= check("4 filas de métricas", df.shape[0] == 4, f"shape={df.shape}")
    ok &= check("4 columnas de fecha", df.shape[1] == 4, f"shape={df.shape}")
    return ok


# ── CASO 3: Tabla con merged cells en header ──────────────────────────────

def caso_3_merged_cells():
    """
    Tabla con merge: [Manager] [Métrica] [── 2024 ──────────]
                                          [ene][feb][mar][abr]
    """
    print("\nCASO 3 — Merged cells en header")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # Fila 1: merge de año
    ws.cell(1, 1, "Manager")
    ws.cell(1, 2, "Métrica")
    ws.cell(1, 3, 2024)
    ws.merge_cells("C1:F1")

    # Fila 2: meses
    ws.cell(2, 1, "Manager")
    ws.cell(2, 2, "Métrica")
    for j, mes in enumerate(["ene-24", "feb-24", "mar-24", "abr-24"], 3):
        ws.cell(2, j, mes)

    # Datos
    row = 3
    for mgr in ["Alpha", "Beta"]:
        for met in ["AUM", "ROA"]:
            ws.cell(row, 1, mgr)
            ws.cell(row, 2, met)
            for j in range(3, 7):
                ws.cell(row, j, float(row * j))
            row += 1

    tabla = tabla_desde_bytes(wb_to_bytes(wb))
    df = tabla["data"]

    ok = True
    ok &= check("DataFrame no es None", df is not None)
    if df is None:
        return ok
    ok &= check("Tiene datos numéricos",
                df.select_dtypes(include='number').shape[1] > 0 or
                any(isinstance(v, (int, float)) for col in df.columns for v in df[col].dropna()),
                f"dtypes={df.dtypes.tolist()}")
    ok &= check("Shape tiene filas de datos", df.shape[0] >= 4, f"shape={df.shape}")
    return ok


# ── CASO 4: Dos tablas en la misma hoja (apiladas) ────────────────────────

def caso_4_dos_tablas_apiladas():
    """
    Dos tablas separadas por filas vacías en la misma hoja.
    """
    print("\nCASO 4 — Dos tablas apiladas en la misma hoja")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    fechas = ["ene-24", "feb-24", "mar-24"]

    # Tabla 1: filas 1-4
    ws.cell(1, 1, "Métrica")
    for j, f in enumerate(fechas, 2):
        ws.cell(1, j, f)
    for i, met in enumerate(["AUM", "ROA"], 2):
        ws.cell(i, 1, met)
        for j in range(2, 5):
            ws.cell(i, j, float(i * j * 10))

    # Filas 5-6 vacías

    # Tabla 2: filas 7-10
    ws.cell(7, 1, "Métrica")
    for j, f in enumerate(fechas, 2):
        ws.cell(7, j, f)
    for i, met in enumerate(["Ingresos", "Costos"], 8):
        ws.cell(i, 1, met)
        for j in range(2, 5):
            ws.cell(i, j, float(i * j * 20))

    tablas = todas_desde_bytes(wb_to_bytes(wb))

    ok = True
    ok &= check("Detecta al menos 2 tablas", len(tablas) >= 2,
                f"detectó {len(tablas)} tabla(s)")
    for i, t in enumerate(tablas):
        df = t["data"]
        ok &= check(f"Tabla {i+1} tiene DataFrame", df is not None)
    return ok


# ── CASO 5: Filas/columnas vacías intercaladas ────────────────────────────

def caso_5_hoja_con_ruido():
    """
    Tabla rodeada de filas y columnas vacías, con título encima.
    """
    print("\nCASO 5 — Tabla con ruido (filas/columnas vacías + título)")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # Título en fila 2, tabla empieza en fila 4, columna B (índice 2)
    ws.cell(2, 2, "Reporte mensual")

    # Header en fila 4
    ws.cell(4, 2, "Métrica")
    for j, f in enumerate(["ene-24", "feb-24", "mar-24"], 3):
        ws.cell(4, j, f)

    # Datos en filas 5-8
    for i, met in enumerate(["AUM", "ROA", "Ingresos", "FN"], 5):
        ws.cell(i, 2, met)
        for j in range(3, 6):
            ws.cell(i, j, float(i * j * 50))

    tabla = tabla_desde_bytes(wb_to_bytes(wb))
    df = tabla["data"]

    ok = True
    ok &= check("DataFrame no es None", df is not None)
    if df is None:
        return ok
    ok &= check("4 métricas detectadas", df.shape[0] == 4, f"shape={df.shape}")
    ok &= check("3 fechas como columnas", df.shape[1] == 3, f"shape={df.shape}")
    return ok


# ── RUNNER ─────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  test_basico.py — excel_analyzer")
    print("=" * 55)

    casos = [
        caso_1_tabla_simple,
        caso_2_fechas_como_strings,
        caso_3_merged_cells,
        caso_4_dos_tablas_apiladas,
        caso_5_hoja_con_ruido,
    ]

    resultados = []
    for caso in casos:
        try:
            ok = caso()
        except Exception as e:
            print(f"  💥 EXCEPCIÓN: {e}")
            ok = False
        resultados.append(ok)

    total = len(resultados)
    pasados = sum(resultados)
    print(f"\n{'=' * 55}")
    print(f"  Resultado: {pasados}/{total} casos pasaron")
    if pasados == total:
        print("  ✅ Todo OK")
    else:
        print("  ❌ Hay fallos — revisar antes de continuar")
    print("=" * 55)

    return 0 if pasados == total else 1


if __name__ == "__main__":
    sys.exit(main())